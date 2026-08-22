#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
成都看雪山 · Flask 单文件版 v2.0（经验知识增强版）

v2.0 升级要点（数值预报 × 观山经验知识融合）：
  1. 洗尘效应强化：前12h大雨+大风的分段打分，"小雨无效、要下透"
     （依据：华西都市报/四川发布/川观新闻 对"在成都遥望雪山"社群 10 年记录的报道）
  2. 新增"霾层积累惩罚"：距上次有效降雨越久，即使预报晴也扣分
     （依据：2021-04-28 龙泉山失败案例——云图连晴三天、久未下雨、全雾）
  3. 观山季月度加权：6/7 月最强（月均 9-10 天）、5/8 月次之、1/12 月最弱
     （依据：四川省气象局/中国气象局统计）
  4. 日出黄金窗口：以太阳高度角为锚的连续函数（5:40-7:00 为经验最佳）
  5. 连续晴日因子：贡嘎等 200km+ 超远山需连续多日晴朗（全年可见率仅约 5%）
  6. 云底过低惩罚：高湿+低云量大时认为云底 <500m，直接罩住峰线
  7. 内置历史成功案例库与经验法则，页面直接展示

安装：
    python -m pip install flask requests
运行：
    python chengdu_snow_mountain.py
浏览器：
    http://127.0.0.1:5000

气溶胶来自 Open-Meteo Air Quality API 的 aerosol_optical_depth（550 nm）；
历史降水（past_days=7）来自 Open-Meteo Forecast API，用于计算洗尘与霾层积累。
无需 API Key、NetCDF、xarray、cdsapi 或 ecCodes。
"""

from __future__ import annotations

import io
import json
import math
import os
import re
import threading
import time
from datetime import datetime, timedelta, timezone

import numpy as np
import requests
from flask import Flask, jsonify, request, render_template_string, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

APP = Flask(__name__)
TZ = timezone(timedelta(hours=8))
DEFAULT_OBSERVER = {"name": "成都", "lat": 30.65957, "lon": 104.05511, "elev": 500}
MOUNTAINS = [
    {"id": "xiling", "name": "西岭雪山·大雪塘", "lat": 30.6167, "lon": 103.1700, "elev": 5364},
    {"id": "siguniang", "name": "四姑娘山·幺妹峰", "lat": 31.109766, "lon": 102.905291, "elev": 6250},
    {"id": "gongga", "name": "贡嘎山主峰", "lat": 29.595833, "lon": 101.879167, "elev": 7508},
    {"id": "jiuding", "name": "九顶山·狮子王峰", "lat": 31.6470, "lon": 103.8270, "elev": 4984},
]
VIEWPOINTS = [
    {"name":"凤凰山公园至真观","district":"金牛区","lat":30.737724,"lon":104.081896,"precision":"园区坐标","target":"西北—西侧群峰"},
    {"name":"成都大熊猫繁育研究基地","district":"成华区","lat":30.742306,"lon":104.135901,"precision":"园区坐标","target":"西北—西侧群峰"},
    {"name":"锦城湖3号湖区广场","district":"高新区","lat":30.573865,"lon":104.039910,"precision":"湖区近似位置","target":"西岭、幺妹峰"},
    {"name":"钟家山美满村三号观景平台","district":"龙泉驿区","lat":30.5620,"lon":104.3300,"precision":"景区近似位置","target":"成都平原与西部群峰"},
    {"name":"丹景山狮子堡观景台","district":"四川天府新区","lat":30.3900,"lon":104.2100,"precision":"景区近似位置","target":"西岭雪山、城市天际线"},
    {"name":"鲁家滩湿地公园","district":"温江区","lat":30.689937,"lon":103.774798,"precision":"园区坐标","target":"西岭与龙门山群峰"},
    {"name":"世界科幻公园·菁蓉湖","district":"郫都区","lat":30.8060,"lon":103.8780,"precision":"园区近似位置","target":"龙门山群峰"},
    {"name":"天府农博岛","district":"新津区","lat":30.504319,"lon":103.801021,"precision":"园区坐标","target":"西岭、贡嘎方向"},
    {"name":"都江堰南桥","district":"都江堰市","lat":30.997597,"lon":103.616407,"precision":"街区坐标","target":"龙门山近景"},
    {"name":"长秋月观景台","district":"蒲江县","lat":30.2500,"lon":103.6000,"precision":"长秋山近似位置","target":"贡嘎、幺妹峰及川西群峰"},
]
WEATHER_MODELS = {
    "best_match": {"name": "智能最佳匹配", "detail": "Open-Meteo 自动选择当地最佳可用模型"},
    "ecmwf_ifs025": {"name": "ECMWF IFS 0.25°", "detail": "欧洲中期天气预报中心全球模式"},
    "gfs_seamless": {"name": "NOAA GFS", "detail": "美国 NOAA 全球预报系统"},
    "icon_seamless": {"name": "DWD ICON", "detail": "德国气象局全球模式"},
    "cma_grapes_global": {"name": "CMA GRAPES", "detail": "中国气象局约15 km全球模式"},
    "jma_seamless": {"name": "JMA", "detail": "日本气象厅全球模式"},
}

# ---- v2.0 经验知识库 -----------------------------------------------------
# 月度平均可见雪山天数（天/月）。依据：四川省气象局《绿镜头系列》、中国气象局
# 《这座超大城市开窗就能看雪山》(2024-07)：6/7 月 9-10 天，5/8 月 7-8 天，1/12 月约 1 天。
MONTH_FREQ = {1: 1.0, 2: 1.5, 3: 2.5, 4: 5.0, 5: 7.5, 6: 9.5, 7: 9.5, 8: 7.5, 9: 5.0, 10: 2.5, 11: 1.5, 12: 1.0}

# 经验法则（页面展示用，均来自公开报道的观山社群/官方统计）
EMPIRICAL_RULES = [
    ("第一定律", "前一天大风大雨（下透）→ 次日清晨雨过天晴，出山概率最高", "华西都市报 · 在成都遥望雪山社群 10 年记录"),
    ("小雨无效", "小雨对霾层起不了决定性作用，必须下透 + 大风", "米拍《2021 成都雪山拍摄日记》"),
    ("霾层积累", "连续 5 天以上无有效降雨，预报晴也容易全雾", "2021-04-28 龙泉山失败案例"),
    ("观山旺季", "6/7 月最强（月均 9-10 天），5/8 月次之，1/12 月最弱", "四川省气象局统计"),
    ("黄金窗口", "清晨 5:40-7:00 空气最通透；日照金山只能清晨看", "电子科大攻略 · 澎湃新闻"),
    ("超远山门槛", "贡嘎（240km）需连续多日晴朗少云，全年可见仅约 5%", "川观新闻 C视觉 · 成都气象"),
]

# 历史成功案例库（公开报道的真实案例：日期 + 天气条件 + 现象）
SUCCESS_CASES = [
    {"date": "2013-07-29", "cond": "前夜大雨+劲风扫净尘霾", "sight": "傍晚幺妹峰日照金山", "src": "四川发布"},
    {"date": "2017-06-05", "cond": "前两日大雨、头天傍晚放晴", "sight": "史诗级 240km 贡嘎长卷，成「开山日」", "src": "华西都市报"},
    {"date": "2020-05-25", "cond": "昨日疾风劲雨带来清凉", "sight": "清晨雪山再现，全网刷屏", "src": "红星新闻"},
    {"date": "2024-05-05", "cond": "前夜狂风暴雨洗去浮尘（立夏·开山节）", "sight": "清晨澄澈如洗，雪山冲破云层", "src": "川观新闻"},
    {"date": "2025-01-28", "cond": "除夕，前一日降雪", "sight": "9 年记录首次除夕见山，全线群山可见", "src": "四川观察"},
    {"date": "2025-03-25", "cond": "早晴（春季开山节）", "sight": "大雪塘/幺妹峰/太子城一线+日照金山", "src": "2025 年度观山报告"},
    {"date": "2025-05-07", "cond": "前日傍晚世纪火烧云后", "sight": "幺妹峰/大雪塘/贡嘎三山齐现", "src": "@成都市气象服务"},
    {"date": "2026-06-03", "cond": "昨夜阵雨雷雨，今晨雨过天晴", "sight": "雪山+朝霞同框", "src": "成都发布"},
    {"date": "2026-07-12/13", "cond": "周末雨水洗礼后通透澄净", "sight": "幺妹峰早晚两次露脸", "src": "川观新闻 C视觉(82)"},
    {"date": "2026-07-19/20", "cond": "连续两天晴朗通透", "sight": "贡嘎+幺妹峰+大雪塘同框", "src": "川观新闻 C视觉(84)"},
    {"date": "2026-08-12", "cond": "昨夜暴雨洗礼", "sight": "清晨通透绝佳，全城十余点位拍到幺妹峰", "src": "川观新闻 C视觉(86)"},
]

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "ChengduSnowMountain/1.0"})
_cache = {}
_cache_lock = threading.Lock()


def clamp(x, lo=0.0, hi=1.0):
    return max(lo, min(hi, x))


# v2.10.21: 观测坐标放开到中国及周边（不再限制在川渝）。
# 依据风云四号红外云图覆盖范围(70-135°E / 10-55°N)取整，超出部分卫星/云图功能不可用但预报仍可出。
GEO_LAT_MIN, GEO_LAT_MAX, GEO_LON_MIN, GEO_LON_MAX = 10.0, 55.0, 70.0, 135.0

def _clamp_geo(lat, lon):
    """将观测坐标钳制到中国及周边范围，并返回 (lat, lon)。"""
    return clamp(lat, GEO_LAT_MIN, GEO_LAT_MAX), clamp(lon, GEO_LON_MIN, GEO_LON_MAX)


def haversine_bearing(a_lat, a_lon, b_lat, b_lon):
    p1, p2 = map(math.radians, (a_lat, b_lat))
    dl = math.radians(b_lon - a_lon)
    y = math.sin(dl) * math.cos(p2)
    x = math.cos(p1) * math.sin(p2) - math.sin(p1) * math.cos(p2) * math.cos(dl)
    bearing = (math.degrees(math.atan2(y, x)) + 360) % 360
    h = math.sin((p2-p1)/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 6371.0088 * 2 * math.asin(math.sqrt(h)), bearing


def interpolate_great_circle(lat1, lon1, lat2, lon2, n=10):
    # 距离小于 300 km 时线性插值误差对 0.4°/0.25°气象格点可忽略。
    return [(lat1+(lat2-lat1)*i/(n-1), lon1+(lon2-lon1)*i/(n-1)) for i in range(n)]


def solar_position(dt_local, lat, lon):
    """NOAA 近似太阳高度角/方位角；dt_local 必须带时区。"""
    u = dt_local.astimezone(timezone.utc)
    jd = u.timestamp()/86400.0 + 2440587.5
    t = (jd-2451545.0)/36525.0
    L0 = (280.46646 + t*(36000.76983 + 0.0003032*t)) % 360
    M = math.radians((357.52911 + t*(35999.05029-0.0001537*t)) % 360)
    C = (1.914602-t*(0.004817+0.000014*t))*math.sin(M) + (0.019993-0.000101*t)*math.sin(2*M) + 0.000289*math.sin(3*M)
    lam = math.radians(L0+C-0.00569-0.00478*math.sin(math.radians(125.04-1934.136*t)))
    eps = math.radians(23.439291-0.0130042*t)
    dec = math.asin(math.sin(eps)*math.sin(lam))
    y = math.tan(eps/2)**2
    L = math.radians(L0)
    eq = 4*math.degrees(y*math.sin(2*L)-2*0.016708634*math.sin(M)+4*0.016708634*y*math.sin(M)*math.cos(2*L)-0.5*y*y*math.sin(4*L)-1.25*0.016708634**2*math.sin(2*M))
    minutes = u.hour*60+u.minute+u.second/60
    tst = (minutes + eq + 4*lon) % 1440
    ha = math.radians(tst/4-180)
    phi = math.radians(lat)
    zen = math.acos(clamp(math.sin(phi)*math.sin(dec)+math.cos(phi)*math.cos(dec)*math.cos(ha), -1, 1))
    elev = 90-math.degrees(zen)
    az = (math.degrees(math.atan2(math.sin(ha), math.cos(ha)*math.sin(phi)-math.tan(dec)*math.cos(phi)))+180) % 360
    return elev, az


def apparent_peak_angle(distance_km, peak_elev, observer_elev):
    # 标准折射取地球等效半径 7/6 R，得到峰顶几何仰角。
    re = 6371008.8 * 7/6
    d = distance_km*1000
    drop = d*d/(2*re)
    return math.degrees(math.atan2(peak_elev-observer_elev-drop, d))


def cache_get(key, max_age):
    with _cache_lock:
        item = _cache.get(key)
        return item[1] if item and time.time()-item[0] < max_age else None


def cache_put(key, value):
    with _cache_lock:
        _cache[key] = (time.time(), value)


def cache_get_stale(key, max_age):
    """返回 (value, is_stale)。缓存存在即返回，is_stale 表示已超过有效期。
    供数据源 429/故障时回退到过期数据，避免整个预报失败。"""
    with _cache_lock:
        item = _cache.get(key)
        if not item:
            return None, False
        return item[1], time.time() - item[0] >= max_age


def http_get_json(url, params, timeout=30, tries=3, wait=1.0):
    """带指数退避的 GET：429/5xx/网络错误自动重试，全部失败后抛出最后一次异常。
    v2.4: 退避缩短（1s、2s）——快速失败优先，避免单请求拖太久触发网关超时。"""
    last = None
    for attempt in range(tries):
        try:
            r = SESSION.get(url, params=params, timeout=timeout)
            if r.status_code == 429 or r.status_code >= 500:
                raise requests.HTTPError(f"{r.status_code} {r.reason}", response=r)
            r.raise_for_status()
            return r.json()
        except (requests.HTTPError, requests.ConnectionError, requests.Timeout) as e:
            last = e
            if attempt < tries - 1:
                time.sleep(wait * (attempt + 1))
    raise last


def open_meteo_elevations(points):
    """Open-Meteo GLO-90 海拔；官方单次最多 100 个坐标。"""
    values = []
    for start in range(0, len(points), 100):
        batch = points[start:start+100]
        params = {
            "latitude": ",".join(f"{p[0]:.6f}" for p in batch),
            "longitude": ",".join(f"{p[1]:.6f}" for p in batch),
        }
        r = http_get_json("https://api.open-meteo.com/v1/elevation", params=params, timeout=30)
        part = r.get("elevation", [])
        if len(part) != len(batch):
            raise RuntimeError(f"海拔 API 返回 {len(part)} 个值，预期 {len(batch)} 个")
        values.extend(None if x is None else float(x) for x in part)
    return values


def observer_elevation(lat, lon):
    key = f"observer-elev:{lat:.5f},{lon:.5f}"
    cached = cache_get(key, 86400 * 30)
    if cached is not None: return cached
    try:
        value = open_meteo_elevations([(lat, lon)])[0]
    except Exception:
        value = None
    if value is None:
        # v2.2: 极端/无数据坐标下兜底为成都平原近似海拔，避免整个预报失败
        value = float(DEFAULT_OBSERVER["elev"])
    cache_put(key, value)
    return value


def path_terrain(observer, mountain, n=10):
    points=interpolate_great_circle(observer["lat"],observer["lon"],mountain["lat"],mountain["lon"],n)
    key="path-terrain:"+":".join(f"{a:.4f},{b:.4f}" for a,b in points)
    cached, stale = cache_get_stale(key, 86400*30)
    if cached is not None and not stale: return cached
    try:
        values=open_meteo_elevations(points)
    except Exception:
        # v2.3: 海拔接口繁忙时用过期地形兜底（地形基本不变，安全）
        if cached is not None: return cached
        raise
    # 端点使用已知观测点/峰顶高程，避免90m DEM格点把尖锐峰顶平滑掉。
    values[0]=float(observer["elev"]); values[-1]=float(mountain["elev"])
    cache_put(key,values)
    return values


# v2.10.21: 峰顶地形遮挡判断——沿观测点到峰顶的高密度视线剖面，
# 若中间任一点地形高于该处视线高度（含地球曲率+标准大气折射修正），则峰顶被山体遮挡、不可见。
TERRAIN_OCC_N = 96              # 剖面采样点数（约每 1.2km 一个点，能捕捉山脊；过密会漏掉尖锐峰点）
TERRAIN_OCC_MARGIN = 80.0       # 容差 m（树木/建筑/DEM 误差），低于视线不足 80m 视为擦边可见
TERRAIN_OCC_RE = 6371008.8 * 7.0 / 6.0   # 等效地球半径 m（7/6R 折射，与前端剖面图一致）

def _los_height_m(e0_m, e1_m, dist_m, d_m):
    """观测点到峰顶的视线在距离 d 处的高度（含折射 bulge = d(D-d)/(2Re)）。"""
    return e0_m + (e1_m - e0_m) * d_m / dist_m + d_m * (dist_m - d_m) / (2.0 * TERRAIN_OCC_RE)

def terrain_occlusion(observer, mountain, n=TERRAIN_OCC_N):
    """判断从观测点到峰顶的视线是否被中间地形遮挡。
    返回 (blocked: bool, info: dict|None)；info 含遮挡最严重处距离/地形高度/视线高度/经纬度。"""
    e0, e1 = float(observer["elev"]), float(mountain["elev"])
    dist = haversine_bearing(observer["lat"], observer["lon"], mountain["lat"], mountain["lon"])[0]
    if dist <= 1:
        return False, None
    points = interpolate_great_circle(observer["lat"], observer["lon"], mountain["lat"], mountain["lon"], n)
    key = "path-terrain:" + ":".join(f"{a:.4f},{b:.4f}" for a, b in points)
    cached, stale = cache_get_stale(key, 86400 * 30)
    if cached is not None and not stale:
        values = cached
    else:
        try:
            values = open_meteo_elevations(points)
        except Exception:
            if cached is not None:
                values = cached
            else:
                return False, None   # 地形数据不可用时保守放行，不误伤
        values[0] = e0; values[-1] = e1
        cache_put(key, values)
    dist_m = dist * 1000.0
    worst = None
    for i in range(1, n - 1):
        d_m = dist_m * i / (n - 1)
        los = _los_height_m(e0, e1, dist_m, d_m)
        t = float(values[i])
        if t > los + TERRAIN_OCC_MARGIN:
            over = t - los
            if worst is None or over > worst["over"]:
                worst = {"over": over, "distance": d_m / 1000.0, "terrain": t, "los": los,
                         "lat": round(points[i][0], 5), "lon": round(points[i][1], 5)}
    if worst is None:
        return False, None
    return True, worst


def open_meteo_corridor(observer, mountains, days=5, model="best_match"):
    points, owners = [], []
    for m in mountains:
        for lat, lon in interpolate_great_circle(observer["lat"], observer["lon"], m["lat"], m["lon"], 10):
            points.append((lat, lon)); owners.append(m["id"])
    key = "om:" + model + ":" + ":".join(f"{x[0]:.3f},{x[1]:.3f}" for x in points)
    cached, stale = cache_get_stale(key, 1800)
    if cached is not None and not stale: return cached, False
    params = {
        "latitude": ",".join(f"{p[0]:.4f}" for p in points),
        "longitude": ",".join(f"{p[1]:.4f}" for p in points),
        "hourly": "cloud_cover,cloud_cover_low,cloud_cover_mid,cloud_cover_high,visibility,relative_humidity_2m,precipitation_probability,precipitation,wind_speed_10m,temperature_2m",
        # v2.0: 回拉 7 天历史降水，用于"前夜大雨洗尘"与"久未降雨霾层积累"两个经验因子
        "forecast_days": days, "past_days": 7, "timezone": "Asia/Shanghai", "models": model
    }
    try:
        raw = http_get_json("https://api.open-meteo.com/v1/forecast", params=params, timeout=30)
    except Exception:
        # v2.3: 限流/故障时回退到过期缓存，保证页面可用
        if cached is not None: return cached, True
        raise
    rows = raw if isinstance(raw, list) else [raw]
    result = {m["id"]: [] for m in mountains}
    for i, row in enumerate(rows): result[owners[i]].append(row["hourly"])
    cache_put(key, result)
    return result, False


def open_meteo_aerosol(observer, mountains, days=5):
    """获取每座山视线上的 10 个 AOD550/PM2.5/沙尘采样点。"""
    points, owners = [], []
    for m in mountains:
        for lat, lon in interpolate_great_circle(observer["lat"], observer["lon"], m["lat"], m["lon"], 10):
            points.append((lat, lon)); owners.append(m["id"])
    key = "air:" + ":".join(f"{x[0]:.3f},{x[1]:.3f}" for x in points)
    cached, stale = cache_get_stale(key, 3600)
    if cached is not None and not stale: return cached, False
    params = {
        "latitude": ",".join(f"{p[0]:.4f}" for p in points),
        "longitude": ",".join(f"{p[1]:.4f}" for p in points),
        "hourly": "aerosol_optical_depth,pm2_5,dust",
        "forecast_days": days,
        "timezone": "Asia/Shanghai",
        "domains": "cams_global",
        "cell_selection": "nearest",
    }
    try:
        raw = http_get_json("https://air-quality-api.open-meteo.com/v1/air-quality", params=params, timeout=35)
    except Exception:
        if cached is not None: return cached, True
        raise
    rows = raw if isinstance(raw, list) else [raw]
    if len(rows) != len(points):
        raise RuntimeError(f"Open-Meteo 气溶胶返回 {len(rows)} 个点，预期 {len(points)} 个点")
    result = {m["id"]: [] for m in mountains}
    for i, row in enumerate(rows):
        result[owners[i]].append(row["hourly"])
    cache_put(key, result)
    return result, False


# ---- v2.6 历史任意时段（Open-Meteo Archive · ERA5 再分析，1940 年至今） ----

def open_meteo_archive(observer, mountains, start_date, end_date):
    """历史任意时段的走廊天气。用 Open-Meteo Archive API（ERA5 再分析，1940-至今）。
    返回与 open_meteo_corridor 相同结构：{山id: [hourly, ...10个走廊点]}。
    注意：ERA5 无能见度变量（API 返回 None），能见度由 score_hour 用湿度/降水/气溶胶/低云物理估算。"""
    points, owners = [], []
    for m in mountains:
        for lat, lon in interpolate_great_circle(observer["lat"], observer["lon"], m["lat"], m["lon"], 10):
            points.append((lat, lon)); owners.append(m["id"])
    params = {
        "latitude": ",".join(f"{p[0]:.4f}" for p in points),
        "longitude": ",".join(f"{p[1]:.4f}" for p in points),
        "hourly": "cloud_cover,cloud_cover_low,cloud_cover_mid,cloud_cover_high,visibility,relative_humidity_2m,precipitation,wind_speed_10m,temperature_2m",
        "start_date": start_date, "end_date": end_date,
        "timezone": "Asia/Shanghai",
    }
    raw = http_get_json("https://archive-api.open-meteo.com/v1/archive", params=params, timeout=45)
    rows = raw if isinstance(raw, list) else [raw]
    if len(rows) != len(points):
        raise RuntimeError(f"历史天气返回 {len(rows)} 个点，预期 {len(points)} 个点")
    result = {m["id"]: [] for m in mountains}
    for i, row in enumerate(rows):
        result[owners[i]].append(row["hourly"])
    return result


def open_meteo_aerosol_history(observer, mountains, start_date, end_date):
    """历史气溶胶。Air Quality API 的 past_days 最多 92 天；超过则返回 None（评分降级）。
    返回 {山id: [hourly, ...]} 或 None。"""
    today = datetime.now(TZ).date()
    s = datetime.strptime(start_date, "%Y-%m-%d").date()
    e = datetime.strptime(end_date, "%Y-%m-%d").date()
    need = (today - s).days + 1          # 从今天往回需要覆盖到 start 的天数
    if need > 92:
        return None
    points, owners = [], []
    for m in mountains:
        for lat, lon in interpolate_great_circle(observer["lat"], observer["lon"], m["lat"], m["lon"], 10):
            points.append((lat, lon)); owners.append(m["id"])
    params = {
        "latitude": ",".join(f"{p[0]:.4f}" for p in points),
        "longitude": ",".join(f"{p[1]:.4f}" for p in points),
        "hourly": "aerosol_optical_depth,pm2_5,dust",
        "past_days": need,
        "timezone": "Asia/Shanghai",
        "domains": "cams_global",
        "cell_selection": "nearest",
    }
    try:
        raw = http_get_json("https://air-quality-api.open-meteo.com/v1/air-quality", params=params, timeout=45)
    except Exception:
        return None
    rows = raw if isinstance(raw, list) else [raw]
    if len(rows) != len(points):
        return None
    result = {m["id"]: [] for m in mountains}
    for i, row in enumerate(rows):
        # 截取查询区间内的数据（past_days 返回的序列比区间长）
        hourly = row["hourly"]
        times = hourly.get("time") or []
        keep = [i for i, t in enumerate(times) if start_date <= t[:10] <= end_date]
        if not keep:
            return None
        cut = {k: (v if not isinstance(v, list) else [v[i] for i in keep]) for k, v in hourly.items()}
        result[owners[i]].append(cut)
    return result


EMPTY_AIR = {"aod": None, "pm2_5": None, "dust": None, "points": []}


# ---- v2.8 天气系统因子（冷空气 / 锋面切变 / 槽脊 / 逆温层 / 温度平流） ----
# 依据：中央气象台天气形势分析规范 + 四川盆地气象博主方法论。
# 数据：Open-Meteo 等压面（850/925/1000hPa）与 10m 地面层对比。

def open_meteo_synoptic(observer, model="best_match", days=5):
    """观测点（成都）等压面形势：850/925/1000hPa 温度、850hPa 风与位势高度。
    past_days=3 提供 24h 变温与高度距平参照；失败返回 None（评分降级，不阻塞）。"""
    key = f"syn:{model}:{observer['lat']:.2f},{observer['lon']:.2f}"
    cached = cache_get(key, 1800)
    if cached is not None: return cached
    params = {
        "latitude": f"{observer['lat']:.4f}", "longitude": f"{observer['lon']:.4f}",
        "hourly": ("temperature_2m,wind_speed_10m,wind_direction_10m,"
                   "temperature_850hPa,temperature_925hPa,temperature_1000hPa,"
                   "wind_speed_850hPa,wind_direction_850hPa,geopotential_height_850hPa"),
        "forecast_days": days, "past_days": 3, "timezone": "Asia/Shanghai", "models": model,
    }
    try:
        raw = http_get_json("https://api.open-meteo.com/v1/forecast", params=params, timeout=30)
        hourly = raw["hourly"]
        cache_put(key, hourly)
        return hourly
    except Exception:
        return None


# ---- v2.9 真实实况接入（免费 API） ----
# Open-Meteo current=（模型当前时次实况，无 key） + 中国天气网/中央气象台观测（真实站点观测，无 key）

CN_CITIES = [  # (城市, cityid, 纬度, 经度)
    ("成都", "101270101", 30.66, 104.06), ("都江堰", "101270108", 30.99, 103.65),
    ("重庆", "101040100", 29.56, 106.55), ("绵阳", "101270401", 31.47, 104.68),
    ("德阳", "101272001", 31.13, 104.40), ("雅安", "101271701", 30.01, 103.04),
    ("眉山", "101271501", 30.05, 103.83), ("乐山", "101271401", 29.55, 103.77),
    ("资阳", "101271901", 30.13, 104.63), ("遂宁", "101270701", 30.51, 105.59),
    ("广元", "101272101", 32.44, 105.84), ("南充", "101270501", 30.80, 106.08),
]

WMO_ZH = {0: "晴", 1: "大致晴朗", 2: "多云", 3: "阴", 45: "雾", 48: "雾凇雾",
          51: "小毛毛雨", 53: "毛毛雨", 55: "浓毛毛雨", 56: "冻毛毛雨", 57: "浓冻毛毛雨",
          61: "小雨", 63: "中雨", 65: "大雨", 66: "冻雨", 67: "强冻雨",
          71: "小雪", 73: "中雪", 75: "大雪", 77: "雪粒",
          80: "小阵雨", 81: "阵雨", 82: "强阵雨", 85: "阵雪", 86: "强阵雪",
          95: "雷暴", 96: "雷暴伴冰雹", 99: "强雷暴伴冰雹"}


def nearest_cn_city(lat, lon):
    best, bd = CN_CITIES[0], 1e9
    for c in CN_CITIES:
        d = (c[2]-lat)**2 + (c[3]-lon)**2
        if d < bd: best, bd = c, d
    return best


def current_weather_om(observer):
    """Open-Meteo 当前实况（模型当前时次，无 key）。失败返回 None。"""
    key = f"cur:om:{observer['lat']:.2f},{observer['lon']:.2f}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        raw = http_get_json("https://api.open-meteo.com/v1/forecast", params={
            "latitude": f"{observer['lat']:.4f}", "longitude": f"{observer['lon']:.4f}",
            "current": ("temperature_2m,relative_humidity_2m,apparent_temperature,is_day,precipitation,"
                        "weather_code,cloud_cover,pressure_msl,wind_speed_10m,wind_direction_10m,visibility"),
            "timezone": "Asia/Shanghai"}, timeout=25)
        cur = raw.get("current") or {}
        if not cur: return None
        out = {"time": cur.get("time"),
               "temperature_2m": cur.get("temperature_2m"),
               "relative_humidity_2m": cur.get("relative_humidity_2m"),
               "weather_code": cur.get("weather_code"),
               "cloud_cover": cur.get("cloud_cover"),
               "visibility": cur.get("visibility"),
               "wind_speed_10m": cur.get("wind_speed_10m"),
               "wind_direction_10m": cur.get("wind_direction_10m"),
               "precipitation": cur.get("precipitation"),
               "pressure_msl": cur.get("pressure_msl"),
               "is_day": cur.get("is_day")}
        cache_put(key, out)
        return out
    except Exception:
        return None


def current_weather_cn(lat, lon):
    """中国天气网/中央气象台 真实站点观测（免费无 key，动态编码）。
    使用 d1.weather.com.cn/sk_2d 逐分钟实况接口（dataSK 对象）。
    按观测点就近匹配城市代码。失败返回 None。"""
    city, cityid = nearest_cn_city(lat, lon)[:2]
    key = f"cur:cn:{cityid}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Referer": "http://www.weather.com.cn/"}
        r = SESSION.get(f"http://d1.weather.com.cn/sk_2d/{cityid}.html", headers=headers, timeout=20)
        # 该接口部分城市/时段返回 UTF-8，部分返回 GBK，按 apparent_encoding 动态检测
        enc = r.apparent_encoding or "utf-8"
        if enc.lower().replace("_", "-") not in ("utf-8", "utf8", "gbk", "gb2312", "gb18030"):
            enc = "utf-8"
        text = r.content.decode(enc, errors="replace")
        # dataSK 对象可能以 "};" 结尾（weather_index 接口）或直接 "}" 结尾（sk_2d 接口）
        m = re.search(r'dataSK\s*=\s*(\{.*?\})\s*;', text, re.S)
        if not m:
            m = re.search(r'dataSK\s*=\s*(\{.*\})', text, re.S)
        if not m: return None
        sk = json.loads(m.group(1))

        def _f(v):  # 提取字符串中的首个数值（支持 "90%"、"13km"、"0级"、"25.3"）
            mm = re.search(r'-?\d+(?:\.\d+)?', str(v or ""))
            return float(mm.group(0)) if mm else None

        vis_km = _f(sk.get("njd"))
        out = {"city": city, "cityid": cityid,
               "time": f'{sk.get("date", "")} {sk.get("time", "")}'.strip(),
               "temperature_2m": _f(sk.get("temp")),
               "humidity": _f(sk.get("SD")),
               "pressure": _f(sk.get("qy")),
               "visibility": vis_km * 1000.0 if vis_km is not None else None,  # km->m，与 Open-Meteo 一致
               "weather": sk.get("weather"),
               "wind": sk.get("WD"),
               "wind_level": sk.get("WS"),
               "aqi": _f(sk.get("aqi")),
               "aqi_pm25": _f(sk.get("aqi_pm25"))}
        cache_put(key, out)
        return out
    except Exception:
        return None


# ---- v2.9.4: 真实观测三源（METAR 机场实况 + 风云四号卫星云图 + 全国雷达拼图） ----
AIRPORTS = [
    ("成都双流", "ZUUU"), ("成都天府", "ZUTF"), ("绵阳南郊", "ZUMY"),
    ("南充高坪", "ZUNC"), ("宜宾五粮液", "ZUYB"), ("泸州云龙", "ZULZ"),
    ("达州金垭", "ZUDA"), ("广元盘龙", "ZUGU"),
]
METAR_CLOUD_ZH = {"FEW": "疏云", "SCT": "散云", "BKN": "多云", "OVC": "阴天",
                  "NSC": "无显著云", "NCD": "无云", "CAVOK": "晴空", "VV": "垂直能见度"}
FLIGHT_ZH = {"VFR": "VFR目视", "MVFR": "MVFR边缘", "IFR": "IFR仪表", "LIFR": "LIFR极低"}


def _vis_to_km(v):
    """METAR 能见度（英里，可为 '6+' 或 '3/4' 分数）→ 公里。"""
    if v is None: return None
    s = str(v).strip()
    plus = s.endswith("+")
    if plus: s = s[:-1]
    if "/" in s:
        try:
            a, b = s.split("/")
            return round(float(a) / float(b) * 1.60934, 1)
        except ValueError:
            return None
    try:
        km = float(s) * 1.60934
        return round(km, 1)
    except ValueError:
        return None


def obs_metar():
    """aviationweather.gov METAR（免费无 key）。返回盆地机场实况：能见度/云底/湿度/风/天气现象。缓存 10 分钟。"""
    key = "obs:metar"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        ids = ",".join(a[1] for a in AIRPORTS)
        r = SESSION.get("https://aviationweather.gov/api/data/metar",
                        params={"ids": ids, "format": "json", "taf": "false", "hours": "24"},
                        headers={"User-Agent": "chengdu-snow-mountain/1.0"}, timeout=25)
        if r.status_code != 200 or "json" not in (r.headers.get("content-type") or ""):
            return None
        arr = r.json()
        latest = {}
        for it in arr:  # hours=5 会返回多时次，按机场保留最新一条
            icao = it.get("icaoId")
            if not icao: continue
            cur = latest.get(icao)
            if cur is None or (it.get("reportTime") or "") >= (cur.get("reportTime") or ""):
                latest[icao] = it
        out = []
        for icao, it in sorted(latest.items()):
            name = dict(AIRPORTS).get(icao, icao)
            temp = it.get("temp"); dewp = it.get("dewp")
            rh = None
            if temp is not None and dewp is not None:
                try:
                    et = 6.112 * math.exp(17.67 * temp / (temp + 243.5))
                    ed = 6.112 * math.exp(17.67 * dewp / (dewp + 243.5))
                    rh = round(min(100.0, ed / et * 100))
                except (ZeroDivisionError, ValueError, TypeError):
                    rh = None
            bases = [c.get("base") for c in (it.get("clouds") or []) if isinstance(c.get("base"), (int, float))]
            min_base = min(bases) if bases else None
            clouds_zh = []
            for c in (it.get("clouds") or []):
                cov = c.get("cover"); base = c.get("base")
                if cov in METAR_CLOUD_ZH and isinstance(base, (int, float)):
                    clouds_zh.append({"text": f"{METAR_CLOUD_ZH[cov]} {int(round(base * 0.3048))}m", "base_m": round(base * 0.3048)})
            raw = it.get("rawOb") or ""
            wx = ""
            m2 = re.search(r"\s([-+]{0,2}(?:VC)?(?:MI|PR|BC|DR|BL|SH|TS|FZ)+[A-Z]{2,4}(?:[A-Z]{2})?)\s", raw)
            if m2: wx = m2.group(1)
            out.append({"icao": icao, "name": name, "time": it.get("reportTime"),
                        "temp": temp, "dewp": dewp, "rh": rh,
                        "visibility_km": _vis_to_km(it.get("visib")),
                        "cloud_base_m": int(round(min_base * 0.3048)) if min_base is not None else None,
                        "clouds": clouds_zh, "wind_dir": it.get("wdir"), "wind_kt": it.get("wspd"),
                        "wx": wx or None, "flight": FLIGHT_ZH.get(it.get("fltCat")) or it.get("fltCat")})
        cache_put(key, out)
        return out
    except Exception:
        return None


def _probe_img(url, timeout=6):
    """HEAD 探测图片是否真实存在（不下载正文）。"""
    try:
        r = SESSION.head(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code != 200:
            return False
        return (r.headers.get("content-type") or "").startswith("image/")
    except Exception:
        return False


def obs_radar_url(now=None):
    """最新全国雷达组合反射率拼图 URL（中央气象台 RDCP，产品文件名用 UTC 时次，6 分钟更新）。缓存 6 分钟。"""
    cached = cache_get("obs:radar", 360)
    if cached is not None: return cached
    now = now or datetime.now(TZ)
    # v2.10.4: 对齐到 6 分钟产品网格（此前从整点探测，拿不到整点之后的时次，图片滞后一个周期）
    utc = now.astimezone(timezone.utc)
    base = utc.replace(minute=utc.minute // 6 * 6, second=0, microsecond=0)
    for back in range(0, 40):  # 最多回退 4 小时
        t = base - timedelta(minutes=6 * back)
        ts = t.strftime("%Y%m%d%H%M") + "00000"
        url = (f"https://image.nmc.cn/product/{t.strftime('%Y/%m/%d')}/RDCP/medium/"
               f"SEVP_AOC_RDCP_SLDAS3_ECREF_ACHN_L88_PI_{ts}.PNG")
        if _probe_img(url):
            out = {"url": url, "time": t.astimezone(TZ).isoformat()}
            cache_put("obs:radar", out)
            return out
    return None


# v2.9.5: 四川单站雷达（中央气象台单站产品，6 分钟更新）。站代码取自 nmc.cn 各雷达页。
SC_RADARS = [
    ("成都", "AZ9280"), ("绵阳", "AZ9816"), ("南充", "AZ9817"),
    ("雅安", "AZ9835"), ("广元", "AZ9839"), ("宜宾", "AZ9837"),
    ("达州", "AZ9834"), ("西昌", "AZ9832"), ("重庆", "AZ9838"),
]


def obs_sc_radar_urls(now=None):
    """四川及周边单站雷达组合反射率 URL（成都/绵阳/南充/雅安等），全部失败返回 None。缓存 6 分钟。"""
    cached = cache_get("obs:radar:sc", 360)
    if cached is not None: return cached
    now = now or datetime.now(TZ)
    # v2.10.4: 对齐到 6 分钟产品网格（与全国拼图一致）
    utc = now.astimezone(timezone.utc)
    base = utc.replace(minute=utc.minute // 6 * 6, second=0, microsecond=0)
    for back in range(0, 40):
        t = base - timedelta(minutes=6 * back)
        ts = t.strftime("%Y%m%d%H%M") + "00000"
        found = []
        for name, code in SC_RADARS:
            url = (f"https://image.nmc.cn/product/{t.strftime('%Y/%m/%d')}/RDCP/"
                   f"SEVP_AOC_RDCP_SLDAS3_ECREF_{code}_L88_PI_{ts}.PNG")
            if _probe_img(url):
                found.append({"name": name, "code": code, "url": url})
        if found:
            out = {"time": t.astimezone(TZ).isoformat(), "stations": found}
            cache_put("obs:radar:sc", out)
            return out
    return None


def obs_sat_url(now=None):
    """最新风云四号 B 星真彩色云图 URL（中央气象台 WXBL，产品文件名用 UTC 时次，15 分钟更新，白天有数据）。缓存 10 分钟。"""
    cached = cache_get("obs:sat", 600)
    if cached is not None: return cached
    now = now or datetime.now(TZ)
    # v2.10.4: 对齐到 15 分钟产品网格（此前从整点探测，夜间时次无法更新到最新）
    utc = now.astimezone(timezone.utc)
    base = utc.replace(minute=utc.minute // 15 * 15, second=0, microsecond=0)
    for back in range(0, 32):  # 最多回退 8 小时
        t = base - timedelta(minutes=15 * back)
        ts = t.strftime("%Y%m%d%H%M") + "00000"
        url = (f"https://image.nmc.cn/product/{t.strftime('%Y/%m/%d')}/WXBL/medium/"
               f"SEVP_NSMC_WXBL_FY4B_ETCC_ACHN_LNO_PY_{ts}.JPG")
        if _probe_img(url):
            out = {"url": url, "time": t.astimezone(TZ).isoformat()}
            cache_put("obs:sat", out)
            return out
    return None


# ---- v2.10.12: 风云四号红外云图（NSMC 官方 WMS 接口，24 小时可用） ----
# 数据源：国家卫星气象中心 WMS 服务 GEOS_IRX（全球静止卫星 10.8μm 红外拼图，含风云四号，逐小时）
# 服务端限制：bbox 经度跨度需 ≥55° 且图片宽度 ≤700px；故拉取大范围图后裁剪出四川/川西区域并放大。
FY4_WMS_BASE = "https://data.nsmc.org.cn/NSMCAPI/v1/nsmc/image/wms/compose"
FY4_WMS_TIME_URL = ("https://data.nsmc.org.cn/nsmcapi/v1/nsmc/image/animation/datatime/mongodb"
                    "?dataCode=GEO_MULT_GBAL_L2_GGM_IRX_GLL_YYYYMMDD_HHmm_4000M.PNG&hourRange=24")
# 请求的大范围（西边界必须 ≤70、东边界-西边界 ≥55°，服务端限制）
FY4_WMS_BBOX = "70,10,135,55"          # 覆盖中国大部 + 西南，满足跨度 ≥55°
FY4_WMS_W, FY4_WMS_H = 700, 700
# 裁剪出的川西走廊范围（含成都平原 + 川西雪山带）
FY4_CROP = {"lon0": 96.5, "lon1": 112.0, "lat0": 24.0, "lat1": 36.5}
FY4_OUT_W = 900                         # 输出图宽度（放大后）
# v2.10.17: GEOS_IRX 二值图为「白=云、透明=晴空」，透明区叠到浅色底图上无法分辨。
# 将透明(晴空)填充为深蓝夜空色，形成「深底+白云」强对比，云况一目了然。
FY4_CLEAR_RGB = (18, 26, 46)           # 晴空底色（深蓝黑）

def _fy4_fill_clear(img):
    """把 RGBA 图中透明像素替换为晴空深色底，返回不透明 RGBA 图。"""
    from PIL import Image as _Img
    if img.mode != "RGBA":
        img = img.convert("RGBA")
    bg = _Img.new("RGBA", img.size, FY4_CLEAR_RGB + (255,))
    return _Img.alpha_composite(bg, img)

def fy4_irx_png(now=None):
    """拉取最新风云四号红外云图并裁剪川西区域，返回 (PNG字节, 时次dict) 或 None。
    缓存 10 分钟。"""
    cached = cache_get("obs:fy4irx", 600)
    if cached is not None: return cached
    try:
        now = now or datetime.now(TZ)
        # 1) 查询最新可用时次
        r = SESSION.get(FY4_WMS_TIME_URL, timeout=20)
        r.raise_for_status()
        ds = r.json().get("ds") or []
        if not ds:
            return None
        latest = ds[-1]
        dt = latest["dataDate"] + (latest["dataTime"] or "")[:4]   # YYYYMMDDHHmm
        # 2) 请求大范围红外云图（白色云 + 透明底）
        r2 = SESSION.get(FY4_WMS_BASE, params={
            "layers": "GEOS_IRX", "datetime": dt, "request": "GetMap",
            "bbox": FY4_WMS_BBOX, "width": FY4_WMS_W, "height": FY4_WMS_H,
            "version": "1.1.0", "format": "png",
        }, timeout=30)
        r2.raise_for_status()
        if not r2.content or len(r2.content) < 2000:
            return None  # 空白图（无数据）
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(r2.content)).convert("RGBA")
        # v2.10.17: 透明(晴空)填充深色底，避免叠浅色底图时"一片白"分不清云
        img = _fy4_fill_clear(img)
        # 3) 裁剪川西范围（等距投影：x 与经度线性、y 与纬度线性）
        lon0, lon1 = float(FY4_WMS_BBOX.split(",")[0]), float(FY4_WMS_BBOX.split(",")[2])
        lat1, lat0 = float(FY4_WMS_BBOX.split(",")[3]), float(FY4_WMS_BBOX.split(",")[1])
        c = FY4_CROP
        px0 = int((c["lon0"] - lon0) / (lon1 - lon0) * img.width)
        px1 = int((c["lon1"] - lon0) / (lon1 - lon0) * img.width)
        py0 = int((lat1 - c["lat1"]) / (lat1 - lat0) * img.height)
        py1 = int((lat1 - c["lat0"]) / (lat1 - lat0) * img.height)
        crop = img.crop((px0, py0, px1, py1))
        # 4) 放大（LANCZOS 保细节）
        oh = int(FY4_OUT_W * crop.height / crop.width)
        out = crop.resize((FY4_OUT_W, oh), Image.LANCZOS)
        buf = _io.BytesIO()
        out.save(buf, format="PNG")
        # 5) 标注时次（北京时间）
        t_utc = datetime.strptime(dt, "%Y%m%d%H%M").replace(tzinfo=timezone.utc)
        t_bj = t_utc.astimezone(TZ)
        info = {"time_utc": dt, "time_bj": t_bj.isoformat(),
                "bbox": [c["lat0"], c["lon0"], c["lat1"], c["lon1"]],
                "width": FY4_OUT_W, "height": oh}
        out_data = {"png": buf.getvalue(), "info": info}
        cache_put("obs:fy4irx", out_data)
        return out_data
    except Exception as e:
        print(f"[fy4_irx] {type(e).__name__}: {e}", flush=True)
        return None


# ---- v2.10.15: 风云四号红外云图多时次动画（复用 GEOS_IRX WMS，近 2 小时 8 帧合成 GIF） ----
FY4_ANIM_FRAMES = 8          # 动画帧数（近 8 个时次，约 2 小时）
FY4_ANIM_DUR_MS = 700        # 每帧显示时长(ms)
FY4_ANIM_OUT_W = 900         # 输出宽度（与单帧一致）

def fy4_irx_anim(now=None):
    """拉取最近 N 个时次的风云四号红外云图，逐帧裁剪川西区域并合成 GIF 动画。
    返回 {"gif": 字节, "info": {times, bbox, width, height, frames}} 或 None。缓存 10 分钟。
    若可用的历史帧 < 3 帧则返回 None（数据不足）。NSMC 偶发 TLS 断连时自动重试一次。"""
    cached = cache_get("obs:fy4anm", 600)
    if cached is not None: return cached
    for attempt in (1, 2):
        try:
            out = _fy4_irx_anim_impl(now)
            if out is not None:
                cache_put("obs:fy4anm", out)
                return out
            break  # 明确无数据，无需重试
        except Exception as e:
            print(f"[fy4_irx_anim] 第{attempt}次失败 {type(e).__name__}: {e}", flush=True)
            if attempt == 1:
                time.sleep(1.5)
                continue
    return None


def _fy4_irx_anim_impl(now):
    """动画实际实现（供重试包装调用），抛异常表示可重试的传输错误。"""
    from PIL import Image
    import io as _io
    now = now or datetime.now(TZ)
    # 1) 查询时次列表（与单帧同一接口，hourRange=24）
    r = SESSION.get(FY4_WMS_TIME_URL, timeout=20)
    r.raise_for_status()
    ds = r.json().get("ds") or []
    if len(ds) < 3:
        return None
    times = [d["dataDate"] + (d["dataTime"] or "")[:4] for d in ds[-FY4_ANIM_FRAMES:]]
    # 2) 逐帧拉取并裁剪
    lon0, lon1 = float(FY4_WMS_BBOX.split(",")[0]), float(FY4_WMS_BBOX.split(",")[2])
    lat1, lat0 = float(FY4_WMS_BBOX.split(",")[3]), float(FY4_WMS_BBOX.split(",")[1])
    c = FY4_CROP
    px0 = int((c["lon0"] - lon0) / (lon1 - lon0) * FY4_WMS_W)
    px1 = int((c["lon1"] - lon0) / (lon1 - lon0) * FY4_WMS_W)
    py0 = int((lat1 - c["lat1"]) / (lat1 - lat0) * FY4_WMS_H)
    py1 = int((lat1 - c["lat0"]) / (lat1 - lat0) * FY4_WMS_H)
    oh = int(FY4_ANIM_OUT_W * (py1 - py0) / (px1 - px0))
    frames, ok_times = [], []
    for dt in times:
        try:
            r2 = SESSION.get(FY4_WMS_BASE, params={
                "layers": "GEOS_IRX", "datetime": dt, "request": "GetMap",
                "bbox": FY4_WMS_BBOX, "width": FY4_WMS_W, "height": FY4_WMS_H,
                "version": "1.1.0", "format": "png",
            }, timeout=30)
            r2.raise_for_status()
            if not r2.content or len(r2.content) < 2000:
                continue  # 该时次空白，跳过
            img = Image.open(_io.BytesIO(r2.content)).convert("RGBA")
            # v2.10.17: 透明(晴空)填充深色底，避免动画帧叠浅色底图时一片白
            img = _fy4_fill_clear(img)
            crop = img.crop((px0, py0, px1, py1))
            frames.append(crop.resize((FY4_ANIM_OUT_W, oh), Image.LANCZOS))
            ok_times.append(dt)
        except Exception:
            continue
    if len(frames) < 3:
        return None  # 有效帧不足
    # 3) 合成 GIF（循环播放，帧间隔 FY4_ANIM_DUR_MS）
    buf = _io.BytesIO()
    frames[0].save(buf, format="GIF", save_all=True, append_images=frames[1:],
                   duration=FY4_ANIM_DUR_MS, loop=0, disposal=2, optimize=False)
    # 4) 时次标注（北京时间）
    bj_times = []
    for dt in ok_times:
        t_utc = datetime.strptime(dt, "%Y%m%d%H%M").replace(tzinfo=timezone.utc)
        bj_times.append(t_utc.astimezone(TZ).strftime("%m-%d %H:%M"))
    return {"gif": buf.getvalue(),
            "info": {"times": bj_times, "frames": len(ok_times),
                     "bbox": [c["lat0"], c["lon0"], c["lat1"], c["lon1"]],
                     "width": FY4_ANIM_OUT_W, "height": oh,
                     "interval_min": 15}}


# ---- v2.10.13: 风云四号实况云况分析（近似云掩膜 CLM + 云顶高度分层） ----
# 云掩膜：NSMC WMS GEOS_IRX 二值云检测图（白=云、黑=晴，已用 18 个机场 METAR 交叉验证）。
# 云顶分层：Open-Meteo 数值模式低/中/高云量近似（无 API Key 的替代方案）。
def _fy4_irx_latest():
    """拉取最新风云四号红外灰度图（700x700，白=云），带重试 + 5 分钟共享缓存。
    v2.10.18: 云况分析与火烧云共用同一张图，避免重复请求并提高成功率。
    返回 (datetime字符串, PIL灰度图) 或 (None, None)。"""
    cached = cache_get("obs:irximg", 300)
    if cached is not None:
        return cached
    for attempt in (1, 2):
        try:
            r = SESSION.get(FY4_WMS_TIME_URL, timeout=20)
            r.raise_for_status()
            ds = (r.json().get("ds") or [])
            if not ds:
                return None, None
            latest = ds[-1]
            dt = latest["dataDate"] + (latest["dataTime"] or "")[:4]
            r2 = SESSION.get(FY4_WMS_BASE, params={
                "layers": "GEOS_IRX", "datetime": dt, "request": "GetMap",
                "bbox": FY4_WMS_BBOX, "width": FY4_WMS_W, "height": FY4_WMS_H,
                "version": "1.1.0", "format": "png",
            }, timeout=30)
            r2.raise_for_status()
            if not r2.content or len(r2.content) < 2000:
                return None, None
            from PIL import Image as _Img
            img = _Img.open(io.BytesIO(r2.content)).convert("L")
            out = (dt, img)
            cache_put("obs:irximg", out)
            return out
        except Exception as e:
            print(f"[fy4_irx] 第{attempt}次拉取失败 {type(e).__name__}: {e}", flush=True)
            if attempt == 1:
                time.sleep(1.5)
    return None, None


def fy4_cloud_analysis(lat, lon, now=None):
    """返回观测点及周边风云四号红外云检测统计 + 云顶高度分层，缓存 10 分钟。失败返回 None。"""
    key = f"obs:fy4clm:{float(lat):.2f},{float(lon):.2f}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        now = now or datetime.now(TZ)
        # 1) 拉最新风云四号红外灰度图（共享缓存 + 重试）
        dt, img = _fy4_irx_latest()
        if img is None:
            return None
        px = img.load()
        W, H = img.size
        # 地理范围：lon 70→135（x 从左到右），lat 55→10（y 从上到下）
        lon0, lon1, lat1, lat0 = 70.0, 135.0, 55.0, 10.0
        cx = int((lon - lon0) / (lon1 - lon0) * W)
        cy = int((lat1 - lat) / (lat1 - lat0) * H)
        cx = max(0, min(W - 1, cx)); cy = max(0, min(H - 1, cy))
        # 中心像元：灰度 ≥128 视为云（白=云）
        center_cloud = bool(px[cx, cy] >= 128)
        # 周围 9×9=81 像元统计（约 90km×90km 范围，4km 源数据经服务端降采样）
        r_ = 4
        n_cloud = n_clear = 0
        for yy in range(max(0, cy - r_), min(H, cy + r_ + 1)):
            for xx in range(max(0, cx - r_), min(W, cx + r_ + 1)):
                if px[xx, yy] >= 128:
                    n_cloud += 1
                else:
                    n_clear += 1
        total = n_cloud + n_clear
        stats = {
            "total": total, "valid": total,
            "valid_rate": round(total / total * 100, 1) if total else 0.0,
            "cloudy": n_cloud,
            "cloudy_rate": round(n_cloud / total * 100, 1) if total else 0.0,
            "clear": n_clear,
            "clear_rate": round(n_clear / total * 100, 1) if total else 0.0,
            "cloud_ratio": round(n_cloud / total * 100, 1) if total else 0.0,
            "clear_ratio": round(n_clear / total * 100, 1) if total else 0.0,
        }
        t_utc = datetime.strptime(dt, "%Y%m%d%H%M").replace(tzinfo=timezone.utc)
        t_bj = t_utc.astimezone(TZ)
        # 3) 云顶高度分层：Open-Meteo 当前时次低/中/高云量（近似）
        top = None
        try:
            om = SESSION.get("https://api.open-meteo.com/v1/forecast", params={
                "latitude": lat, "longitude": lon,
                "current": "cloud_cover_low,cloud_cover_mid,cloud_cover_high",
            }, timeout=20)
            om.raise_for_status()
            c = om.json().get("current") or {}
            top = {"low": c.get("cloud_cover_low"), "mid": c.get("cloud_cover_mid"),
                   "high": c.get("cloud_cover_high")}
        except Exception as e:
            print(f"[fy4_cloud top] {type(e).__name__}: {e}", flush=True)
        out = {
            "time_utc": dt, "time_bj": t_bj.isoformat(),
            "center": {"lat": round(lat, 4), "lon": round(lon, 4),
                       "row": cy, "col": cx,
                       "cloud": center_cloud,
                       "text": "有云" if center_cloud else "晴空"},
            "stats": stats, "top": top,
            "note": "云况基于国家卫星气象中心风云四号红外云检测（GEOS_IRX，白=云）；"
                    "云顶分层基于 Open-Meteo 数值模式（低/中/高云量）近似",
        }
        cache_put(key, out)
        return out
    except Exception as e:
        print(f"[fy4_cloud] {type(e).__name__}: {e}", flush=True)
        return None


# ---- v2.10.19: 实时火烧云潜力预报（支持 风云四号 / KMA 千里眼2A 双数据源） ----
# 原理：火烧云 = 日落/日出方向存在「中高云」+ 云量适中 + 太阳接近地平线。
#   - 中高云识别：
#       · 风云四号：红外云图灰度≈亮温，云顶越冷(灰度高)越高，灰度 160-255 视为中高云。
#       · KMA GK2A：CIRA SLIDER band_13(10.3µm 红外)，ircimss2 增强色标索引 >=176
#         （亮温约 -30℃ 以下）视为中高云。
#   - 太阳方位扇区：取太阳方位角 ±70° 扇形内的云，是火烧云真正会出现的天空范围。
#   - 评分：扇区内中高云覆盖率 15%-75% 最佳 + 云顶冷度 + 太阳高度角窗口(-6°~+6°)。
# 输出：潜力评分 + 可绘制范围（中高云像元集合，前端 Canvas 渐变"丝滑"叠加）。

def _azimuth_diff(a, b):
    """两方位角(0-360)之差，返回 0-180。"""
    d = abs(a - b) % 360
    return d if d <= 180 else 360 - d

# ---- v2.10.20: 火烧云"云底受光"几何判断 ----
# 几何模型（观测者为原点，x=水平球面距离(km)，h=海拔高度(km)，R=地球半径）：
#   视线高度公式   h(x) = θ·x + x²/(2R)     （θ=视线仰角(弧度)，第一项为仰角上升，第二项为地球曲率修正）
#   由上式反推云边界(云底)高度角：θ_edge = h_base/x - x/(2R)
#   判断：只有太阳光线高度角 θ_sun 满足
#       -x/(2R) < θ_sun < θ_edge
#   时，阳光才能从云边界下方的空隙穿过、照射到云层下方照亮云底 → 真火烧云；
#   否则云底在阴影中（光线高于云底=照云顶，低于地表=被地球遮挡）。
FIRE_R = 6371.0                # 地球半径 km
FIRE_CLOUD_THICK_KM = 1.5      # 中高云平均厚度估计（云顶→云底，卫星只见云顶）

def _cloud_top_km_fy4(g):
    """风云四号红外灰度 → 云顶高度(km)。灰度≈亮温(越高越冷)，标准大气递减率反推。"""
    t = -32.0 - (g - 160.0) / 95.0 * 53.0     # g=160→-32℃, g=255→-85℃
    return max(1.5, (15.0 - t) / 6.5)         # 地面15℃, 6.5℃/km

def _cloud_top_km_kma(iv):
    """KMA ircimss2 色标索引 → 云顶高度(km)。索引≥176 对应亮温≤-30℃。"""
    t = -30.0 - (iv - 176.0) / 79.0 * 55.0    # iv=176→-30℃, iv=255→-85℃
    return max(1.5, (15.0 - t) / 6.5)

def _cloud_base_lit(sun_elev_deg, dist_km, h_top_km):
    """几何模型：给定太阳高度角、云距离、云顶高度，判断云底能否被低角度阳光照亮。
    返回 (受光bool, 云底高度km, 云边界高度角°)。"""
    if dist_km <= 0:
        return False, 0.0, 0.0
    h_base = max(0.5, h_top_km - FIRE_CLOUD_THICK_KM)
    theta_sun = math.radians(sun_elev_deg)
    theta_edge = h_base / dist_km - dist_km / (2.0 * FIRE_R)
    lit = theta_edge > theta_sun > -dist_km / (2.0 * FIRE_R)
    return lit, h_base, math.degrees(theta_edge)

def _fire_cloud_score(elev, az, sector_cloud_rate, sector_lit_rate, avg_g):
    """双源共用的火烧云评分（基于"云底受光"中高云占比）。返回 (score, level, is_window, phase)。"""
    if 15 <= sector_lit_rate <= 75:
        s_cover = 100 - abs(sector_lit_rate - 40) * 1.2   # 40% 附近最优
    else:
        s_cover = max(0, 100 - abs(sector_lit_rate - 40) * 2.5)
    if avg_g:
        s_cold = min(100, (avg_g - 150) * 1.8)
    else:
        s_cold = 0
    if -8 <= elev <= 8:
        s_sun = 100 - abs(elev) * 9
    else:
        s_sun = max(0, 55 - abs(elev) * 4)
    score = int(round(0.5 * s_cover + 0.3 * s_cold + 0.2 * s_sun))
    score = max(0, min(100, score))
    is_window = -10 <= elev <= 10
    phase = ("黎明（日出）" if az < 180 else "黄昏（日落）") if is_window else \
            ("白天（太阳偏高）" if elev > 10 else "夜间（太阳低于地平线）")
    level = "高" if score >= 70 else ("中" if score >= 45 else ("低" if score >= 25 else "无"))
    return score, level, is_window, phase

# ---------- KMA GK2A 数据源（CIRA SLIDER，免登录，Web Mercator 瓦片） ----------
SLIDER_CFG = dict(lon0=128.0, sat_alt=42171.7, max_rad_x=0.150618, max_rad_y=0.150485,
                  disk_radius_x_z0=337, disk_radius_y_z0=336, tile_size=678, z_max=5)
SLIDER_H_ALT, SLIDER_R, SLIDER_LON0 = 42171.7, 6378.1, math.radians(128.0)

def _slider_inv_latlon(e, s):
    """GK2A GEOS 扫描角(东正,北正) → (lat, lon)。球面近似。"""
    ce, se, cs, ss = math.cos(e), math.sin(e), math.cos(s), math.sin(s)
    A = ce * cs
    disc = SLIDER_H_ALT*SLIDER_H_ALT*A*A - (SLIDER_H_ALT*SLIDER_H_ALT - SLIDER_R*SLIDER_R)
    if disc < 0: return None
    r = SLIDER_H_ALT*A - math.sqrt(disc)
    if r <= 0: return None
    x = SLIDER_H_ALT - r*A; y = r*se; z = r*ce*ss
    return math.degrees(math.asin(z/SLIDER_R)), math.degrees(math.atan2(y, x)) + 128.0

def _slider_latest_ts():
    r = SESSION.get("https://slider.cira.colostate.edu/data/json/gk2a/full_disk/band_13/latest_times.json", timeout=15)
    r.raise_for_status()
    return str(r.json()["timestamps_int"][0])

def _slider_full_png(ts, z=1):
    """下载 GK2A 全盘 band_13 瓦片拼图（P 模式，索引即 ircimss2 色标值）。缓存 10 分钟。"""
    key = f"obs:kmaimg:{ts}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    ymd = f"{ts[:4]}/{ts[4:6]}/{ts[6:8]}"
    n = 2 ** z
    from PIL import Image as _Img
    imgs = []
    for rr in range(n):
        for cc in range(n):
            r = SESSION.get(
                f"https://slider.cira.colostate.edu/data/imagery/{ymd}/gk2a---full_disk/band_13/{ts}/{z:02d}/{rr:03d}_{cc:03d}.png",
                timeout=25)
            r.raise_for_status()
            imgs.append(_Img.open(io.BytesIO(r.content)).convert("P"))
    W = imgs[0].width
    full = _Img.new("P", (W*n, W*n))
    for i, im in enumerate(imgs):
        full.paste(im, ((i % n)*W, (i // n)*W))
    cache_put(key, full)
    return full

def _fire_cloud_analyze_kma(lat, lon, elev, az):
    """GK2A 太阳扇区中高云分析（含"云底受光"几何过滤）。
    返回 (cloud_rate, mh_rate, lit_rate, avg_g, cells, ts) 或 None。"""
    SECTOR, MAX_DIST = 70.0, 1500.0
    ts = _slider_latest_ts()
    full = _slider_full_png(ts)
    arr = np.array(full)
    Hh, Ww = arr.shape
    z = 1
    scale = 2 ** (SLIDER_CFG["z_max"] - z)
    cx = cy = (SLIDER_CFG["tile_size"]/2) * 2**SLIDER_CFG["z_max"] / scale
    sx_r = SLIDER_CFG["disk_radius_x_z0"] * 2**SLIDER_CFG["z_max"] / scale
    sy_r = SLIDER_CFG["disk_radius_y_z0"] * 2**SLIDER_CFG["z_max"] / scale
    max_x, max_y = SLIDER_CFG["max_rad_x"], SLIDER_CFG["max_rad_y"]
    CELL = 2
    n_tot = n_cld = n_mh = n_lit = 0
    mid_high = []
    ys, xs = np.mgrid[CELL//2:Hh:CELL, CELL//2:Ww:CELL]
    for yy, xx in zip(ys.ravel(), xs.ravel()):
        iv = int(arr[yy, xx])
        if iv <= 0: continue
        e = (xx - cx)/sx_r * max_x
        s = -(yy - cy)/sy_r * max_y
        if e*e + s*s > 0.0232: continue   # 圆盘外（≈0.151²）
        ll = _slider_inv_latlon(e, s)
        if ll is None: continue
        plat, plon = ll
        if not (5 <= plat <= 65 and 60 <= plon <= 190): continue
        dist, brg = haversine_bearing(lat, lon, plat, plon)
        if dist > MAX_DIST or _azimuth_diff(brg, az) > SECTOR: continue
        n_tot += 1
        if iv >= 128: n_cld += 1
        if iv >= 176:
            n_mh += 1
            lit, _, _ = _cloud_base_lit(elev, dist, _cloud_top_km_kma(iv))
            if lit:
                n_lit += 1
                mid_high.append((plon, plat, iv))
    if n_tot < 20:
        return None
    if mid_high:
        avg_g = sum(x[2] for x in mid_high) / len(mid_high)
    else:
        avg_g = 0
    return (n_cld/n_tot*100, n_mh/n_tot*100, n_lit/n_tot*100, avg_g, mid_high, ts)

def _fire_cloud_analyze_fy4(lat, lon, elev, az):
    """风云四号太阳扇区中高云分析（含"云底受光"几何过滤）。
    返回 (cloud_rate, mh_rate, lit_rate, avg_g, cells, dt) 或 None。"""
    SECTOR, MAX_DIST = 70.0, 1500.0
    dt, img = _fy4_irx_latest()
    if img is None:
        return None
    px = img.load(); W, H = img.size
    lon0, lon1, lat1, lat0 = 70.0, 135.0, 55.0, 10.0
    CELL = 3
    mid_high = []
    n_sector_total = n_sector_cloud = n_sector_midhigh = n_sector_lit = 0
    for yy in range(0, H, CELL):
        lat_y = lat1 - (lat1 - lat0) * (yy + 0.5) / H
        for xx in range(0, W, CELL):
            g = px[xx, yy]
            if g < 40: continue
            lon_x = lon0 + (lon1 - lon0) * (xx + 0.5) / W
            dist, bearing = haversine_bearing(lat, lon, lat_y, lon_x)
            if dist > MAX_DIST or _azimuth_diff(bearing, az) > SECTOR: continue
            n_sector_total += 1
            if g >= 128: n_sector_cloud += 1
            if g >= 160:
                n_sector_midhigh += 1
                lit, _, _ = _cloud_base_lit(elev, dist, _cloud_top_km_fy4(g))
                if lit:
                    n_sector_lit += 1
                    mid_high.append((lon_x, lat_y, g))
    if n_sector_total < 20:
        return None
    if mid_high:
        avg_g = sum(x[2] for x in mid_high) / len(mid_high)
    else:
        avg_g = 0
    return (n_sector_cloud/n_sector_total*100, n_sector_midhigh/n_sector_total*100,
            n_sector_lit/n_sector_total*100, avg_g, mid_high, dt)

def fire_cloud_forecast(lat, lon, now=None, src="fy4"):
    """实时火烧云潜力预报。src: fy4(风云四号) / kma(GK2A 千里眼2A)。缓存 10 分钟。"""
    src = src if src in ("fy4", "kma") else "fy4"
    key = f"obs:firec:{src}:{float(lat):.2f},{float(lon):.2f}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        now = now or datetime.now(TZ)
        elev, az = solar_position(now, lat, lon)
        if src == "kma":
            res = _fire_cloud_analyze_kma(lat, lon, elev, az)
            time_tag = None
            src_name = "KMA 千里眼2A (GK2A)"
            src_tag = "kma"
        else:
            res = _fire_cloud_analyze_fy4(lat, lon, elev, az)
            time_tag = None
            src_name = "风云四号"
            src_tag = "fy4"
        if res is None:
            return None
        sector_cloud_rate, sector_mh_rate, sector_lit_rate, avg_g, mid_high, time_tag = res
        score, level, is_window, phase = _fire_cloud_score(
            elev, az, sector_cloud_rate, sector_lit_rate, avg_g)
        if len(mid_high) > 2500:
            step = math.ceil(len(mid_high) / 2500)
            mid_high = mid_high[::step]
        cells = [{"lon": round(x, 3), "lat": round(y, 3), "g": g_} for x, y, g_ in mid_high]
        t_utc = datetime.strptime(time_tag[:12], "%Y%m%d%H%M").replace(tzinfo=timezone.utc)
        out = {
            "ok": True,
            "src": src_tag, "src_name": src_name,
            "time_utc": time_tag, "time_bj": t_utc.astimezone(TZ).isoformat(),
            "sun": {"elev": round(elev, 1), "az": round(az, 1), "phase": phase},
            "score": score,
            "level": level,
            "sector": {"az_center": round(az, 1), "half_width": 70.0, "max_dist_km": 1500,
                       "cloud_rate": round(sector_cloud_rate, 1),
                       "midhigh_rate": round(sector_mh_rate, 1),
                       "lit_rate": round(sector_lit_rate, 1),
                       "avg_gray": round(avg_g, 1) if mid_high else None},
            "window": is_window,
            "cells": cells,
            "note": f"火烧云潜力 = 观测点周边1500km内、日落/日出方向中高云覆盖率 + 云底受光率 + 云顶冷度 + 太阳高度角综合；"
                    f"数据源：{src_name}红外云图；中高云（卷云/高层云）最易被染红，云量适中(15-75%)最佳。"
                    f"绘制范围仅为太阳扇区内『云底受光』的中高云：阳光须从云边界下方空隙穿过照亮云底"
                    f"（几何模型 θ_sun∈(-D/2R, h_base/D-D/2R)），未受光云底不会形成火烧云，已剔除。",
        }
        cache_put(key, out)
        return out
    except Exception as e:
        print(f"[fire_cloud:{src}] {type(e).__name__}: {e}", flush=True)
        return None


# ---- v2.10.14: NASA GIBS 卫星云顶高度（免登录，Himawari 红外 + MODIS CTH） ----
# 数据源：NASA GIBS WMS（gibs.earthdata.nasa.gov），免登录免费。
#  - Himawari_AHI_Band13_Clean_Infrared：10.4µm 增强红外，灰度 0-255 线性映射 TBB
#    -92.6℃(白/冷) ~ +57℃(黑/暖)，灰度越高=云顶越冷越高，10 分钟时次（GIBS 有数小时延迟）
#  - MODIS_Aqua/Terra_Cloud_Top_Height_Day：彩虹色标 0-20km 定量云顶高度，日级（每天过境 2-4 次）
# 注意：WMS 1.3.0 + EPSG:4326 的 BBOX 轴序为 纬度优先(lat,lon)。
GIBS_WMS = "https://gibs.earthdata.nasa.gov/wms/epsg4326/best/"
GIBS_BBOX = "24,96,37,112"      # 川西（纬度,经度 轴序）
GIBS_W, GIBS_H = 640, 520
# MODIS CTH 彩虹色标锚点（从 GIBS 官方图例提取，0%→100% 对应 0→20km）
GIBS_CTH_ANCHORS = [(0, (241, 0, 0)), (6.25, (170, 0, 0)), (12.5, (110, 0, 0)),
                    (18.75, (112, 1, 2)), (25, (124, 91, 5)), (31.25, (240, 190, 64)),
                    (37.5, (255, 255, 0)), (43.75, (0, 220, 0)), (50, (0, 136, 0)),
                    (56.25, (0, 80, 0)), (62.5, (0, 80, 0)), (68.75, (0, 136, 238)),
                    (75, (0, 0, 255)), (81.25, (0, 0, 170)), (87.5, (0, 0, 100)),
                    (93.75, (183, 15, 141)), (100, (102, 0, 119))]
GIBS_CTH_MAX = 20.0  # km


def _gibs_color_pos(r, g, b, anchors):
    """在锚点色标中找最近颜色对应的位置百分比(0-100)。"""
    best, best_d = 0.0, 1e18
    for pos, (ar, ag, ab) in anchors:
        d = (r - ar) ** 2 + (g - ag) ** 2 + (b - ab) ** 2
        if d < best_d:
            best_d, best = d, pos
    return best


def gibs_cloud_analysis(lat, lon):
    """NASA GIBS 卫星云顶高度分析：Himawari 红外灰度分级 + MODIS CTH 定量统计。缓存 15 分钟。"""
    key = f"obs:gibs:{float(lat):.2f},{float(lon):.2f}"
    cached = cache_get(key, 900)
    if cached is not None: return cached
    try:
        from PIL import Image as _Img
        # bbox 常量解析：GIBS_BBOX="lat_min,lon_min,lat_max,lon_max"
        bbox = [float(v) for v in GIBS_BBOX.split(",")]
        lat_bot, lon_left, lat_top, lon_right = bbox
        out = {"ir13": None, "cth": None, "note": ""}
        # 1) Himawari 红外（最新时次，近实时参考）
        try:
            r = SESSION.get(GIBS_WMS, params={
                "SERVICE": "WMS", "REQUEST": "GetMap", "VERSION": "1.3.0",
                "LAYERS": "Himawari_AHI_Band13_Clean_Infrared", "FORMAT": "image/png",
                "TRANSPARENT": "TRUE", "BBOX": GIBS_BBOX, "CRS": "EPSG:4326",
                "WIDTH": GIBS_W, "HEIGHT": GIBS_H,
            }, timeout=40)
            r.raise_for_status()
            img = _Img.open(io.BytesIO(r.content)).convert("L")
            g = img.load(); W, H = img.size
            cx = int((lon - lon_left) / (lon_right - lon_left) * W)
            cy = int((lat_top - lat) / (lat_top - lat_bot) * H)
            cx = max(0, min(W - 1, cx)); cy = max(0, min(H - 1, cy))
            # 9x9 像元灰度统计（约 55km 范围）
            rr = 4
            vals = []
            for yy in range(max(0, cy - rr), min(H, cy + rr + 1)):
                for xx in range(max(0, cx - rr), min(W, cx + rr + 1)):
                    vals.append(g[xx, yy])
            if vals:
                # 灰度 → TBB（0=+57℃暖 → 255=-92.6℃冷），越高=云顶越冷越高
                tbbs = [57.0 - (v / 255.0) * 149.6 for v in vals]
                mean_tbb = sum(tbbs) / len(tbbs)
                def band_of(tbb):
                    if tbb < -35: return "high"
                    if tbb < -12: return "mid"
                    return "low"
                n_high = sum(1 for t in tbbs if t < -35)
                n_mid = sum(1 for t in tbbs if -35 <= t < -12)
                n_low = sum(1 for t in tbbs if t >= -12)
                n = len(tbbs)
                out["ir13"] = {
                    "time": "最新时次", "total": n,
                    "mean_tbb": round(mean_tbb, 1),
                    "high": n_high, "high_rate": round(n_high / n * 100, 1),
                    "mid": n_mid, "mid_rate": round(n_mid / n * 100, 1),
                    "low": n_low, "low_rate": round(n_low / n * 100, 1),
                    "note": "Himawari-8/9 10.4μm 增强红外（NASA GIBS），灰度→亮温线性映射，"
                            "亮温越低=云顶越高；仅供相对分级参考",
                }
        except Exception as e:
            print(f"[gibs ir13] {type(e).__name__}: {e}", flush=True)
        # 2) MODIS CTH（白天定量云顶高度）
        for layer in ["MODIS_Aqua_Cloud_Top_Height_Day", "MODIS_Terra_Cloud_Top_Height_Day"]:
            try:
                r = SESSION.get(GIBS_WMS, params={
                    "SERVICE": "WMS", "REQUEST": "GetMap", "VERSION": "1.3.0",
                    "LAYERS": layer, "FORMAT": "image/png", "TRANSPARENT": "TRUE",
                    "BBOX": GIBS_BBOX, "CRS": "EPSG:4326",
                    "WIDTH": GIBS_W, "HEIGHT": GIBS_H,
                }, timeout=40)
                r.raise_for_status()
                img = _Img.open(io.BytesIO(r.content)).convert("RGBA")
                px = img.load(); W, H = img.size
                cx = int((lon - lon_left) / (lon_right - lon_left) * W)
                cy = int((lat_top - lat) / (lat_top - lat_bot) * H)
                cx = max(0, min(W - 1, cx)); cy = max(0, min(H - 1, cy))
                rr = 4
                heights = []
                for yy in range(max(0, cy - rr), min(H, cy + rr + 1)):
                    for xx in range(max(0, cx - rr), min(W, cx + rr + 1)):
                        pr, pg, pb, pa = px[xx, yy]
                        if pa < 128:  # 透明=无数据/晴空
                            continue
                        pos = _gibs_color_pos(pr, pg, pb, GIBS_CTH_ANCHORS)
                        heights.append(pos / 100.0 * GIBS_CTH_MAX)
                if len(heights) >= 5:
                    heights.sort()
                    n = len(heights)
                    n_low = sum(1 for h in heights if h < 3)
                    n_mid = sum(1 for h in heights if 3 <= h < 7)
                    n_high = sum(1 for h in heights if h >= 7)
                    out["cth"] = {
                        "source": "Aqua" if "Aqua" in layer else "Terra",
                        "total": n, "max": round(heights[-1], 1),
                        "min": round(heights[0], 1),
                        "mean": round(sum(heights) / n, 1),
                        "median": round(heights[n // 2], 1),
                        "low": n_low, "low_rate": round(n_low / n * 100, 1),
                        "mid": n_mid, "mid_rate": round(n_mid / n * 100, 1),
                        "high": n_high, "high_rate": round(n_high / n * 100, 1),
                        "note": "MODIS 云顶高度（NASA GIBS 彩虹色标 0-20km），日间过境定量产品",
                    }
                    break  # Aqua 有数据就够，否则用 Terra
            except Exception as e:
                print(f"[gibs cth {layer}] {type(e).__name__}: {e}", flush=True)
        if not out["ir13"] and not out["cth"]:
            return None
        cache_put(key, out)
        return out
    except Exception as e:
        print(f"[gibs] {type(e).__name__}: {e}", flush=True)
        return None


# ---- v2.9.6: 云图未来 3 小时趋势 + 当前云层判断 ----
# v2.10.7: 实况云型识别（Python 版）——与前端剖面 11 云属多因子判定一致
GN_ZH_PY = {"ci": "卷云", "cc": "卷积云", "cs": "卷层云", "ac": "高积云", "as": "高层云",
            "ns": "雨层云", "sc": "层积云", "st": "层云", "cu": "积云", "cb": "积雨云", "fn": "碎积云"}


def genus_of_pt(key, cover, rh, pop, low_cov):
    """实况云型多因子判定：高度层 × 云量 × 湿度 × 降水概率 × 低云总量。
    pop 为 Open-Meteo 降水概率（0-100%），rh 为相对湿度（0-100%）。
    返回云属键（ci/cc/cs/ac/as/ns/sc/st/cu/cb/fn）或 None；cb/ns 视为跨层云塔。"""
    if cover is None or cover <= 3:
        return None
    if key == "high":
        if pop >= 60 and cover >= 70: return "cs"   # 系统性降水伴随的卷层云
        if pop >= 35 and cover >= 45: return "cc"   # 对流性降水伴随的卷积云
        if cover >= 80 and rh >= 75: return "cs"    # 大片均匀卷层云
        if cover >= 50 and rh < 70: return "cc"     # 成行/成波状的卷积云
        return "ci"
    if key == "mid":
        if pop >= 60 and cover >= 70: return "ns"   # 连续性降水的中层（雨层云上界）
        if cover >= 75 and rh >= 80: return "as"    # 均匀幕状高层云
        if cover >= 55: return "ac"                 # 块状高积云
        if cover >= 25 and rh < 65: return "ac"
        return "as" if cover >= 15 else None
    # 低云：先判跨层强对流/系统性降水
    if pop >= 65 and cover >= 75: return "ns"       # 连续性降水 → 雨层云
    if pop >= 35: return "cb"                       # 对流性强降水 → 积雨云
    if cover >= 80 and low_cov >= 80 and rh >= 90: return "ns"  # 低云极厚+近饱和
    if cover >= 78 and rh >= 90: return "st"        # 厚而均匀、近饱和 → 层云
    if cover >= 80 and rh >= 80 and pop >= 30: return "cb"      # 厚积云伴有降水
    if cover >= 50 and rh >= 78: return "sc"        # 层积云（较稳定）
    if cover >= 30 and rh < 72: return "cu"         # 低湿度块状 → 积云
    if cover >= 15 and rh >= 85: return "fn"        # 零散碎云+高湿 → 碎积云
    if cover >= 12: return "cu"
    return None


def cloud_trend(lat, lon):
    """未来 3 小时云量趋势与当前云层判断。
    采样点：观测点 + 沿西岭方向走廊中段/远端（Open-Meteo forecast 单次多站请求）。
    返回 4 小时（当前+3）的分层云量序列与趋势/判断文案。缓存 10 分钟。"""
    key = f"obs:ct:{lat:.2f},{lon:.2f}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        seg = interpolate_great_circle(lat, lon, MOUNTAINS[0]["lat"], MOUNTAINS[0]["lon"], 10)
        pts = [(lat, lon), seg[2], seg[4]]  # 观测点 / 走廊中段 / 远端
        params = {
            "latitude": ",".join(f"{p[0]:.4f}" for p in pts),
            "longitude": ",".join(f"{p[1]:.4f}" for p in pts),
            "hourly": ("cloud_cover,cloud_cover_low,cloud_cover_mid,cloud_cover_high,"
                       "precipitation_probability,relative_humidity_2m"),
            "forecast_days": 1, "timezone": "Asia/Shanghai",
        }
        raw = http_get_json("https://api.open-meteo.com/v1/forecast", params=params, timeout=25)
        rows = raw if isinstance(raw, list) else [raw]
        if not rows: return None
        now0 = datetime.now(TZ).replace(minute=0, second=0, microsecond=0)
        ref = now0.strftime("%Y-%m-%dT%H:%M")
        hours, n = [], 4
        for h in range(n):
            hours.append({"time": (now0 + timedelta(hours=h)).isoformat(),
                          "cloud": 0, "low": 0, "mid": 0, "high": 0, "pop": 0, "rh": 0})
        for row in rows:
            hh = row.get("hourly") or {}
            times = hh.get("time") or []
            if not times: continue
            idx = next((i for i, t in enumerate(times) if t >= ref), 0)
            for h in range(n):
                j = idx + h
                if j >= len(times): break
                def _v(arr):
                    try: return float(arr[j]) if arr and j < len(arr) and arr[j] is not None else 0.0
                    except (TypeError, ValueError, IndexError): return 0.0
                hours[h]["cloud"] += _v(hh.get("cloud_cover")) / len(rows)
                hours[h]["low"] += _v(hh.get("cloud_cover_low")) / len(rows)
                hours[h]["mid"] += _v(hh.get("cloud_cover_mid")) / len(rows)
                hours[h]["high"] += _v(hh.get("cloud_cover_high")) / len(rows)
                hours[h]["pop"] += _v(hh.get("precipitation_probability")) / len(rows)
                hours[h]["rh"] += _v(hh.get("relative_humidity_2m")) / len(rows)
        for h in hours:
            for k in ("cloud", "low", "mid", "high", "pop", "rh"):
                h[k] = round(h[k])
            dl = {"low": h["low"], "mid": h["mid"], "high": h["high"]}
            dom = max(dl, key=dl.get)
            h["genus"] = genus_of_pt(dom, dl[dom], h["rh"], h["pop"], h["low"])
            h["genus_zh"] = GN_ZH_PY.get(h["genus"], "")
        cur = hours[0]
        fut = hours[1:]
        avg = sum(x["cloud"] for x in fut) / len(fut) if fut else cur["cloud"]
        delta = round(avg - cur["cloud"])
        if delta > 12: dir_, arrow = "inc", "↑ 增多"
        elif delta < -12: dir_, arrow = "dec", "↓ 减少"
        else: dir_, arrow = "stable", "→ 稳定"
        # 当前云型识别（v2.10.7：11 云属多因子判定 + 跨层云塔，与剖面识别体系一致）
        layers = {"low": cur["low"], "mid": cur["mid"], "high": cur["high"]}
        dom = max(layers, key=layers.get)
        gd = genus_of_pt(dom, layers[dom], cur["rh"], cur["pop"], cur["low"])
        if gd == "cb": layer = "积雨云/强对流"
        elif gd == "ns": layer = "雨层云/系统性降水"
        elif gd == "st": layer = "层云/低云底"
        elif gd == "sc": layer = "层积云"
        elif gd == "cu": layer = "积云"
        elif gd == "fn": layer = "碎积云"
        elif gd == "ac": layer = "高积云"
        elif gd == "as": layer = "高层云"
        elif gd == "cc": layer = "卷积云"
        elif gd == "cs": layer = "卷层云"
        elif gd == "ci": layer = "卷云"
        elif cur["cloud"] < 20: layer = "晴空"
        elif max(cur["low"], cur["mid"], cur["high"]) < 25: layer = "疏云"
        else: layer = {"low": "低云为主", "mid": "中云为主"}.get(dom, "高云为主")
        judge_by_genus = {
            "cb": "积雨云（强对流云塔）发展，局地阵雨/雷暴概率高，观山条件差且注意天气突变。",
            "ns": "雨层云（深厚云塔）控制，系统性降水持续，洗尘明显但观山窗口需等雨止云散。",
            "st": "层云云底低，易罩住平原与峰线，观山前景偏差。",
            "sc": "层积云铺展，云底偏低，峰顶方向易被遮蔽。",
            "cu": "淡积云为主，云间有空隙，峰顶方向大概率可见。",
            "fn": "碎积云零散飘移，偶有云团遮挡，总体可看山。",
            "ac": "高积云（中云）为主，可能遮住山腰/峰顶，能否见山取决于云底高度。",
            "as": "高层云幕铺开，天色阴沉，峰顶轮廓可能被遮。",
            "cc": "卷积云波状排列，不影响峰顶可见性，通透度尚可。",
            "cs": "卷层云幕覆盖，峰顶仍可见但通透度降低，可能出现日晕。",
            "ci": "卷云（高云）为主，一般不影响峰顶轮廓，通透度略降。",
        }
        if gd: judge = judge_by_genus[gd]
        elif layer == "晴空": judge = "云量很少，天空通透，适合观山（重点留意能见度与气溶胶）。"
        elif layer == "疏云": judge = "云量不多，峰顶方向大概率可见，偶有云团飘过。"
        elif dom == "low": judge = "低云层厚且云底低，容易罩住成都平原与峰线，观山前景偏差。"
        elif dom == "mid": judge = "中云可能遮住山腰/峰顶，能否见山取决于云底高度。"
        else: judge = "高云为主，一般不影响峰顶轮廓，但会降低通透度与霞光效果。"
        if cur["pop"] >= 50: judge += " 未来降水概率高，注意洗尘/降雨对通透度的影响。"
        trend_text = f"未来 3 小时总云量{arrow}（当前 {cur['cloud']}% → 均 {avg:.0f}%）；"
        if max(x["pop"] for x in fut) >= 50: trend_text += "降水概率偏高，有降雨/阵雨可能。"
        elif max(x["pop"] for x in fut) >= 25: trend_text += "午后局地降水概率中等。"
        else: trend_text += "无明显降水信号。"
        out = {"points": ["观测点", "走廊中段", "走廊远端"], "hours": hours,
               "trend": {"delta": delta, "dir": dir_, "arrow": arrow, "text": trend_text},
               "judge": {"layer": layer, "text": judge, "genus": gd}, "generated": datetime.now(TZ).isoformat()}
        cache_put(key, out)
        return out
    except Exception:
        return None


# ---- v2.9.7: 云图影响范围标注（观测点+雪山+视线走廊+走廊云量） ----
# 风云四号 ACHN 等距投影（全国源图范围）与前端裁剪显示的西南范围
SAT_SRC_BBOX = {"lon_min": 65.0, "lon_max": 145.0, "lat_min": 10.0, "lat_max": 60.0}
SAT_VIEW_BBOX = {"lon_min": 97.0, "lon_max": 110.0, "lat_min": 26.0, "lat_max": 36.0}  # 西南地区（覆盖成都+四座雪山+川西）


def cloud_map_data(lat, lon, model="best_match"):
    """为云图叠加标注提供数据：观测点、四座雪山、每条视线走廊的采样点与当前分层云量。
    复用 Open-Meteo 走廊预报（缓存），取当前时次；失败则返回无云量的走廊点。缓存 10 分钟。"""
    key = f"obs:cm:{lat:.2f},{lon:.2f}"
    cached = cache_get(key, 600)
    if cached is not None: return cached
    try:
        pts, owners = [], []
        corridors = []
        for m in MOUNTAINS:
            seg = interpolate_great_circle(lat, lon, m["lat"], m["lon"], 10)
            corridors.append(seg)
            for p in seg:
                pts.append(p); owners.append(m["id"])
        params = {
            "latitude": ",".join(f"{p[0]:.4f}" for p in pts),
            "longitude": ",".join(f"{p[1]:.4f}" for p in pts),
            "hourly": ("cloud_cover,cloud_cover_low,cloud_cover_mid,cloud_cover_high,"
                       "precipitation_probability,relative_humidity_2m"),
            "forecast_days": 1, "timezone": "Asia/Shanghai",
        }
        raw = http_get_json("https://api.open-meteo.com/v1/forecast", params=params, timeout=30)
        rows = raw if isinstance(raw, list) else [raw]
        now0 = datetime.now(TZ).replace(minute=0, second=0, microsecond=0)
        ref = now0.strftime("%Y-%m-%dT%H:%M")  # Open-Meteo time 无时区后缀，用本地整点字符串比较
        idx = 0
        cur = {"time": None, "cloud": None, "low": None, "mid": None, "high": None}
        if rows and rows[0].get("hourly", {}).get("time"):
            times = rows[0]["hourly"]["time"]
            idx = next((i for i, t in enumerate(times) if t >= ref), 0)
            cur["time"] = times[idx] if idx < len(times) else None
        def _g(row, field):
            arr = row.get("hourly", {}).get(field) or []
            try: return float(arr[idx]) if idx < len(arr) and arr[idx] is not None else None
            except (TypeError, ValueError, IndexError): return None
        corridors_out = []
        for ci, m in enumerate(MOUNTAINS):
            seg = corridors[ci]
            points = []
            for pi, p in enumerate(seg):
                row = rows[pi] if pi < len(rows) else {}
                low_c = _g(row, "cloud_cover_low") or 0
                mid_c = _g(row, "cloud_cover_mid") or 0
                high_c = _g(row, "cloud_cover_high") or 0
                rh = _g(row, "relative_humidity_2m") or 0
                pop = _g(row, "precipitation_probability") or 0
                # v2.10.7: 走廊点云型识别——取主导层做 11 云属多因子判定
                dl = {"low": low_c, "mid": mid_c, "high": high_c}
                dom = max(dl, key=dl.get)
                genus = genus_of_pt(dom, dl[dom], rh, pop, low_c)
                points.append({"lat": round(p[0], 4), "lon": round(p[1], 4),
                               "cloud": _g(row, "cloud_cover"), "low": low_c,
                               "mid": mid_c, "high": high_c, "rh": round(rh),
                               "pop": round(pop), "genus": genus,
                               "genus_zh": GN_ZH_PY.get(genus, "")})
            corridors_out.append({"id": m["id"], "name": m["name"], "elev": m["elev"],
                                  "lat": m["lat"], "lon": m["lon"], "points": points})
        out = {"observer": {"lat": lat, "lon": lon},
               "src_bbox": SAT_SRC_BBOX, "bbox": SAT_VIEW_BBOX,
               "hour": cur["time"], "corridors": corridors_out,
               "generated": datetime.now(TZ).isoformat()}
        cache_put(key, out)
        return out
    except Exception:
        return None


def _num(v):
    try:
        return float(str(v).replace("℃", "").replace("°", "").strip())
    except (TypeError, ValueError):
        return None


def current_check(om, today_morning_score):
    """实况 vs 今日预报的一致性校验。返回 (是否一致, 结论文案)。"""
    if not om: return True, "实况数据暂不可用"
    vis = om.get("visibility"); cc = om.get("cloud_cover")
    wc = om.get("weather_code"); rain = om.get("precipitation") or 0
    foggy = wc in (45, 48) or (vis is not None and vis < 5000)
    cloudy = cc is not None and cc > 75
    wet = rain > 0.1 or wc in (51, 53, 55, 61, 63, 65, 80, 81, 82, 95, 96, 99)
    if today_morning_score is None:
        return True, "今日预报未生成"
    blocked = foggy or cloudy or wet
    if blocked and today_morning_score >= 65:
        return False, f"实况{('有雾/能见度'+str(round(vis/1000,1))+'km' if foggy else '')}{('云量'+str(cc)+'%' if cloudy else '')}{('有降水' if wet else '')}，预报晨间 {today_morning_score} 分偏乐观，出发前请再核实"
    if not blocked and today_morning_score < 45:
        return False, f"实况通透（能见度 {round(vis/1000,1) if vis else '?'}km/云量 {cc if cc is not None else '?'}%），预报 {today_morning_score} 分偏悲观，或可等待窗口"
    return True, f"实况与预报基本一致（预报晨间 {today_morning_score} 分）"


def open_meteo_archive_synoptic(observer, start_date, end_date):
    """历史时段的形势数据。ERA5 在中国区域不提供等压面（850hPa 等为 None），
    改用地面要素做近似检测（2m 变温→冷空气、晴夜静风→辐射逆温、风向急转→锋面）。"""
    try:
        params = {
            "latitude": f"{observer['lat']:.4f}", "longitude": f"{observer['lon']:.4f}",
            "hourly": ("temperature_2m,wind_speed_10m,wind_direction_10m,cloud_cover,"
                       "temperature_850hPa,temperature_925hPa,temperature_1000hPa,"
                       "wind_speed_850hPa,wind_direction_850hPa,geopotential_height_850hPa"),
            "start_date": start_date, "end_date": end_date, "timezone": "Asia/Shanghai",
        }
        raw = http_get_json("https://archive-api.open-meteo.com/v1/archive", params=params, timeout=40)
        return raw["hourly"]
    except Exception:
        return None


def synoptic_factors(syn, dt):
    """检测天气系统因子。返回 (facts字典, 加分, 理由列表)。
    等压面数据可用→完整检测；仅地面数据→近似检测（历史回顾）。"""
    if not syn:
        return {"available": False}, 0.0, []
    times = syn.get("time") or []
    if not times:
        return {"available": False}, 0.0, []
    idx = min(range(len(times)), key=lambda i: abs(datetime.fromisoformat(times[i]).replace(tzinfo=TZ).timestamp()-dt.timestamp()))
    def at(field, back=0):
        arr = syn.get(field) or []
        i = idx - back
        if not (0 <= i < len(arr)): return None
        v = arr[i]
        try: return None if v is None else float(v)
        except (TypeError, ValueError): return None
    # 等压面可用 → 完整检测
    if any(x is not None for x in (syn.get("temperature_850hPa") or [])[:6]):
        return _synoptic_full(at, syn, dt)
    # 仅地面 → 近似检测
    return _synoptic_approx(at, dt)


def _synoptic_full(at, syn, dt):
    t2m, t850, t925, t1000 = at("temperature_2m"), at("temperature_850hPa"), at("temperature_925hPa"), at("temperature_1000hPa")
    ws850, wd850 = at("wind_speed_850hPa"), at("wind_direction_850hPa")
    ws10, wd10 = at("wind_speed_10m"), at("wind_direction_10m")
    gph, gph24 = at("geopotential_height_850hPa"), at("geopotential_height_850hPa", 24)
    t850_24 = at("temperature_850hPa", 24)
    facts = {"available": True}; bonus = 0.0; notes = []
    # ---- 逆温层：T1000−T850 正常 6~9°C；≤4 即有逆温/等温倾向（霾层滞留的物理机制） ----
    if t1000 is not None and t850 is not None:
        diff = t1000 - t850
        facts["inversion_strength"] = round(diff, 1)
        if diff <= 1:
            bonus -= 8; facts["inversion"] = "强"; notes.append("强逆温层：低层污染物难扩散")
        elif diff <= 4:
            bonus -= 4; facts["inversion"] = "弱"; notes.append("逆温/等温层：霾易滞留")
        else:
            facts["inversion"] = "无"
    # ---- 冷空气：24h 850hPa 变温 + 季节温度水平（槽后冷平流） ----
    if t850 is not None and t850_24 is not None:
        dT = t850 - t850_24
        facts["dt850_24h"] = round(dT, 1)
        cold_base = 14 if 5 <= dt.month <= 9 else 4
        if dT <= -5 or t850 <= cold_base - 6:
            bonus += 3; facts["cold_air"] = "强"; notes.append("强冷空气活动")
        elif dT <= -3 or t850 <= cold_base - 2:
            bonus += 1.5; facts["cold_air"] = "弱"; notes.append("弱冷空气渗透")
    # ---- 锋面/切变线：850hPa 与地面风向切变 + 风速差 ----
    if wd850 is not None and wd10 is not None and ws850 is not None and ws10 is not None:
        shear = abs(((wd850-wd10+540) % 360)-180)
        speed_shear = ws850 - ws10
        facts["wind_shear"] = round(shear)
        if shear >= 100 and speed_shear >= 6:
            bonus -= 5; facts["front"] = "锋面/切变线"; notes.append("锋面/切变线附近，云雨风险")
        elif shear >= 80 or (180 <= wd850 <= 260 and ws850 >= 12):
            bonus -= 2.5; facts["front"] = "切变区"; notes.append("切变区，天气不稳定")
    # ---- 槽脊：850hPa 位势高度变化 + 风向扇区 ----
    if gph is not None:
        anom = gph - (gph24 if gph24 is not None else gph)
        facts["gph850_anom"] = round(anom, 0)
    if wd850 is not None:
        westerly = 180 <= wd850 <= 270      # 槽前西南-西风（暖湿上升，云雨多）
        northerly = wd850 >= 270 or wd850 <= 60  # 槽后/脊前西北-北风（干爽通透）
        anom = facts.get("gph850_anom", 0)
        if northerly and anom >= -5:
            bonus += 2; facts["system"] = "脊前西北气流"; notes.append("脊前西北气流，晴好稳定")
        elif westerly and anom <= 8:
            bonus -= 2.5; facts["system"] = "槽前西南气流"; notes.append("槽前暖湿气流，云量偏多")
        elif northerly:
            bonus += 1; facts["system"] = "偏北气流"
        else:
            facts["system"] = "偏南气流"
    return facts, round(bonus, 1), notes


def _synoptic_approx(at, dt):
    """历史回顾的地面近似检测：2m 变温→冷空气、晴夜静风→辐射逆温、风向急转→锋面。"""
    t2m = at("temperature_2m"); t24 = at("temperature_2m", 24)
    ws = at("wind_speed_10m"); wd = at("wind_direction_10m"); wd24 = at("wind_direction_10m", 24)
    cc = at("cloud_cover")
    facts = {"available": True}; bonus = 0.0; notes = []
    # 冷空气：2m 温度 24h 降温明显
    if t2m is not None and t24 is not None:
        dT = t2m - t24
        facts["dt2m_24h"] = round(dT, 1)
        if dT <= -6:
            bonus += 2; facts["cold_air"] = "强"; notes.append("冷空气活动（2m 24h降温）")
        elif dT <= -4:
            bonus += 1; facts["cold_air"] = "弱"; notes.append("弱冷空气渗透")
    # 锋面：24h 风向急转 ≥120°
    if wd is not None and wd24 is not None:
        turn = abs(((wd-wd24+540) % 360)-180)
        facts["wind_turn_24h"] = round(turn)
        if turn >= 120:
            bonus -= 2; facts["front"] = "风向急转(近似锋面)"; notes.append("风向急转，近似锋面活动")
    # 辐射逆温：夜间 + 少云 + 微风（近似，标注）
    if dt.hour >= 20 or dt.hour <= 8:
        night_clear = (cc is not None and cc < 25)
        calm = (ws is not None and ws < 2.0)
        if night_clear and calm:
            bonus -= 3; facts["inversion"] = "辐射逆温(近似)"; notes.append("晴夜静风，辐射逆温风险")
        elif night_clear or calm:
            facts["inversion"] = "倾向"
    # 风向扇区（近似槽脊）
    if wd is not None:
        if wd >= 270 or wd <= 60:
            bonus += 1; facts["system"] = "偏北气流"
        elif 180 <= wd <= 270:
            bonus -= 1.5; facts["system"] = "偏南气流"
        else:
            facts["system"] = "偏东气流"
    return facts, round(bonus, 1), notes


def aggregate_daily(hours):
    """把逐小时评分聚合为每日（晨 5-10 点 / 晚 16-20 点 / 全天最佳 + 金山最佳）。"""
    daily = []
    dates = sorted(set(x["time"][:10] for x in hours))
    for d in dates:
        day = [x for x in hours if x["time"].startswith(d)]
        morning = [x for x in day if 5 <= datetime.fromisoformat(x["time"]).hour <= 10]
        evening = [x for x in day if 16 <= datetime.fromisoformat(x["time"]).hour <= 20]
        best = max(day, key=lambda x: x["score"])
        daily.append({"date": d,
                      "morning": max(morning, key=lambda x: x["score"]) if morning else best,
                      "evening": max(evening, key=lambda x: x["score"]) if evening else best,
                      "gold": max(day, key=lambda x: x["gold"])})
    return daily


def build_history(observer, start_date, end_date):
    """任意日期范围的每日评分回顾。AOD 超过 92 天时缺失，评分自动降级（仍可用）。"""
    s = datetime.strptime(start_date, "%Y-%m-%d").date()
    e = datetime.strptime(end_date, "%Y-%m-%d").date()
    # 多拉一天作洗尘/霾层判断的前置缓冲
    buf_start = (s - timedelta(days=1)).isoformat()
    meteo = open_meteo_archive(observer, MOUNTAINS, buf_start, end_date)
    aerosol = open_meteo_aerosol_history(observer, MOUNTAINS, start_date, end_date)
    syn = open_meteo_archive_synoptic(observer, buf_start, end_date)   # v2.8: 历史等压面形势
    aod_available = aerosol is not None
    result = []
    for m in MOUNTAINS:
        dist, bearing = haversine_bearing(observer["lat"], observer["lon"], m["lat"], m["lon"])
        terrain = path_terrain(observer, m, len(meteo[m["id"]]))
        times = [datetime.fromisoformat(t).replace(tzinfo=TZ) for t in meteo[m["id"]][0]["time"]]
        hours = []
        for dt in times:
            air = air_at_time(aerosol[m["id"]], dt) if aerosol else EMPTY_AIR
            hours.append(score_hour(dt, m, meteo[m["id"]], observer, air, terrain, syn))
        daily = [d for d in aggregate_daily(hours) if d["date"] >= start_date]
        result.append({**m, "distance": round(dist, 1), "bearing": round(bearing, 1),
                       "peak_angle": round(apparent_peak_angle(dist, m["elev"], observer["elev"]), 2),
                       "daily": daily})
    return {"observer": observer, "mountains": result,
            "aod_available": aod_available,
            "span": {"start": start_date, "end": end_date},
            "generated": datetime.now(TZ).isoformat()}


# ---- v2.7 历史导出 Excel（历史数据 vs 预报数据 评分差异趋势） ----

XLSX_HEADER_FILL = PatternFill("solid", fgColor="2D6A8F")
XLSX_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
XLSX_BODY_FONT = Font(name="Arial", size=11)
XLSX_BOLD = Font(name="Arial", size=11, bold=True)
XLSX_KPI_FILL = PatternFill("solid", fgColor="EAF2FF")
XLSX_ZEBRA = PatternFill("solid", fgColor="F7F9FC")
XLSX_BORDER = Border(
    left=Side(style="thin", color="D9DEE7"), right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"), bottom=Side(style="thin", color="D9DEE7"))
XLSX_TOP = Border(
    left=Side(style="thin", color="D9DEE7"), right=Side(style="thin", color="D9DEE7"),
    top=Side(style="medium", color="AAB4C5"), bottom=Side(style="thin", color="D9DEE7"))


def _xlsx_header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, c, h)
        cell.fill = XLSX_HEADER_FILL; cell.font = XLSX_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.freeze_panes = "A2"


def export_history_xlsx(observer, start_date, end_date):
    """导出 xlsx：Sheet1 每日评分（历史范围）；Sheet2 历史 vs 预报趋势对比（均值/差值用公式）。"""
    hist = build_history(observer, start_date, end_date)
    # 预报段：今天起 5 天（30 分钟缓存，通常命中）
    try:
        meteo, _ = open_meteo_corridor(observer, MOUNTAINS, model="best_match")
        aerosol, _ = open_meteo_aerosol(observer, MOUNTAINS)
    except Exception:
        meteo = aerosol = None
    fc_daily = {}
    if meteo:
        today = datetime.now(TZ).date().isoformat()
        syn = open_meteo_synoptic(observer, model="best_match")
        for m in MOUNTAINS:
            try:
                terrain = path_terrain(observer, m, len(meteo[m["id"]]))
                times = [datetime.fromisoformat(t).replace(tzinfo=TZ) for t in meteo[m["id"]][0]["time"]]
                hours = [score_hour(dt, m, meteo[m["id"]], observer, air_at_time(aerosol[m["id"]], dt) if aerosol else EMPTY_AIR, terrain, syn) for dt in times]
                fc_daily[m["id"]] = [d for d in aggregate_daily(hours) if d["date"] >= today]
            except Exception:
                fc_daily[m["id"]] = []

    wb = Workbook()
    # ---- Sheet1: 每日评分 ----
    ws1 = wb.active; ws1.title = "每日评分"
    cols1 = ["日期"]
    for m in hist["mountains"]:
        cols1 += [f"{m['name']} 晨分", f"{m['name']} 晚分"]
    _xlsx_header(ws1, cols1, [12] + [13] * (len(cols1) - 1))
    r = 2
    for d in hist["mountains"][0]["daily"]:
        ws1.cell(r, 1, d["date"]).font = XLSX_BODY_FONT
        for i, m in enumerate(hist["mountains"]):
            day = next((x for x in m["daily"] if x["date"] == d["date"]), None)
            if day:
                c1 = ws1.cell(r, 2 + i * 2, day["morning"]["score"])
                c2 = ws1.cell(r, 3 + i * 2, day["evening"]["score"])
                for c in (c1, c2):
                    c.font = XLSX_BODY_FONT; c.alignment = Alignment(horizontal="center")
        r += 1
    last = r - 1
    for row in ws1.iter_rows(min_row=2, max_row=last, min_col=1, max_col=len(cols1)):
        for cell in row:
            cell.border = XLSX_BORDER
            if (cell.row - 2) % 2: cell.fill = XLSX_ZEBRA
    # 末行加 KPI 均值（公式）
    r += 1
    ws1.cell(r, 1, "历史均值").font = XLSX_BOLD
    for i, m in enumerate(hist["mountains"]):
        a = ws1.cell(r, 2 + i * 2); b = ws1.cell(r, 3 + i * 2)
        ca = chr(66 + i * 2)   # 晨分列（B,D,F,H）
        cb = chr(67 + i * 2)   # 晚分列（C,E,G,I）
        a.value = f"=AVERAGE({ca}2:{ca}{last})"
        b.value = f"=AVERAGE({cb}2:{cb}{last})"
        for c in (a, b):
            c.font = XLSX_BOLD; c.fill = XLSX_KPI_FILL; c.alignment = Alignment(horizontal="center")
    ws1.cell(r, 1).border = XLSX_TOP
    for c in range(2, len(cols1) + 1):
        ws1.cell(r, c).border = XLSX_TOP

    # ---- Sheet2: 历史 vs 预报 趋势对比 ----
    ws2 = wb.create_sheet("历史vs预报")
    _xlsx_header(ws2, ["山峰", "历史天数", "历史晨均值", "历史晚均值", "预报天数", "预报晨均值", "预报晚均值", "晨分差值", "晚分差值"], [16, 10, 12, 12, 10, 12, 12, 12, 12])
    r = 2
    for m in hist["mountains"]:
        days = m["daily"]
        fc = fc_daily.get(m["id"], [])
        n_hist = len(days); n_fc = len(fc)
        row0 = r
        ws2.cell(r, 1, m["name"]).font = XLSX_BOLD
        ws2.cell(r, 2, n_hist).font = XLSX_BODY_FONT
        # 均值用公式引用 Sheet1 对应列
        col_am = 2 + hist["mountains"].index(m) * 2   # 晨分列
        col_ev = col_am + 1                            # 晚分列
        ws2.cell(r, 3).value = f"=IF({chr(64+col_am)}{2}=\"\",\"\",ROUND(AVERAGE('每日评分'!{chr(64+col_am)}2:{chr(64+col_am)}{1+n_hist}),1))"
        ws2.cell(r, 4).value = f"=IF(ISERROR(AVERAGE('每日评分'!{chr(64+col_ev)}2:{chr(64+col_ev)}{1+n_hist})),\"\",ROUND(AVERAGE('每日评分'!{chr(64+col_ev)}2:{chr(64+col_ev)}{1+n_hist}),1))"
        ws2.cell(r, 5, n_fc).font = XLSX_BODY_FONT
        if fc and n_fc:
            fc_morn = sum(x["morning"]["score"] for x in fc) / n_fc
            fc_eve = sum(x["evening"]["score"] for x in fc) / n_fc
            ws2.cell(r, 6, round(fc_morn, 1)); ws2.cell(r, 7, round(fc_eve, 1))
        else:
            ws2.cell(r, 6, "—"); ws2.cell(r, 7, "—")
        # 差值：预报均值 - 历史均值（公式，引用本行）
        ws2.cell(r, 8).value = f"=IF(OR(ISNUMBER(F{r}),ISNUMBER(C{r})),ROUND(F{r}-C{r},1),\"—\")" if fc and n_fc else "—"
        ws2.cell(r, 9).value = f"=IF(OR(ISNUMBER(G{r}),ISNUMBER(D{r})),ROUND(G{r}-D{r},1),\"—\")" if fc and n_fc else "—"
        for c in range(1, 10):
            cell = ws2.cell(r, c)
            cell.border = XLSX_BORDER
            if cell.font is None or not cell.font.bold: cell.font = XLSX_BODY_FONT
            if (r - 2) % 2: cell.fill = XLSX_ZEBRA
        # KPI 高亮差值列
        for c in (8, 9):
            ws2.cell(r, c).fill = XLSX_KPI_FILL
        r += 1
    # 汇总行：各列均值（公式）
    ws2.cell(r, 1, "全部山峰均值").font = XLSX_BOLD
    for c in range(3, 10):
        col = chr(64 + c)
        cell = ws2.cell(r, c)
        cell.value = f"=ROUND(AVERAGE({col}2:{col}{r-1}),1)"
        cell.font = XLSX_BOLD; cell.fill = XLSX_KPI_FILL
        cell.border = XLSX_TOP
    ws2.cell(r, 2).border = XLSX_TOP; ws2.cell(r, 1).border = XLSX_TOP

    for ws in (ws1, ws2):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.alignment is None or not cell.alignment.horizontal:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def air_at_time(corridor, local_dt):
    """返回视线最不利 AOD，以及用于解释的 PM2.5/沙尘峰值。"""
    if not corridor: return {"aod": None, "pm2_5": None, "dust": None, "points": []}
    times = corridor[0].get("time", [])
    if not times: return {"aod": None, "pm2_5": None, "dust": None, "points": []}
    idx = min(range(len(times)), key=lambda i: abs(datetime.fromisoformat(times[i]).replace(tzinfo=TZ).timestamp()-local_dt.timestamp()))
    def maximum(field):
        values=[]
        for point in corridor:
            arr=point.get(field, [])
            if idx < len(arr) and arr[idx] is not None:
                try: values.append(float(arr[idx]))
                except (TypeError, ValueError): pass
        return max(values) if values else None
    points=[]
    for point in corridor:
        item={}
        for field in ("aerosol_optical_depth","pm2_5","dust"):
            arr=point.get(field,[]); value=arr[idx] if idx<len(arr) else None
            try: item[field]=None if value is None else float(value)
            except (TypeError,ValueError): item[field]=None
        points.append(item)
    return {"aod": maximum("aerosol_optical_depth"), "pm2_5": maximum("pm2_5"), "dust": maximum("dust"), "points": points}


# ---- v2.0 经验知识层（数值预报 × 观山社群 10 年经验） ---------------------

def washout_score(rain12, wind_avg):
    """前12h降水+风的洗尘得分 0~12。
    经验：'小雨无效、要下透'；大雨+大风是'次日清晨出山'第一定律。"""
    if rain12 >= 10: base = 1.0                 # 大雨/暴雨
    elif rain12 >= 3: base = 0.55 + 0.45*(rain12-3)/7   # 中到大雨
    elif rain12 >= 1: base = 0.20 + 0.35*(rain12-1)/2   # 小到中雨（效果有限）
    else: base = 0.0                             # 小雨/无雨：对霾层基本无效
    wind_factor = 1.0 if wind_avg >= 12 else (0.6 if wind_avg >= 6 else 0.35)
    return round(12.0*base*wind_factor, 1)


def haze_streak_hours(precip_arr, idx):
    """距上次有效降水（单小时≥1mm，近似中雨）的小时数/24（天）。
    数据不可用（有效值不足 1 天）时返回 None。"""
    if not precip_arr or idx <= 0:
        return None
    useful = sum(1 for i, v in enumerate(precip_arr[:idx]) if v is not None)
    if useful < 24:          # 历史降水数据缺失（部分模型 past_days 无降水变量）
        return None
    for i in range(idx-1, -1, -1):
        v = precip_arr[i]
        if v is not None:
            try:
                if float(v) >= 1.0:
                    return (idx - i) / 24.0
            except (TypeError, ValueError):
                pass
    return idx / 24.0        # 整个回看窗口无雨 → 取窗口长度


def haze_penalty(streak_days):
    """霾层积累惩罚 0~12。连续无有效降雨越久，盆地静稳霾层越厚，
    即使预报晴天也可能'看着蓝，看不见山'（2021-04-28 失败案例）。"""
    if streak_days is None: return 0.0
    if streak_days >= 7:  return 12.0
    if streak_days >= 5:  return 8.0 + 2.0*(streak_days-5)
    if streak_days >= 3:  return 3.0 + 2.5*(streak_days-3)
    if streak_days >= 1.5: return 1.0
    return 0.0


def season_bonus(month):
    """观山季月度加权 -2~+4。依据四川省气象局月均可见天数统计。"""
    f = MONTH_FREQ.get(month, 2.5)
    if f >= 9: return 4
    if f >= 7: return 3
    if f >= 4: return 1
    if f >= 2: return 0
    return -2


def window_bonus(sun_elev, hour):
    """日出黄金窗口 0~5 分：以太阳高度角为锚（经验最佳 5:40-7:00），
    傍晚日落前少量加分（夏季傍晚偶尔可见雪山晚霞）。清晨与傍晚互斥。"""
    morning = 0.0
    if hour <= 12 and -6 <= sun_elev <= 12:
        if sun_elev <= 4:
            k = 0.30 + 0.70*(sun_elev+6)/10.0
        else:
            k = 1.00 - 0.60*(sun_elev-4)/8.0
        morning = 5.0*max(0.0, min(1.0, k))
    evening = 0.0
    if hour > 12 and 0 <= sun_elev <= 8:
        evening = 2.0*(1.0 - sun_elev/8.0)
    return round(morning + evening, 1)


def consecutive_clear(corridor, idx, dist):
    """连续晴日因子 0~2.5（超远山×1.6）。贡嘎等 200km+ 远山需连续
    多日晴朗少云、高空通透（全年可见率仅约 5%）。"""
    p = corridor[0]
    prec = p.get("precipitation") or []
    low = p.get("cloud_cover_low") or []
    lo, hi = max(0, idx-24), min(len(prec), idx+24)
    if hi <= lo:
        return 0.0
    total = 0.0
    for v in prec[lo:hi]:
        if v is not None:
            try: total += float(v)
            except (TypeError, ValueError): pass
    seg = low[lo:hi]
    max_low = max((float(v) for v in seg if v is not None), default=99)
    if total < 0.5 and max_low <= 45:
        return round(2.5*(1.6 if dist > 200 else 1.0), 1)
    if total < 2.0 and max_low <= 60:
        return 1.0
    return 0.0


def radiation_fog_risk(dt, corridor, idx, rh3, wind3, cloud3, syn_facts):
    """辐射雾风险评估（v2.8）。形成条件：晴夜（少云利于地表辐射冷却）+ 微风（近地面
    降温但不破坏层结）+ 近地层高湿（露点接近气温）+ 逆温层；日出前后（5-10 时）最浓。
    返回 (等级, 扣分, 理由)；扣分计入总分。依据：中央气象台雾预报规范、四川盆地气象博主
    "晴夜静风高湿必防辐射雾"经验。"""
    if not (4 <= dt.hour <= 10):
        return None, 0.0, ""
    # 晴夜条件：前夜（0-8 时）云量少，利于辐射冷却
    cc = [x for x in (cloud3 or []) if x is not None]
    clear_night = (not cc) or (sum(cc)/len(cc) <= 35)
    calm = (not wind3) or (max(wind3) <= 3.0)
    humid = (not rh3) or (max(rh3) >= 88)
    inversion = bool(syn_facts and syn_facts.get("inversion") and syn_facts.get("inversion") != "无")
    # 多日未雨有利于雾形成（水汽在地面附近，无降水扰动）
    dry = True
    p0 = corridor[0]
    prec = p0.get("precipitation") or []
    lo = max(0, idx-48)
    dry = sum(float(v or 0) for v in prec[lo:idx+1]) < 1.0
    score = (1 if clear_night else 0) + (1 if calm else 0) + (1 if humid else 0) + (1 if inversion else 0) + (1 if dry else 0)
    if score >= 4:
        level = "高"
        penalty = 9.0 if (humid and clear_night and calm) else 6.0
    elif score >= 3:
        level = "中"
        penalty = 4.0
    elif score >= 2:
        level = "低"
        penalty = 1.5
    else:
        return None, 0.0, ""
    note = f"辐射雾风险{level}（{('晴夜' if clear_night else '多云')}·{('静风' if calm else '有风')}·{('高湿' if humid else '湿度一般')}{'·逆温' if inversion else ''}）"
    return level, penalty, note


def estimate_visibility(rh, precip, aod=None, cloud_low=None):
    """v2.10.8: ERA5 再分析无能见度变量（Archive API 返回 None，导致历史回算能见度恒为默认 10km）。
    改用可得的物理量经验估算：降水强度 > 相对湿度（雾/轻雾/霾分级）> 气溶胶（Koschmieder 近似）> 低云罩顶。
    返回米，范围 200-30000。"""
    try: rh = float(rh)
    except (TypeError, ValueError): rh = 70.0
    try: precip = float(precip or 0)
    except (TypeError, ValueError): precip = 0.0
    # 降水强度直接决定能见度上限（雨强-能见度经验关系）
    if precip >= 4.0: return 1500.0
    if precip >= 2.0: return 2500.0
    if precip >= 1.0: return 4000.0
    if precip >= 0.4: return 5500.0
    if precip >= 0.1: return 8000.0
    # 湿度主导（雾/轻雾/霾 分级，参考能见度观测规范）
    if rh >= 99: return 300.0
    if rh >= 97: return 700.0
    if rh >= 95: return 1400.0
    if rh >= 93: return 2500.0
    if rh >= 90: return 4500.0
    if rh >= 87: return 7000.0
    if rh >= 84: return 9500.0
    if rh >= 78: return 13000.0
    base = 22000.0 if rh >= 60 else 28000.0
    # 气溶胶衰减（Koschmieder：V = 3.912 / b_ext，消光系数 ∝ AOD）
    if aod is not None:
        try:
            aod = float(aod)
            if aod > 0.05:
                base = min(base, 18000.0 * math.exp(-2.2 * aod))
        except (TypeError, ValueError):
            pass
    # 低云罩顶：低云量大且近饱和时视野受限
    try:
        if float(cloud_low or 0) >= 85 and rh >= 85:
            base = min(base, 3000.0)
    except (TypeError, ValueError):
        pass
    return max(200.0, min(30000.0, base))


def score_hour(dt, mountain, corridor, observer, air, terrain, syn=None):
    idx = min(range(len(corridor[0]["time"])), key=lambda i: abs(datetime.fromisoformat(corridor[0]["time"][i]).replace(tzinfo=TZ).timestamp()-dt.timestamp()))
    def vals(field, default=0):
        out=[]
        for p in corridor:
            arr=p.get(field) or []
            value=arr[idx] if idx < len(arr) and arr[idx] is not None else default
            try: out.append(float(value))
            except (TypeError,ValueError): out.append(float(default))
        return out
    low, mid, high = vals("cloud_cover_low"), vals("cloud_cover_mid"), vals("cloud_cover_high")
    rh = vals("relative_humidity_2m", 70); pop = vals("precipitation_probability")
    # v2.10.8: ERA5 再分析无能见度（Archive API 返回 None，此前恒为默认 10km）→
    # 逐点用湿度/降水/气溶胶/低云物理估算；预报路径（模型真实值）不受影响。
    air_points = air.get("points", [])
    vis = []
    for i in range(len(corridor)):
        arr = corridor[i].get("visibility") or []
        raw = arr[idx] if idx < len(arr) and arr[idx] is not None else None
        if raw is not None:
            try: vis.append(float(raw)); continue
            except (TypeError, ValueError): pass
        pre = corridor[i].get("precipitation") or []
        pv = pre[idx] if idx < len(pre) and pre[idx] is not None else 0.0
        ap = air_points[i] if i < len(air_points) else {}
        vis.append(estimate_visibility(rh[i], pv, ap.get("aerosol_optical_depth"), low[i]))
    wind = vals("wind_speed_10m")
    # v2.0: 前12h降水只看近观测端 3 个点（盆地内），避免走廊全线求和夸大。
    rain12 = 0.0
    for p in corridor[:3]:
        arr = p.get("precipitation") or []
        rain12 = max(rain12, sum(float(v or 0) for v in arr[max(0, idx-12):idx]))
    wind_avg = sum(wind[:3])/max(1, len(wind[:3]))
    # ---- 数值物理层（实时云量/能见度/湿度/气溶胶是主约束） ----
    cloud_pen = 0.34*max(low[:3])/100 + 0.28*max(low[3:])/100 + 0.17*max(mid[3:])/100 + 0.06*sum(high)/len(high)/100
    # v2.0: 云底过低惩罚——高湿+低云量大时，云底很可能低于 500m 直接罩住峰线（彭县数据：全年大多数云底<500m）
    if max(rh[:3]) > 85 and max(low[:3]) > 60:
        cloud_pen = min(1.0, cloud_pen + 0.10)
    vis_factor = clamp((min(vis[:3])-5000)/45000)
    humidity_pen = clamp((max(rh[:3])-65)/35)*0.14
    rain_pen = max(pop)/100*0.12
    aod = air.get("aod")
    aerosol_factor = None if aod is None else math.exp(-2.15*max(0,aod))
    aerosol_pen = 0.18 if aerosol_factor is None else (1-aerosol_factor)*0.36
    # v2.0: 前夜大雨+大风洗尘时，CAMS 对湿沉降模拟常滞后（AOD 虚高），经验性打折
    if rain12 >= 3 and wind_avg >= 8:
        aerosol_pen *= 0.6
    base = clamp(1-cloud_pen-humidity_pen-rain_pen-aerosol_pen)
    # 超远山对能见度和气溶胶更敏感。
    dist, bearing = haversine_bearing(observer["lat"],observer["lon"],mountain["lat"],mountain["lon"])
    range_factor = clamp(vis_factor + 0.35, 0.18, 1) ** (dist/180)
    # ---- v2.0 经验知识层（观山社群 10 年经验，非空想） ----
    washout = washout_score(rain12, wind_avg)                 # +0~12 前夜大雨洗尘
    streak = haze_streak_hours(corridor[0].get("precipitation") or [], idx)
    haze = haze_penalty(streak)                               # -0~12 久未降雨霾层积累
    season = season_bonus(dt.month)                           # -2~+4 观山季月度加权
    elev, sunaz = solar_position(dt, observer["lat"], observer["lon"])
    window = window_bonus(elev, dt.hour)                      # +0~5 日出黄金窗口
    clear = consecutive_clear(corridor, idx, dist)            # +0~2.5(远山×1.6) 连续晴日
    wind_b = 3.0 if 8 <= wind_avg <= 28 else 0
    empirical_bonus = washout - haze + season + window + clear + wind_b
    # v2.8: 天气系统因子（冷空气/锋面切变/槽脊/逆温层）独立计入总分
    syn_facts, syn_bonus, syn_notes = synoptic_factors(syn, dt)
    # v2.8: 辐射雾风险（晴夜+静风+高湿+逆温 → 清晨能见度骤降，直接扣分）
    fog_level, fog_penalty, fog_note = radiation_fog_risk(dt, corridor, idx, rh[:3], wind[:3], vals("cloud_cover"), syn_facts)
    score = round(clamp(100*base*range_factor+empirical_bonus+syn_bonus-fog_penalty,0,100))
    gold = 0
    if -1.5 <= elev <= 6:
        # 太阳位于山峰相反方向附近时，山体正面受光（容差随散射放宽）。
        sep = abs((sunaz-bearing+180)%360-180)
        front = clamp((sep-70)/70)
        gold = round(score*front*clamp((elev+1.5)/3.5 if elev<2 else (6-elev)/4+0.35))
    reasons=[]
    if max(low)>65: reasons.append("通道低云偏多")
    if max(mid[3:])>65: reasons.append("峰区中云遮挡")
    if min(vis[:3])<15000: reasons.append("盆地能见度不足")
    if aod is not None and aod>0.4: reasons.append("AOD550 很高")
    elif aod is not None and aod>0.2: reasons.append("气溶胶偏多")
    if max(rh[:3])>88: reasons.append("近地层湿度高")
    if haze >= 5: reasons.append("久未有效降雨，警惕霾层积累")
    if washout >= 8: reasons.append("前夜大雨+大风洗尘，出山概率高")
    elif washout >= 3: reasons.append("前夜有降水，空气较通透")
    if season >= 3: reasons.append("正值观山旺季")
    if window >= 4: reasons.append("日出黄金窗口")
    if clear > 0 and dist > 200: reasons.append("连续晴日，远山可期")
    if wind_b: reasons.append("近地风有利于盆地扩散")
    if syn_notes: reasons.extend(syn_notes)
    if fog_note: reasons.append(fog_note)
    profile=[]
    air_points=air.get("points",[])
    def _pv(i, field, default=0.0):
        arr = corridor[i].get(field) or []
        v = arr[idx] if idx < len(arr) and arr[idx] is not None else default
        try: return float(v)
        except (TypeError, ValueError): return float(default)
    for i in range(len(corridor)):
        ap=air_points[i] if i<len(air_points) else {}
        profile.append({"distance":round(dist*i/max(1,len(corridor)-1),1),"terrain":round(float(terrain[i]),1),"low":round(low[i],1),"mid":round(mid[i],1),"high":round(high[i],1),"rh":round(rh[i],1),"visibility":round(vis[i]/1000,1),"aod":ap.get("aerosol_optical_depth"),"precip":round(_pv(i,"precipitation"),2),"wind":round(_pv(i,"wind_speed_10m"),1),"temp":round(_pv(i,"temperature_2m",15),1)})
    # v2.1: 视线遮挡点检测——与"天空形态模拟"同一套判定（云量>=45% 且视线高度落在云层高度带内），
    # 输出经纬度供地图红色标注。云层高度带：低云 0.2-3km / 中云 3-7km / 高云 7-12km AGL。
    blocked=[]
    pts=interpolate_great_circle(observer["lat"],observer["lon"],mountain["lat"],mountain["lon"],len(corridor))
    los0=float(observer["elev"]); los1=float(mountain["elev"])
    for i,p in enumerate(profile):
        los=los0+(los1-los0)*(p["distance"]/max(1,dist))
        for key,lo,hi in (("low",200,3000),("mid",3000,7000),("high",7000,12000)):
            cover=p[key] or 0
            base=p["terrain"]+lo; top=p["terrain"]+hi
            if cover>=45 and los>=base and los<=top:
                blocked.append({"lat":round(pts[i][0],5),"lon":round(pts[i][1],5),"distance":p["distance"],"band":key,"cover":round(cover,1)})
                break
    empirical = {"washout": washout, "haze": haze, "season": season, "window": window,
                 "clear": clear, "wind": wind_b, "notes": reasons}
    return {"time":dt.isoformat(),"score":score,"gold":gold,"empirical":empirical,"empirical_bonus":round(empirical_bonus,1),"synoptic":syn_facts,"synoptic_bonus":round(syn_bonus,1),"fog":None if fog_level is None else {"level":fog_level,"penalty":fog_penalty},"aod":None if aod is None else round(aod,3),"pm2_5":None if air.get("pm2_5") is None else round(air["pm2_5"],1),"dust":None if air.get("dust") is None else round(air["dust"],1),"low":round(max(low),1),"mid":round(max(mid[3:]),1),"high":round(sum(high)/len(high),1),"visibility":round(min(vis[:3])/1000,1),"rh":round(max(rh[:3]),1),"sun_elev":round(elev,1),"profile":profile,"blocked":blocked,"reasons":reasons or ["云量与通透度较好"]}


def build_forecast(observer, model="best_match"):
    if model not in WEATHER_MODELS: model="best_match"
    meteo, meteo_stale = open_meteo_corridor(observer, MOUNTAINS, model=model)
    aerosol, aero_stale = open_meteo_aerosol(observer, MOUNTAINS)
    syn = open_meteo_synoptic(observer, model=model)   # v2.8: 等压面形势（冷空气/锋面/槽脊/逆温）
    warnings=[]
    # v2.3: 数据源限流/故障时回退到缓存，页面给出提示
    if meteo_stale: warnings.append("天气数据源繁忙，本次为缓存预报（可能滞后）")
    if aero_stale: warnings.append("气溶胶数据源繁忙，本次为缓存数据（可能滞后）")
    result=[]
    for m in MOUNTAINS:
        dist,bearing=haversine_bearing(observer["lat"],observer["lon"],m["lat"],m["lon"])
        terrain=path_terrain(observer,m,len(meteo[m["id"]]))
        hours=[]
        # 每小时评价，页面展示每天清晨/傍晚最佳值。
        times=[datetime.fromisoformat(t).replace(tzinfo=TZ) for t in meteo[m["id"]][0]["time"]]
        for dt in times:
            air=air_at_time(aerosol[m["id"]],dt)
            hours.append(score_hour(dt,m,meteo[m["id"]],observer,air,terrain,syn))
        # v2.6: 聚合逻辑统一走 aggregate_daily；今天及以后进预报卡片
        today = datetime.now(TZ).date().isoformat()
        all_days = aggregate_daily(hours)
        daily = [d for d in all_days if d["date"] >= today]
        # v2.10.5: 历史保留 profile（含视线走廊逐点云量/地形/温度等），供「早晚剖面图」渲染。
        # 此前 slim() 剥离 profile 导致历史卡片无法查看剖面。
        history = [{"date": d["date"], "morning": d["morning"], "evening": d["evening"]} for d in all_days if d["date"] < today]
        # v2.10.21: 峰顶地形遮挡判断——被中间山体挡住则该山不参与预报展示，前端给出提示
        occ_blocked, occ_info = terrain_occlusion(observer, m)
        item = {**m, "distance": round(dist, 1), "bearing": round(bearing, 1),
                "peak_angle": round(apparent_peak_angle(dist, m["elev"], observer["elev"]), 2),
                "daily": daily, "history": history,
                "terrain_blocked": occ_blocked}
        if occ_blocked:
            item["terrain_note"] = (f"视线被 {occ_info['distance']:.0f}km 处海拔 {occ_info['terrain']:.0f}m 的地形"
                                    f"阻挡（该处视线高度仅 {occ_info['los']:.0f}m），峰顶被山体遮挡，当前观测点看不到该峰")
        result.append(item)
    return {"observer":observer,"mountains":result,"weather_model":{"id":model,**WEATHER_MODELS[model]},"aerosol":{"status":"ready","message":"Open-Meteo CAMS Global · AOD550/PM2.5/沙尘 · 自动缓存1小时"},"warnings":warnings,"generated":datetime.now(TZ).isoformat(),"method":"所选天气模型 + Open-Meteo AOD550 + 通道云量/湿度/能见度 + v2.0 经验知识层（前夜大雨洗尘/霾层积累/观山季/日出窗口/连续晴日） + v2.8 天气系统层（850hPa 冷空气/锋面切变/槽脊/逆温层）"}


@APP.get("/")
def index(): return render_template_string(HTML, observer=DEFAULT_OBSERVER, viewpoints=VIEWPOINTS, success_cases=SUCCESS_CASES, rules=EMPIRICAL_RULES)


@APP.get("/api/forecast")
def api_forecast():
    try:
        lat=float(request.args.get("lat",DEFAULT_OBSERVER["lat"])); lon=float(request.args.get("lon",DEFAULT_OBSERVER["lon"]))
        model=request.args.get("model","best_match")
        if model not in WEATHER_MODELS: raise ValueError("不支持的天气数据源")
        # v2.2: 坐标合法性校验 + 放开到中国及周边（拖拽出范围时不至于请求无意义区域）
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            raise ValueError("经纬度超出有效范围")
        lat, lon = _clamp_geo(lat, lon)
        observer={"name":request.args.get("name","成都"),"lat":lat,"lon":lon,"elev":round(observer_elevation(lat,lon),1)}
        return jsonify({"ok":True,"data":build_forecast(observer,model)})
    except Exception as e:
        return jsonify({"ok":False,"error":f"{type(e).__name__}: {e}"}),500


@APP.get("/api/current")
def api_current():
    """真实实况：Open-Meteo 当前时次 + 中国天气网/中央气象台站点观测（免费无 key）。
    附与今日晨间预报的一致性校验。"""
    try:
        lat=float(request.args.get("lat",DEFAULT_OBSERVER["lat"])); lon=float(request.args.get("lon",DEFAULT_OBSERVER["lon"]))
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            raise ValueError("经纬度超出有效范围")
        lat, lon = _clamp_geo(lat, lon)
        observer={"name":request.args.get("name","成都"),"lat":lat,"lon":lon,"elev":round(observer_elevation(lat,lon),1)}
        om = current_weather_om(observer)
        cn = current_weather_cn(lat, lon)
        # 今日晨间预报分由前端传入（lastData 中今天最高晨评分），避免后端重复计算
        today_score = request.args.get("today_score", type=float)
        match, note = current_check(om, today_score)
        return jsonify({"ok":True,"data":{
            "observer":observer,
            "om":om,
            "cn":cn,
            "wmo_zh": (WMO_ZH.get(om.get("weather_code")) if om else None),
            "check":{"match":match,"note":note},
            "today_score":today_score,
            "generated":datetime.now(TZ).isoformat()}})
    except Exception as e:
        return jsonify({"ok":False,"error":f"{type(e).__name__}: {e}"}),400


@APP.get("/api/obs")
def api_obs():
    """v2.9.4 真实观测三源：METAR 机场实况（能见度/云底/雾/湿度）+ 风云四号卫星云图 + 全国雷达拼图。"""
    try:
        metar = obs_metar()
        radar = obs_radar_url()
        sc_radar = obs_sc_radar_urls()
        sat = obs_sat_url()
        now = datetime.now(TZ)
        sat_note = None
        if not sat:
            hh = now.hour
            sat_note = ("当前时段无风云四号可见光云图（日出前/日落后无数据），"
                        "请以雷达拼图与机场 METAR 实况为准") if hh < 6 or hh >= 19 else "风云四号云图暂不可用（数据源无响应）"
        resp = jsonify({"ok": True, "data": {
            "metar": metar, "radar": radar, "sc_radar": sc_radar, "sat": sat,
            "sat_note": sat_note, "generated": now.isoformat()}})
        # v2.10.2: 禁用浏览器缓存，确保刷新页面时实况云图/雷达拿到最新时次
        resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/sat")
def api_sat():
    """v2.10.11 返回风云四号 B 星真彩色卫星云图 URL（供 Windy 云图图层叠加真实云图）。"""
    try:
        sat = obs_sat_url()
        if not sat:
            return jsonify({"ok": False, "error": "风云四号云图暂不可用（当前为日出前/日落后或数据源无响应）"}), 404
        resp = jsonify({"ok": True, "data": sat})
        resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/fy4-irx")
def api_fy4_irx():
    """v2.10.12 返回风云四号红外云图 PNG（NSMC WMS GEOS_IRX，裁剪川西区域放大，24 小时可用）。"""
    try:
        data = fy4_irx_png()
        if not data:
            return jsonify({"ok": False, "error": "风云四号红外云图暂不可用（数据源无响应或暂无最新时次）"}), 502
        return send_file(io.BytesIO(data["png"]), mimetype="image/png",
                         download_name="fy4_irx.png", max_age=0)
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/fy4-irx/info")
def api_fy4_irx_info():
    """v2.10.12 风云四号红外云图元信息（时次/边界/尺寸）。"""
    try:
        data = fy4_irx_png()
        if not data:
            return jsonify({"ok": False, "error": "风云四号红外云图暂不可用"}), 502
        resp = jsonify({"ok": True, "data": data["info"]})
        resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/fy4-irx/anim")
def api_fy4_irx_anim():
    """v2.10.15 风云四号红外云图多时次动画 GIF（近 2 小时 8 帧，川西区域裁剪）。"""
    try:
        data = fy4_irx_anim()
        if not data:
            return jsonify({"ok": False, "error": "风云四号云图动画暂不可用（历史帧不足或数据源无响应）"}), 502
        resp = send_file(io.BytesIO(data["gif"]), mimetype="image/gif",
                         download_name="fy4_irx_anim.gif", max_age=0)
        resp.headers["X-Fy4-Anim-Frames"] = str(data["info"]["frames"])
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/fy4-cloud")
def api_fy4_cloud():
    """v2.10.13 风云四号实况云况分析：中心像元 + 周围 81 像元云掩膜统计 + 云顶高度分层。"""
    try:
        lat = float(request.args.get("lat", DEFAULT_OBSERVER["lat"]))
        lon = float(request.args.get("lon", DEFAULT_OBSERVER["lon"]))
        data = fy4_cloud_analysis(lat, lon)
        if not data:
            return jsonify({"ok": False, "error": "风云四号云况分析暂不可用（数据源无响应或暂无最新时次）"}), 502
        resp = jsonify({"ok": True, "data": data})
        resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/fire-cloud")
def api_fire_cloud():
    """v2.10.19 实时火烧云潜力预报：太阳方位扇区内中高云识别 + 潜力评分 + 绘制范围。
    src: fy4(风云四号) / kma(韩国 GK2A 千里眼2A)。"""
    try:
        lat = float(request.args.get("lat", DEFAULT_OBSERVER["lat"]))
        lon = float(request.args.get("lon", DEFAULT_OBSERVER["lon"]))
        src = request.args.get("src", "fy4")
        data = fire_cloud_forecast(lat, lon, src=src)
        if not data:
            return jsonify({"ok": False, "error": "火烧云预报暂不可用（数据源无响应或太阳扇区内云数据不足）"}), 502
        resp = jsonify({"ok": True, "data": data})
        resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/fy4-cloud/gibs")
def api_fy4_cloud_gibs():
    """v2.10.14 NASA GIBS 卫星云顶高度：Himawari 红外分级 + MODIS CTH 定量统计。"""
    try:
        lat = float(request.args.get("lat", DEFAULT_OBSERVER["lat"]))
        lon = float(request.args.get("lon", DEFAULT_OBSERVER["lon"]))
        data = gibs_cloud_analysis(lat, lon)
        if not data:
            return jsonify({"ok": False, "error": "卫星云顶高度暂不可用（NASA GIBS 数据源无响应）"}), 502
        resp = jsonify({"ok": True, "data": data})
        resp.headers["Cache-Control"] = "no-store"
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/cloud-trend")
def api_cloud_trend():
    """v2.9.6 云图趋势：未来 3 小时云量趋势 + 当前云层判断（Open-Meteo 走廊采样）。"""
    try:
        lat = float(request.args.get("lat", DEFAULT_OBSERVER["lat"]))
        lon = float(request.args.get("lon", DEFAULT_OBSERVER["lon"]))
        lat, lon = _clamp_geo(lat, lon)
        data = cloud_trend(lat, lon)
        if not data:
            return jsonify({"ok": False, "error": "云量趋势数据源暂不可用"}), 502
        return jsonify({"ok": True, "data": data})
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/cloud-map")
def api_cloud_map():
    """v2.9.7 云图影响范围标注数据：观测点+雪山+视线走廊+走廊当前分层云量（用于在卫星云图上叠加绘制）。"""
    try:
        lat = float(request.args.get("lat", DEFAULT_OBSERVER["lat"]))
        lon = float(request.args.get("lon", DEFAULT_OBSERVER["lon"]))
        lat, lon = _clamp_geo(lat, lon)
        data = cloud_map_data(lat, lon)
        if not data:
            return jsonify({"ok": False, "error": "云图标注数据源暂不可用"}), 502
        return jsonify({"ok": True, "data": data})
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 400


@APP.get("/api/history")
def api_history():
    """历史回顾：任意日期范围（ERA5 再分析，1940 至今）。跨度上限 90 天防滥用。"""
    try:
        lat=float(request.args.get("lat",DEFAULT_OBSERVER["lat"])); lon=float(request.args.get("lon",DEFAULT_OBSERVER["lon"]))
        start=request.args.get("start",""); end=request.args.get("end","")
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            raise ValueError("经纬度超出有效范围")
        lat, lon = _clamp_geo(lat, lon)
        s=datetime.strptime(start,"%Y-%m-%d").date(); e=datetime.strptime(end,"%Y-%m-%d").date()
        if s > e: s, e = e, s
        if (e - s).days > 90: raise ValueError("历史跨度最大 90 天")
        today=datetime.now(TZ).date()
        if e > today: e = today
        if s > e: raise ValueError("日期范围无效")
        observer={"name":request.args.get("name","成都"),"lat":lat,"lon":lon,"elev":round(observer_elevation(lat,lon),1)}
        return jsonify({"ok":True,"data":build_history(observer, s.isoformat(), e.isoformat())})
    except Exception as ex:
        return jsonify({"ok":False,"error":f"{type(ex).__name__}: {ex}"}),400


@APP.get("/api/history/export")
def api_history_export():
    """导出历史回顾为 Excel（含历史 vs 预报趋势对比）。"""
    try:
        lat=float(request.args.get("lat",DEFAULT_OBSERVER["lat"])); lon=float(request.args.get("lon",DEFAULT_OBSERVER["lon"]))
        start=request.args.get("start",""); end=request.args.get("end","")
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            raise ValueError("经纬度超出有效范围")
        lat, lon = _clamp_geo(lat, lon)
        s=datetime.strptime(start,"%Y-%m-%d").date(); e=datetime.strptime(end,"%Y-%m-%d").date()
        if s > e: s, e = e, s
        if (e - s).days > 90: raise ValueError("历史跨度最大 90 天")
        today=datetime.now(TZ).date()
        if e > today: e = today
        if s > e: raise ValueError("日期范围无效")
        observer={"name":request.args.get("name","成都"),"lat":lat,"lon":lon,"elev":round(observer_elevation(lat,lon),1)}
        buf = export_history_xlsx(observer, s.isoformat(), e.isoformat())
        return send_file(buf, as_attachment=True, download_name=f"成都看雪山历史回顾_{s.isoformat()}_{e.isoformat()}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as ex:
        return jsonify({"ok":False,"error":f"{type(ex).__name__}: {ex}"}),400


@APP.get("/api/elevation-grid")
def api_elevation_grid():
    try:
        south=float(request.args["south"]); west=float(request.args["west"])
        north=float(request.args["north"]); east=float(request.args["east"])
        n=max(6,min(14,int(request.args.get("n",10))))
        if not (-90 <= south < north <= 90 and -180 <= west < east <= 180):
            raise ValueError("地图范围无效")
        # 采用格心采样；每格返回边界，前端直接绘制半透明分区。
        dy=(north-south)/n; dx=(east-west)/n
        points=[(south+(iy+.5)*dy,west+(ix+.5)*dx) for iy in range(n) for ix in range(n)]
        key=f"elev-grid:{south:.3f},{west:.3f},{north:.3f},{east:.3f},{n}"
        elevs=cache_get(key,86400*30)
        if elevs is None:
            elevs=open_meteo_elevations(points); cache_put(key,elevs)
        cells=[]
        for k,((lat,lon),z) in enumerate(zip(points,elevs)):
            iy=k//n; ix=k%n
            band="unknown" if z is None else ("low" if z<1500 else "mid" if z<3500 else "high")
            cells.append({"south":south+iy*dy,"north":south+(iy+1)*dy,"west":west+ix*dx,"east":west+(ix+1)*dx,"elevation":z,"band":band})
        return jsonify({"ok":True,"cells":cells,"source":"Open-Meteo GLO-90","thresholds":{"low":"<1500 m","mid":"1500–3500 m","high":"≥3500 m"}})
    except Exception as e:
        return jsonify({"ok":False,"error":f"{type(e).__name__}: {e}"}),400


HTML = r'''<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>成都看雪山 · Open-Meteo气溶胶版</title><link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<style>
:root{--bg:#f1f4f7;--panel:#ffffff;--line:#e1e7ec;--line2:#d3dce3;--text:#1b2836;--muted:#62798c;--accent:#2d6a8f;--accent2:#24587a;--soft:#eaf2f7;--gold:#c07f2a;--red:#c25548;--green:#3e8e6e}*{box-sizing:border-box}html{-webkit-text-size-adjust:100%}body{margin:0;background:var(--bg);color:var(--text);font-family:-apple-system,"PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif;-webkit-font-smoothing:antialiased;line-height:1.5}.wrap{max-width:1180px;margin:auto;padding:18px 16px 36px}.hero{display:flex;justify-content:space-between;align-items:flex-end;gap:14px;margin:4px 0 16px;padding-bottom:14px;border-bottom:1px solid var(--line)}.hero h1{margin:0;font-family:"Songti SC","Noto Serif SC","STSong",serif;font-size:28px;font-weight:600;letter-spacing:1px;color:#16222e}.hero p{margin:5px 0 0;color:var(--muted);font-size:13px;letter-spacing:.2px}button{border:1px solid var(--line2);background:#fff;color:#2c3d4b;border-radius:9px;padding:9px 15px;font-weight:600;font-size:13px;cursor:pointer;transition:background .15s,border-color .15s}.primary{background:var(--accent);border-color:var(--accent);color:#fff}.primary:hover{background:var(--accent2);border-color:var(--accent2)}.grid{display:grid;grid-template-columns:330px 1fr;gap:16px;align-items:start}.panel{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:16px;box-shadow:0 1px 2px rgba(23,43,58,.05)}.inputs{display:grid;grid-template-columns:1fr 1fr;gap:10px}.inputs label{font-size:12px;color:var(--muted);font-weight:500}.inputs .wide{grid-column:1/-1}input,select{width:100%;margin-top:5px;background:#fff;border:1px solid var(--line2);color:var(--text);border-radius:8px;padding:9px 10px;font-size:13px;outline:none;transition:border-color .15s}input:focus,select:focus{border-color:var(--accent)}.actions{display:flex;gap:8px;margin-top:12px}.status{margin-top:12px;padding:9px 12px;border-left:3px solid var(--accent);background:var(--soft);color:#41586b;font-size:12.5px;word-break:break-word;border-radius:0 8px 8px 0}#map{height:340px;border-radius:10px;background:#e9eef2}.legend{display:flex;gap:14px;flex-wrap:wrap;font-size:12px;color:var(--muted);margin-top:9px}#liveMap{height:270px;border-radius:10px;background:#e9eef2}.cards{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px;margin-top:16px}.mountain h3{margin:0 0 3px;font-size:16px}.meta{font-size:12px;color:var(--muted)}.days{display:grid;grid-template-columns:repeat(5,minmax(112px,1fr));gap:8px;overflow:auto;margin-top:12px;padding-bottom:2px}.day{background:#fafbfc;border:1px solid var(--line);border-radius:10px;padding:10px;min-width:112px;transition:border-color .15s}.day:hover{border-color:var(--line2)}.day strong{font-size:12px;color:var(--muted);font-weight:600}.score{font-size:30px;font-weight:800;margin:6px 0 8px;font-variant-numeric:tabular-nums;letter-spacing:-.5px}.bar{height:4px;background:#e7edf2;border-radius:4px;overflow:hidden}.bar i{display:block;height:100%;background:var(--green);border-radius:4px}.small{font-size:11.5px;color:var(--muted);line-height:1.7}.metrics{display:grid;grid-template-columns:1fr 1fr;gap:2px 10px;margin:8px 0;font-size:11.5px;color:var(--muted);line-height:1.6}.metrics b{color:var(--text);font-weight:600;font-variant-numeric:tabular-nums}.tabs{display:flex;gap:5px;margin:7px 0}.tabs span{font-size:11px;padding:3px 7px;background:#eef2f5;color:#4a6172;border-radius:6px}.sim-actions{display:flex;gap:5px;margin-top:8px}.sim-actions button{padding:6px 8px;font-size:11px;flex:1;border-radius:7px;background:#f4f7f9;border-color:#e0e7ec;color:#3b5163}.sim-actions button:hover{background:#e9eff3}.loading{opacity:.55;pointer-events:none}.rules{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:9px;margin:0 0 14px}.rule{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:10px 12px;font-size:12px;color:#3d5566;box-shadow:0 1px 2px rgba(23,43,58,.04)}.rule b{display:block;color:var(--gold);margin-bottom:3px;font-size:12px;letter-spacing:.3px}.rule small{color:#8aa0b0;font-size:11px}.emp-chip{display:inline-block;margin-top:4px;padding:2px 7px;border-radius:6px;background:var(--soft);color:var(--accent);font-size:10.5px}.sys-chip{display:inline-block;margin-top:4px;padding:2px 7px;border-radius:6px;background:#fdf3e3;color:#a06a1f;font-size:10.5px}.sys-chip.pos{background:#eaf6ef;color:#2f7d5c}.fog-chip{display:inline-block;margin-top:4px;padding:2px 7px;border-radius:6px;background:#f3e8fb;color:#7d3fb0;font-size:10.5px}
/* v2.9: 实况面板 */
.cur-panel{display:flex;gap:18px;align-items:center;flex-wrap:wrap;padding:14px 16px}.cur-main{display:flex;align-items:center;gap:12px}.cur-icon{font-size:30px;line-height:1}.cur-temp{font-size:30px;font-weight:800;font-variant-numeric:tabular-nums}.cur-cond{font-size:12.5px;color:var(--muted)}.cur-metrics{display:flex;gap:14px;flex-wrap:wrap;font-size:12px;color:var(--muted)}.cur-metrics b{color:var(--text);font-weight:600;font-variant-numeric:tabular-nums}.cur-check{margin-top:9px;padding:8px 12px;border-radius:8px;font-size:12.5px;background:var(--soft);color:#3d5566}.cur-check.warn{background:#fdf3e3;color:#9a5f13}.cur-check.good{background:#eaf6ef;color:#2f7d5c}.cur-src{font-size:11px;color:#8aa0b0;margin-top:6px}.cur-src b{color:var(--accent);font-weight:600}
/* v2.9.4: 真实观测三源面板（METAR + 卫星云图 + 雷达） */
.obs-panel{margin-top:12px}.obs-title{font-size:13px;font-weight:700;color:var(--text);margin:0 0 10px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}.obs-title .tag2{font-size:10.5px;color:var(--muted);background:var(--soft);padding:2px 8px;border-radius:6px;font-weight:500}.obs-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:8px;margin-bottom:12px}.metar-card{background:#fafbfc;border:1px solid var(--line);border-radius:10px;padding:9px 11px}.metar-card .ap{font-size:12px;font-weight:700;color:var(--text);display:flex;justify-content:space-between;align-items:center}.metar-card .ap small{color:var(--muted);font-weight:500;font-size:10.5px}.metar-row{display:flex;justify-content:space-between;font-size:11.5px;color:var(--muted);margin-top:5px;line-height:1.6}.metar-row b{color:var(--text);font-weight:600}.flt{font-size:10px;padding:1px 6px;border-radius:5px;font-weight:700}.flt.VFR{background:#eaf6ef;color:#2f7d5c}.flt.MVFR{background:#e8f1fb;color:#2d6a8f}.flt.IFR{background:#fdf3e3;color:#a06a1f}.flt.LIFR{background:#fbe9e7;color:#c25548}.obs-imgs{display:grid;grid-template-columns:1fr 1fr;gap:10px}.obs-imgs figure{margin:0;background:#fafbfc;border:1px solid var(--line);border-radius:10px;padding:8px;text-align:center}.obs-imgs img{width:100%;border-radius:8px;background:#eef3f6;display:block;min-height:120px;object-fit:contain}.obs-imgs figcaption{font-size:11px;color:var(--muted);margin-top:6px;line-height:1.6}.obs-imgs figcaption b{color:var(--accent)}.obs-note{font-size:11.5px;color:var(--muted);margin-top:10px;line-height:1.7}.obs-load{display:flex;align-items:center;gap:8px;color:var(--muted);font-size:12.5px;padding:6px 0}
/* v2.9.7: 云图叠加标注层 */
.sat-wrap{position:relative;display:block;border-radius:8px;overflow:hidden}.sat-wrap canvas{position:absolute;inset:0;width:100%;height:100%;pointer-events:none}.sat-legend{display:flex;gap:12px;justify-content:center;flex-wrap:wrap;font-size:10.5px;color:var(--muted);margin-top:5px}.sat-legend i{display:inline-block;width:14px;height:3px;border-radius:2px;vertical-align:middle;margin-right:3px}.sat-legend .pt{display:inline-block;width:8px;height:8px;border-radius:50%;vertical-align:middle;margin-right:3px;background:#d64541;border:1.5px solid #fff;box-shadow:0 0 0 1px #d64541}
/* v2.9.5: 四川单站雷达 */
.obs-sc{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:10px}.obs-sc figure{margin:0;background:#fafbfc;border:1px solid var(--line);border-radius:10px;padding:8px;text-align:center}.obs-sc figure.sc-main{border:2px solid var(--accent);box-shadow:0 2px 8px rgba(45,106,143,.12)}.obs-sc img{width:100%;border-radius:8px;background:#eef3f6;display:block;min-height:150px;object-fit:contain}.obs-sc figcaption{font-size:11px;color:var(--muted);margin-top:6px;line-height:1.6}.obs-sc figcaption b{color:var(--accent)}
/* v2.9.6: 云图趋势 + 当前云层判断 */
.obs-trend{display:grid;grid-template-columns:1.2fr 1fr;gap:12px;margin-top:12px;align-items:stretch}.obs-trend .card{background:#fafbfc;border:1px solid var(--line);border-radius:10px;padding:11px 13px}.obs-trend .card h4{margin:0 0 9px;font-size:12.5px;color:var(--text)}.ct-hours{display:grid;grid-template-columns:repeat(4,1fr);gap:6px}.ct-h{background:#fff;border:1px solid var(--line);border-radius:8px;padding:7px 6px;text-align:center}.ct-h .t{font-size:10.5px;color:var(--muted)}.ct-h .v{font-size:16px;font-weight:800;color:var(--text);font-variant-numeric:tabular-nums}.ct-h .g{font-size:10px;color:#5a7487;margin-top:1px}.ct-h .pop{font-size:10px;color:var(--gold)}.ct-bar{height:5px;border-radius:3px;overflow:hidden;background:#e7edf2;margin-top:5px;display:flex}.ct-bar i{height:100%;display:block}.ct-bar i.low{background:#8ab0c9}.ct-bar i.mid{background:#8ac4c9}.ct-bar i.high{background:#c9c2e8}.ct-trend{font-size:12px;color:#3d5566;margin-top:8px;line-height:1.6;padding-top:7px;border-top:1px dashed var(--line)}.ct-judge{display:flex;flex-direction:column;justify-content:center;gap:7px}.ct-layer{display:inline-block;align-self:flex-start;padding:2px 9px;border-radius:6px;font-size:11px;font-weight:700;background:var(--soft);color:var(--accent)}.ct-layer.low{background:#eaf3f8;color:#3a7ca5}.ct-layer.mid{background:#eaf7f7;color:#2f8f94}.ct-layer.high{background:#f1eefb;color:#6a56b0}.ct-layer.clear{background:#eaf6ef;color:#2f7d5c}.ct-layer.sct{background:#f2f6f8;color:#5a7487}.ct-layer.rain{background:#e8eef3;color:#3d5a75}.ct-layer.cb{background:#f0e8f5;color:#7d3fb0}.ct-text{font-size:12px;color:#41586b;line-height:1.7}
@media(max-width:780px){.obs-trend{grid-template-columns:1fr}}.foot{color:var(--muted);font-size:12px;line-height:1.8;margin:16px 0}.modal{position:fixed;inset:0;background:rgba(20,32,42,.45);z-index:9999;display:none;align-items:center;justify-content:center;padding:12px;backdrop-filter:blur(2px)}.modal.open{display:flex}.sim-box{width:min(900px,100%);max-height:96vh;overflow:auto;background:#fff;border:1px solid var(--line);border-radius:14px;padding:16px;box-shadow:0 18px 50px rgba(20,32,42,.25)}.sim-head{display:flex;justify-content:space-between;align-items:flex-start;gap:10px}.sim-head h3{margin:0;font-size:16px}.sim-close{padding:7px 12px}.sim-canvas{width:100%;aspect-ratio:16/9;display:block;background:#eef4f8;border-radius:10px;margin-top:12px;border:1px solid var(--line)}.sim-metrics{display:grid;grid-template-columns:repeat(6,1fr);gap:6px;margin-top:10px}.sim-metrics div{background:#f4f7f9;border:1px solid var(--line);border-radius:8px;padding:7px;text-align:center;font-size:11px;color:var(--muted)}.sim-metrics b{display:block;color:var(--text);font-size:14px;font-variant-numeric:tabular-nums}.sim-note{font-size:12px;color:var(--muted);line-height:1.7;margin-top:9px}.evidence{margin-top:16px}.evidence summary{cursor:pointer;font-weight:700;font-size:13px;color:#2c3d4b;padding:10px 12px;background:var(--panel);border:1px solid var(--line);border-radius:10px;list-style:none}.evidence summary::-webkit-details-marker{display:none}.evidence summary::after{content:"＋";float:right;color:#8aa0b0}.evidence[open] summary::after{content:"－"}.sample-list{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px;margin-top:9px}.sample{background:var(--panel);border:1px solid var(--line);border-radius:9px;padding:10px 12px;font-size:12px}.sample b{color:var(--accent);font-weight:600}.tag{display:inline-block;margin-top:6px;padding:2px 7px;border-radius:6px;background:var(--soft);color:#5a7487;font-size:10.5px}
/* v2.5: Tab 导航与板块 */
.tabs-nav{display:flex;gap:6px;margin:0 0 16px;border-bottom:1px solid var(--line);padding-bottom:10px}.tab-btn{border:none;background:transparent;color:var(--muted);padding:8px 16px;font-size:14px;font-weight:600;border-radius:8px 8px 0 0;position:relative;cursor:pointer}.tab-btn:hover{background:var(--soft);color:var(--accent)}.tab-btn.active{color:var(--accent);background:var(--soft)}.tab-btn.active::after{content:"";position:absolute;left:12px;right:12px;bottom:-11px;height:3px;background:var(--accent);border-radius:2px}.tab-pane{display:none}.tab-pane.active{display:block}.hist-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:8px;margin-top:10px}.hist-day{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:9px 11px;font-size:11.5px}.hist-day b{font-size:15px;font-variant-numeric:tabular-nums}.hist-tag{display:inline-block;font-size:9.5px;color:#b07a2a;background:#fdf3e3;border:1px solid #f0dfc0;border-radius:4px;padding:0 4px;margin-left:4px;vertical-align:1px;font-weight:600}.hist-sim{display:flex;gap:5px;margin-top:7px}.hist-sim button{flex:1;padding:4px 6px;font-size:10.5px;border-radius:6px;background:#f4f7f9;border-color:#e0e7ec;color:#3b5163;cursor:pointer}.hist-sim button:hover{background:#e9eff3;border-color:var(--accent);color:var(--accent)}.hist-head{font-weight:700;color:var(--text);font-size:13px;margin:16px 0 2px;padding-bottom:6px;border-bottom:1px solid var(--line)}.hist-note{color:var(--muted);font-size:12px;line-height:1.7;margin:8px 0 0}
/* v2.6: 历史日期选择 */
.hist-ctl{display:flex;gap:8px;align-items:flex-end;flex-wrap:wrap;padding:12px;background:var(--soft);border:1px solid var(--line);border-radius:10px;margin-bottom:10px}.hist-ctl label{font-size:12px;color:var(--muted);font-weight:500}.hist-ctl input[type=date]{margin-top:4px;width:auto}.hist-ctl .spacer{flex:1}.hist-ctl .hint{font-size:11.5px;color:var(--muted);width:100%}
/* v2.10: 响应式布局（自适应桌面/平板/手机） */
@media(max-width:980px){.obs-trend{grid-template-columns:1fr}.obs-imgs{grid-template-columns:1fr 1fr}}
@media(max-width:780px){.grid{grid-template-columns:1fr}.cards{grid-template-columns:1fr}.hero{align-items:flex-start;flex-direction:column}.hero .primary{width:100%}#map{height:290px}#liveMap{height:240px}.sim-metrics{grid-template-columns:repeat(3,1fr)}.sample-list{grid-template-columns:1fr}.obs-imgs{grid-template-columns:1fr}.tabs-nav{overflow-x:auto;flex-wrap:nowrap;-webkit-overflow-scrolling:touch;padding-bottom:8px;scrollbar-width:none}.tabs-nav::-webkit-scrollbar{display:none}.tab-btn{white-space:nowrap;padding:8px 13px;font-size:13.5px}.cur-panel{gap:12px;padding:12px}.cur-metrics{gap:10px}.cur-temp{font-size:26px}.inputs{grid-template-columns:1fr}.actions{flex-wrap:wrap}.actions button{flex:1}.ct-hours{grid-template-columns:repeat(2,1fr)}.hist-ctl button{flex:1}.hist-ctl input[type=date]{flex:1;min-width:0}.sat-legend{gap:8px;font-size:10px}.panel{padding:13px}.wrap{padding:14px 12px 28px}.hero h1{font-size:24px}}
@media(max-width:420px){.cur-metrics{font-size:11.5px}.ct-hours{grid-template-columns:1fr 1fr}.tab-btn{padding:8px 10px;font-size:13px}.hist-head{font-size:12.5px}}
/* v2.10.9: Windy 实况嵌入式板块 */
.windy-panel{width:100%}.windy-ctl{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px}.windy-ctl button{flex:1;min-width:76px;padding:8px 6px;font-size:12px;border-radius:8px;background:#f4f7f9;border-color:#e0e7ec;color:#3b5163;white-space:nowrap}.windy-ctl button.on{background:var(--accent);border-color:var(--accent);color:#fff}.windy-box{position:relative;width:100%;aspect-ratio:16/9;background:#0b1220;border-radius:10px;overflow:hidden;border:1px solid var(--line)}.windy-box iframe{position:absolute;inset:0;width:100%;height:100%;border:0;display:block}.windy-note{font-size:11.5px;color:var(--muted);margin-top:9px;line-height:1.7}.windy-note b{color:var(--accent)}.windy-load{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;color:#8aa0b0;font-size:13px;letter-spacing:1px;background:#0b1220;pointer-events:none;transition:opacity .3s}.windy-box.loading .windy-load{opacity:1}.windy-box .windy-load{opacity:0}
/* v2.10.10: Windy.app 数值云图瓦片板块（数据源/云量类型切换） */
.wm-ctl{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:8px}.wm-ctl .grp{display:flex;gap:4px;align-items:center;flex-wrap:wrap}.wm-ctl .grp b{font-size:11px;color:var(--muted);font-weight:600;margin-right:2px}.wm-ctl button{min-width:0;padding:6px 10px;font-size:12px;border-radius:8px;background:#f4f7f9;border-color:#e0e7ec;color:#3b5163;white-space:nowrap}.wm-ctl button.on{background:var(--accent);border-color:var(--accent);color:#fff}.wm-map{height:420px;border-radius:10px;background:#0b1220;border:1px solid var(--line);position:relative}.wm-note{font-size:11.5px;color:var(--muted);margin-top:9px;line-height:1.7}.wm-note b{color:var(--accent)}.wm-status{font-size:11px;color:var(--muted);margin:2px 0 8px}
@media(max-width:780px){.wm-map{height:320px}.wm-ctl .grp{width:100%}}
</style></head><body><div class="wrap"><div class="hero"><div><h1>成都 · 看雪山</h1><p>数值预报 × 观山经验 综合测算观测指数</p></div><button class="primary" onclick="loadAll()">刷新全部数据</button></div>
<nav class="tabs-nav"><button class="tab-btn active" data-tab="fc" onclick="switchTab('fc')">天气预报</button><button class="tab-btn" data-tab="live" onclick="switchTab('live')">实况观测</button><button class="tab-btn" data-tab="hist" onclick="switchTab('hist')">历史回顾</button><button class="tab-btn" data-tab="kb" onclick="switchTab('kb')">观山知识</button><button class="tab-btn" data-tab="windy" onclick="switchTab('windy')">Windy 实况</button><button class="tab-btn" data-tab="wm" onclick="switchTab('wm')">卫星云图</button></nav>
<div id="pane-live" class="tab-pane"><div class="panel"><div class="obs-title">观测点地图 <span class="tag2">拖动蓝点调整观测位置 · 松手自动刷新实况与云量趋势</span></div><div id="liveMap"></div><div class="legend"><span>● 观测点（可拖动）</span><span style="color:#8a6d3b">● 雪山</span><span>线段 = 实际观测方向</span></div></div><div class="panel" style="margin-top:12px"><div class="obs-title">当前天气实况 <span class="tag2">中央气象台站点观测 + Open-Meteo 当前时次</span></div><div id="currentBox"><div class="cur-panel" style="justify-content:center;color:var(--muted);font-size:13px">正在获取实况观测…</div></div></div><div id="obsBox" class="panel obs-panel" style="display:none"></div><div class="panel" style="margin-top:12px"><div class="obs-title">风云四号实况云况分析 <span class="tag2">红外云检测（白=云）· 跟随当前观测点 · 自动刷新</span></div><div id="fy4CloudBox"><div class="hist-note" style="color:var(--muted)">正在获取风云四号云况分析…</div></div></div><div class="panel" style="margin-top:12px"><div class="obs-title">火烧云潜力预报 <span class="tag2">风云四号红外云图 · 太阳方向中高云识别 · 10 分钟缓存</span></div><div id="fireCloudBox"><div class="hist-note" style="color:var(--muted)">正在分析太阳方向中高云分布…</div></div></div></div>
<div id="pane-fc" class="tab-pane active"><div class="grid"><section class="panel"><div class="obs-title">预报设置 <span class="tag2">观测点 · 海拔 · 模型</span></div><div class="inputs"><label>纬度<input id="lat" value="{{observer.lat}}"></label><label>经度<input id="lon" value="{{observer.lon}}"></label><label>海拔 m（自动）<input id="elev" value="读取中" readonly></label><label>地点<input id="name" value="{{observer.name}}"></label><label class="wide">天气预报数据源<select id="model" onchange="loadAll()"><option value="best_match">智能最佳匹配（推荐）</option><option value="ecmwf_ifs025">ECMWF IFS 0.25°</option><option value="gfs_seamless">NOAA GFS</option><option value="icon_seamless">DWD ICON</option><option value="cma_grapes_global">中国气象局 CMA GRAPES</option><option value="jma_seamless">日本气象厅 JMA</option></select></label></div><div class="actions"><button onclick="locate()">手机定位</button><button onclick="loadAll()">更新全部数据</button></div><div id="modelStatus" class="status" style="display:none">天气模型：正在获取…</div><div id="aerosol" class="status" style="display:none">气溶胶：正在获取 Open-Meteo 数据…</div></section><section class="panel"><div class="obs-title">观测地图 <span class="tag2">可拖动观测点 · 海拔分区</span></div><div id="map"></div><div class="legend"><span>● 观测点</span><span style="color:#8a6d3b">● 雪山</span><span>线段 = 实际观测方向</span><span style="color:#d64541">● 云层遮挡视线</span><span style="color:#4a9d78">■ 低海拔 &lt;1500m</span><span style="color:#c98a2b">■ 中海拔 1500–3500m</span><span style="color:#8a6bbd">■ 高海拔 ≥3500m</span></div></section></div>
<div id="cards" class="cards"></div></div>
<div id="pane-hist" class="tab-pane"><div class="hist-ctl"><label>起始日期<input type="date" id="hStart"></label><label>结束日期<input type="date" id="hEnd"></label><button class="primary" onclick="queryHistory()">查询回顾</button><button onclick="exportHistory()">导出为 Excel</button><div class="spacer"></div><div class="hint">支持任意历史日期（ERA5 再分析，1940 年至今），单次跨度最多 90 天；近 92 天含气溶胶 AOD 数据，更早则无 AOD（评分自动降级）。导出文件含「每日评分」与「历史 vs 预报趋势对比」两个工作表。</div></div>
<div class="panel"><div class="hist-head" style="margin-top:0">评分趋势：历史（回算） vs 预报</div><div class="hist-note">实线+实心点 = 过去 7 天历史回算（ERA5 再分析，虚线右侧为预报）；空心点虚线 = 未来 5 天预报。竖线右侧即预报区间。</div><canvas id="trendCanvas" style="width:100%;height:280px;display:block;margin-top:8px"></canvas></div>
<div id="histBox" class="panel" style="margin-top:12px"><div class="hist-note">选择日期范围后点击「查询回顾」，这里会展示各座山的每日评分。</div></div></div>
<div id="pane-kb" class="tab-pane"><div class="rules">{% for t,c,s in rules %}<div class="rule"><b>{{ t }}</b>{{ c }}<small>来源：{{ s }}</small></div>{% endfor %}</div>
<details class="evidence"><summary>历史成功案例库（{{ success_cases|length }} 例 · 均来自公开报道，可对照学习经验规律）</summary><div class="sample-list">{% for c in success_cases %}<div class="sample"><b>{{ c.date }}</b><div>{{ c.cond }} → {{ c.sight }}</div><span class="tag">{{ c.src }}</span></div>{% endfor %}</div></details></div></div>
<div id="pane-windy" class="tab-pane"><div class="windy-panel"><div class="panel"><div class="obs-title">Windy 实况地图 <span class="tag2">嵌入式实时气象图层 · 自动跟随当前观测点</span></div><div class="windy-ctl"><button data-windylayer="wind" class="on" onclick="setWindyLayer('wind',this)">风场</button><button data-windylayer="satellite" onclick="setWindyLayer('satellite',this)">卫星云图</button><button data-windylayer="clouds" onclick="setWindyLayer('clouds',this)">云量</button><button data-windylayer="temp" onclick="setWindyLayer('temp',this)">温度</button><button data-windylayer="rain" onclick="setWindyLayer('rain',this)">降水</button><button data-windylayer="rh" onclick="setWindyLayer('rh',this)">湿度</button><button data-windylayer="pressure" onclick="setWindyLayer('pressure',this)">气压</button></div><div id="windyBox" class="windy-box loading"><div class="windy-load">正在加载 Windy 实时地图…</div></div><div class="windy-note">数据由 <b>Windy.com</b> 通过嵌入式 iframe 提供（ECMWF/GFS/ICON 模型叠加实时观测与卫星云图），随你的观测点与所选图层实时更新。卫星云图可直观观察西部雪山区域云况。</div></div></div></div>
<div id="pane-wm" class="tab-pane"><div class="panel"><div class="obs-title">风云四号卫星云图 <span class="tag2">红外动画（多时次）· 真彩色叠加 · 跟随当前观测点</span></div><div class="wm-ctl"><span class="grp"><b>实况</b><button data-wm-sat="fy4" onclick="toggleFy4Sat(this)">风云四号红外</button><button data-wm-sat="wxbl" onclick="toggleFy4Sat(this)">真彩色云图</button></span></div><div id="wmStatus" class="wm-status">正在初始化…</div><div id="wmMap" class="wm-map"></div><div class="wm-note">「<b>风云四号红外</b>」叠加国家卫星气象中心 WMS 红外云图（24 小时可用，白色为云，近 2 小时多时次动画循环），「<b>真彩色云图</b>」叠加中央气象台风云四号 B 星真彩色（仅白天）。点击左上角图层按钮可切换高德底图（标准/卫星）。</div></div></div>
<div id="simModal" class="modal" onclick="if(event.target===this)closeSimulation()"><div class="sim-box"><div class="sim-head"><div><h3 id="simTitle">天空形态模拟</h3><div id="simSub" class="meta"></div></div><button class="sim-close" onclick="closeSimulation()">关闭</button></div><canvas id="simCanvas" class="sim-canvas"></canvas><div id="simMetrics" class="sim-metrics"></div><div id="simNote" class="sim-note"></div></div></div>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script><script>
const gaodeOpt={subdomains:'1234',maxZoom:18,attribution:'地图 © 高德'};
// 部分手机网络无法访问 webrd 域；标准图优先使用与卫星图相同、已验证可达的 webst 域。
const gaodeRoad=L.tileLayer('https://webst0{s}.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=7&x={x}&y={y}&z={z}',gaodeOpt);
// 若当地节点不在 webst 提供 style=7，则逐瓦片回退到传统 webrd 标准图地址。
gaodeRoad.on('tileerror',e=>{const t=e.tile;if(t.dataset.gaodeRetry)return;t.dataset.gaodeRetry='1';t.src=t.src.replace(/webst0[1-4]/,'webrd02')});
const gaodeSat=L.tileLayer('https://webst0{s}.is.autonavi.com/appmaptile?style=6&x={x}&y={y}&z={z}',gaodeOpt);
const gaodeLabel=L.tileLayer('https://webst0{s}.is.autonavi.com/appmaptile?style=8&x={x}&y={y}&z={z}',gaodeOpt);
const gaodeSatellite=L.layerGroup([gaodeSat,gaodeLabel]);
const elevationLayer=L.layerGroup();
const viewpointLayer=L.layerGroup();
let map=L.map('map',{layers:[gaodeRoad,elevationLayer,viewpointLayer]}).setView([30.657,104.058],8),layers=[];
map.createPane('terrainPane');map.getPane('terrainPane').style.zIndex=350;
L.control.layers({'高德标准地图':gaodeRoad,'高德卫星影像':gaodeSatellite},{'成都观山季推荐点':viewpointLayer,'海拔分区':elevationLayer},{position:'topright',collapsed:true}).addTo(map);
const $=x=>document.getElementById(x),obs=()=>({name:$('name').value,lat:+$('lat').value,lon:+$('lon').value,model:$('model').value});
// v2.3: 全局请求互斥——预报请求进行中时忽略新的触发，避免拖拽连点打爆数据源限流
let _forecastBusy=false,_lastReqKey='',_retry429=0,_dragTimer=null;
const viewpoints={{ viewpoints|tojson }};
function color(s){return s>=80?'#3e8e6e':s>=65?'#5f9b78':s>=45?'#c98a2b':'#c25548'}
function fmtTime(t){return new Date(t).toLocaleTimeString('zh-CN',{hour:'2-digit',minute:'2-digit'})}
function outOfChina(lat,lon){return lon<72.004||lon>137.8347||lat<0.8293||lat>55.8271}
function transformLat(x,y){let r=-100+2*x+3*y+.2*y*y+.1*x*y+.2*Math.sqrt(Math.abs(x));r+=(20*Math.sin(6*x*Math.PI)+20*Math.sin(2*x*Math.PI))*2/3;r+=(20*Math.sin(y*Math.PI)+40*Math.sin(y/3*Math.PI))*2/3;r+=(160*Math.sin(y/12*Math.PI)+320*Math.sin(y*Math.PI/30))*2/3;return r}
function transformLon(x,y){let r=300+x+2*y+.1*x*x+.1*x*y+.1*Math.sqrt(Math.abs(x));r+=(20*Math.sin(6*x*Math.PI)+20*Math.sin(2*x*Math.PI))*2/3;r+=(20*Math.sin(x*Math.PI)+40*Math.sin(x/3*Math.PI))*2/3;r+=(150*Math.sin(x/12*Math.PI)+300*Math.sin(x/30*Math.PI))*2/3;return r}
function wgsToGcj(lat,lon){if(outOfChina(lat,lon))return[lat,lon];const a=6378245,ee=.00669342162296594323,dLat=transformLat(lon-105,lat-35),dLon=transformLon(lon-105,lat-35),rad=lat/180*Math.PI,magic=1-ee*Math.sin(rad)**2,sqrt=Math.sqrt(magic);return[lat+dLat*180/((a*(1-ee))/(magic*sqrt)*Math.PI),lon+dLon*180/(a/sqrt*Math.cos(rad)*Math.PI)]}
function gcjToWgs(lat,lon){if(outOfChina(lat,lon))return[lat,lon];const g=wgsToGcj(lat,lon);return[lat*2-g[0],lon*2-g[1]]}
function setViewpoint(i){const v=viewpoints[i];$('name').value=v.name;$('lat').value=v.lat.toFixed(6);$('lon').value=v.lon.toFixed(6);$('elev').value='读取中';map.closePopup();syncCloudMapOverlay();loadAll()}
viewpoints.forEach((v,i)=>{const p=wgsToGcj(v.lat,v.lon);L.circleMarker(p,{radius:6,color:'#4a9d78',weight:2,fillColor:'#3e8e6e',fillOpacity:.95}).bindPopup(`<b>${v.name}</b><br>${v.district} · ${v.target}<br><small>${v.precision}；出发前请以高德导航入口为准</small><br><button onclick="setViewpoint(${i})" style="margin-top:7px;padding:6px 8px">设为观测点并预报</button>`).addTo(viewpointLayer)});
// v2.10.3: 实况板块观测点地图（与预报板块主地图双向联动，拖动即刷新云量趋势/云图/实况/预报）
const liveOmIcon=L.divIcon({className:'',html:'<div style="width:22px;height:22px;border-radius:50%;background:#2d6a8f;border:3px solid #fff;box-shadow:0 1px 4px rgba(23,43,58,.45);cursor:grab"></div>',iconSize:[22,22],iconAnchor:[11,11]});
const _liveObs=wgsToGcj(+$('lat').value,+$('lon').value);
// v2.10.3: 实况地图使用独立瓦片实例（Leaflet 图层不能同时挂在两个地图上）
const gaodeRoadLive=L.tileLayer('https://webst0{s}.is.autonavi.com/appmaptile?lang=zh_cn&size=1&scale=1&style=7&x={x}&y={y}&z={z}',gaodeOpt);
let liveMap=L.map('liveMap',{layers:[gaodeRoadLive]}).setView(_liveObs,9);
const liveOm=L.marker(_liveObs,{icon:liveOmIcon,draggable:true,bubblingMouseEvents:false,autoPan:true}).addTo(liveMap).bindPopup('拖动我调整观测位置，松手自动刷新实况与云量趋势');
liveOm.on('dragstart',()=>{liveOm._wasDrag=liveMap.dragging.enabled();liveMap.dragging.disable();liveOm.getElement().style.cursor='grabbing'});
liveOm.on('drag',()=>{const g=liveOm.getLatLng();let la=g.lat,lo=g.lng,chg=false;
// v2.10.21: 放开到中国及周边（风云四号云图覆盖范围），不再限制在川渝
if(la<10){la=10;chg=true}if(la>55){la=55;chg=true}
if(lo<70){lo=70;chg=true}if(lo>135){lo=135;chg=true}
if(chg)liveOm.setLatLng([la,lo])});
liveOm.on('dragend',()=>{liveOm.getElement().style.cursor='grab';if(liveOm._wasDrag)liveMap.dragging.enable();const g=liveOm.getLatLng(),w=gcjToWgs(g.lat,g.lng);$('lat').value=w[0].toFixed(6);$('lon').value=w[1].toFixed(6);$('name').value='地图拖拽点';$('elev').value='读取中';liveMap.closePopup();
// 同步主地图观测点标记
if(window._mainOm)window._mainOm.setLatLng(wgsToGcj(w[0],w[1]));
// v2.10.3: 实况联动——云图叠加标注 + 未来云量趋势 + 当前实况 + 预报
syncCloudMapOverlay();loadCurrent();loadFy4Cloud();
clearTimeout(_dragTimer);_dragTimer=setTimeout(()=>loadAll(),1200)});
// v2.10.18: 实况地图缩放时重绘火烧云范围（保持渐变清晰）
liveMap.on('zoomend',()=>{if(_fcData)drawFireCloudMap(_fcData,true)});
function syncLiveOmFromMain(){if(typeof liveOm!=='undefined'){const g=wgsToGcj(+$('lat').value,+$('lon').value);liveOm.setLatLng(g)}}
let lastData=null;
let blockedLayer=L.layerGroup();
function bandName(k){return k==='low'?'低云':k==='mid'?'中云':'高云'}
// v2.1: 地图红色标注视线遮挡位置；null 参数=画所有山今天清晨，否则画指定山某时段
function paintBlocked(mi,di,period){
  map.removeLayer(blockedLayer);
  blockedLayer=L.layerGroup();
  const add=b=>{const g=wgsToGcj(b.lat,b.lon);L.circleMarker(g,{radius:7,color:'#c0392b',weight:2,fillColor:'#d64541',fillOpacity:.9}).addTo(blockedLayer).bindPopup(`<b style="color:#c0392b">云层遮挡视线</b><br>${b.distance} km 处 · ${bandName(b.band)} ${b.cover}%<br><small>该处云层高度带与观测视线相交，挡住峰景</small>`)};
  if(mi==null){lastData.mountains.forEach(m=>(m.daily[0].morning.blocked||[]).forEach(add))}
  else{const x=lastData.mountains[mi].daily[di][period];(x.blocked||[]).forEach(add)}
  blockedLayer.addTo(map);
}
function draw(d){try{_draw(d)}catch(e){console.warn('渲染失败',e);$('modelStatus').textContent='渲染失败：'+e.message;$('modelStatus').style.borderColor='#c25548'}}
function _draw(d){lastData=d;layers.forEach(x=>map.removeLayer(x));layers=[];let o=d.observer,og=wgsToGcj(o.lat,o.lon),bounds=[og];$('elev').value=o.elev.toFixed(0);
// v2.2: 观测点可拖拽。用 L.marker+divIcon 替代 circleMarker：Leaflet 对 marker 的拖拽是原生实现，
// bubblingMouseEvents:false 让拖动时地图不会跟着动。
const omIcon=L.divIcon({className:'',html:'<div style="width:22px;height:22px;border-radius:50%;background:#2d6a8f;border:3px solid #fff;box-shadow:0 1px 4px rgba(23,43,58,.45);cursor:grab"></div>',iconSize:[22,22],iconAnchor:[11,11]});
const om=L.marker(og,{icon:omIcon,draggable:true,bubblingMouseEvents:false,autoPan:true}).addTo(map).bindPopup(`${o.name}<br>自动海拔 ${o.elev}m<br><small>拖动我调整观测位置，松手自动预测</small>`);
om.on('dragstart',()=>{const was=map.dragging.enabled();window._dragMap=was;map.dragging.disable();om.getElement().style.cursor='grabbing'});
om.on('drag',()=>{const g=om.getLatLng();let la=g.lat,lo=g.lng,chg=false;
// v2.10.21: 放开到中国及周边（风云四号云图覆盖范围），防止拖到无效区域导致预报失败
if(la<10){la=10;chg=true}if(la>55){la=55;chg=true}
if(lo<70){lo=70;chg=true}if(lo>135){lo=135;chg=true}
if(chg)om.setLatLng([la,lo])});
om.on('dragend',()=>{om.getElement().style.cursor='grab';if(window._dragMap)map.dragging.enable();const g=om.getLatLng(),w=gcjToWgs(g.lat,g.lng);$('lat').value=w[0].toFixed(6);$('lon').value=w[1].toFixed(6);$('name').value='地图拖拽点';$('elev').value='读取中';map.closePopup();window._dragFit=true;syncCloudMapOverlay();syncLiveOmFromMain();
// v2.3: 防抖 1.2s——连续拖动只发一次预报，避免连点触发限流
clearTimeout(_dragTimer);_dragTimer=setTimeout(()=>loadAll(),1200)});
layers.push(om);window._mainOm=om;syncLiveOmFromMain();
d.mountains.forEach(m=>{let mg=wgsToGcj(m.lat,m.lon);bounds.push(mg);
if(m.terrain_blocked){
  // v2.10.21: 地形遮挡——灰显虚线 + 遮挡提示弹窗，不展示该山预测
  layers.push(L.polyline([og,mg],{color:'#c9ccd1',weight:1.2,dashArray:'2 7',interactive:false}).addTo(map));
  layers.push(L.circleMarker(mg,{radius:6,color:'#9aa0a6',weight:1.5,fillColor:'#c0c4c8',fillOpacity:.75}).addTo(map).bindPopup(`${m.name}<br>${m.elev}m · ${m.distance}km<br><b style="color:#c25548">地形遮挡</b>：${m.terrain_note||'峰顶被中间山体遮挡'}`));
}else{
  layers.push(L.polyline([og,mg],{color:'#8ba3b5',weight:1.8,dashArray:'6 7'}).addTo(map));
  layers.push(L.circleMarker(mg,{radius:7,color:'#8a6d3b',weight:1.5,fillColor:'#b8964e',fillOpacity:.9}).addTo(map).bindPopup(`${m.name}<br>${m.elev}m · ${m.distance}km`));
}});if(window._dragFit){window._dragFit=false}else{map.fitBounds(bounds,{padding:[25,25]})}paintBlocked();drawHistory(d);drawTrend(d);
$('cards').innerHTML=d.mountains.map((m,mi)=>m.terrain_blocked?
`<section class="panel mountain"><h3>${m.name} <span class="tag2" style="background:#c25548;color:#fff">地形遮挡</span></h3><div class="meta">${m.distance} km · 方位 ${m.bearing}° · 峰顶仰角约 ${m.peak_angle}°</div><div class="hist-note" style="color:#c25548;margin-top:8px">⛰ ${m.terrain_note||'峰顶被中间山体遮挡'}。已取消显示该山预测。</div></section>`
:`<section class="panel mountain"><h3>${m.name}</h3><div class="meta">${m.distance} km · 方位 ${m.bearing}° · 峰顶仰角约 ${m.peak_angle}°</div><div class="days">${m.daily.map((x,di)=>{let b=x.morning;return `<div class="day"><strong>${x.date.slice(5)}</strong><div class="score" style="color:${color(b.score)}">${b.score}</div><div class="bar"><i style="width:${b.score}%"></i></div><div class="tabs"><span>晨 ${fmtTime(b.time)}</span><span>晚 ${x.evening.score}</span></div><div class="metrics"><span>AOD550 <b>${b.aod==null?'暂无':b.aod.toFixed(2)}</b></span><span>低云 <b>${b.low}%</b></span><span>能见 <b>${b.visibility}km</b></span><span>湿度 <b>${b.rh}%</b></span></div><div class="small">金山最高 ${x.gold.gold} 分<br><span class="emp-chip">经验 ${b.empirical_bonus>=0?'+':''}${b.empirical_bonus} · 洗尘+${b.empirical.washout} 霾−${b.empirical.haze} 季${b.empirical.season>=0?'+':''}${b.empirical.season} 窗+${b.empirical.window} 晴+${b.empirical.clear}</span>${synTags(b.synoptic)}${fogTag(b.fog)}<br>${b.reasons.join('、')}</div><div class="sim-actions"><button onclick="openSimulation(${mi},${di},'morning')">晨间形态</button><button onclick="openSimulation(${mi},${di},'evening')">傍晚形态</button></div></div>`}).join('')}</div></section>`).join('')}

// v2.10: 板块整合——天气预报(fc)/实况观测(live)/观山知识(kb)
function switchTab(t){document.querySelectorAll('.tab-btn').forEach(b=>b.classList.toggle('active',b.dataset.tab===t));document.querySelectorAll('.tab-pane').forEach(p=>p.classList.toggle('active',p.id==='pane-'+t));if(t==='fc'){setTimeout(()=>map.invalidateSize(),60)}else if(t==='hist'){setTimeout(()=>{if(lastData){drawTrend(lastData);drawHistory(lastData)}},60)}else if(t==='live'){setTimeout(()=>{liveMap.invalidateSize();const _g=wgsToGcj(+$('lat').value,+$('lon').value);liveMap.setView(_g,Math.max(liveMap.getZoom(),9));const im=document.querySelector('.sat-wrap img');if(im&&im.complete&&im.naturalWidth)zoomSat(im);const cv=$('satOverlay');if(cv&&cv.width>10)drawCloudOverlay();loadFy4Cloud()},60)}else if(t==='windy'){renderWindy()}else if(t==='wm'){setTimeout(()=>{wmInit()},60)}}
// v2.10.9: Windy 实况嵌入——按观测点 + 图层生成 iframe
let _windyLayer='wind';
const WINDY_ZOOM=7;
function windyUrl(){
  const lat=+$('lat').value||30.657,lon=+$('lon').value||104.058;
  const overlay=_windyLayer;
  // satellite 层在官方 embed 中自动使用实时卫星云图（东半球 Himawari），无需 product 参数
  const product=overlay==='satellite'?'':'product=ecmwf';
  return 'https://embed.windy.com/embed.html?type=map&location=coordinates&metricRain=default&metricTemp=default&metricWind=default&zoom='+WINDY_ZOOM+
    '&overlay='+overlay+(product?('&'+product):'')+'&level=surface&lat='+lat.toFixed(4)+'&lon='+lon.toFixed(4);
}
function renderWindy(){
  const box=$('windyBox');if(!box)return;
  box.innerHTML='<div class="windy-load">正在加载 Windy 实时地图…</div>';
  box.classList.add('loading');
  const f=document.createElement('iframe');
  f.src=windyUrl();f.loading='lazy';f.allowFullscreen=true;
  f.onload=()=>box.classList.remove('loading');
  box.appendChild(f);
}
function setWindyLayer(layer,btn){
  _windyLayer=layer;
  document.querySelectorAll('.windy-ctl button').forEach(b=>b.classList.toggle('on',b===btn));
  renderWindy();
}
// v2.10.16: 卫星云图底图（移除 windy.app 数值云量瓦片，仅保留风云四号红外/真彩色叠加）
let _wmMap=null;
function wmInit(){
  if(_wmMap){_wmMap.invalidateSize();return}
  const el=$('wmMap');if(!el)return;
  _wmMap=L.map('wmMap',{layers:[gaodeRoad]}).setView([+$('lat').value||30.657,+$('lon').value||104.058],6);
  _wmMap.createPane('wmPane');_wmMap.getPane('wmPane').style.zIndex=400;
  L.control.layers({'高德标准地图':gaodeRoad,'高德卫星影像':gaodeSatellite},null,{position:'topright',collapsed:true}).addTo(_wmMap);
  _wmMap.on('zoomend',()=>setWmStatus());
  setWmStatus('底图已加载');
}
function setWmStatus(extra){
  const s=$('wmStatus');if(!s)return;
  const mz=_wmMap?_wmMap.getZoom():6;
  s.textContent=`缩放级别 ${mz}${extra?(' · '+extra):''}`;
}
// v2.10.10: 标签页切换时若卫星云图已初始化过则保持状态
function wmRefresh(){if(_wmMap)_wmMap.invalidateSize()}
// v2.10.12: 风云四号实况云图叠加（红外 WMS 裁剪图 / 真彩色云图）
let _wmSat=null,_wmSatLayer=null,_wmAnimUrl=null;
function toggleFy4Sat(btn){
  if(!_wmMap){wmInit()}
  const mode=btn.dataset.wmSat;
  document.querySelectorAll('.wm-ctl [data-wm-sat]').forEach(b=>b.classList.toggle('on',b===btn&&_wmSat!==mode||(b===btn&&_wmSat===null)));
  if(_wmSat===mode){_wmSat=null;if(_wmSatLayer){_wmMap.removeLayer(_wmSatLayer);_wmSatLayer=null}if(_wmAnimUrl){URL.revokeObjectURL(_wmAnimUrl);_wmAnimUrl=null}setWmStatus('已取消实况叠加');return}
  _wmSat=mode;
  if(mode==='fy4'){loadFy4Irx(btn)}else{loadWxblSat(btn)}
}
function loadFy4Irx(btn){
  setWmStatus('正在加载风云四号红外云图…');
  fetch('/api/fy4-irx/info').then(r=>r.json()).then(j=>{
    if(!j.ok)throw Error(j.error||'加载失败');
    const d=j.data,b=d.bbox;
    const bounds=[[b[0],b[1]],[b[2],b[3]]];
    const apply=(url,label)=>{
      if(_wmSatLayer)_wmMap.removeLayer(_wmSatLayer);
      _wmSatLayer=L.imageOverlay(url,bounds,{opacity:.85,pane:'wmPane',interactive:false}).addTo(_wmMap);
      setWmStatus(label);
      if(_wmMap.getZoom()<6)_wmMap.setZoom(6,{animate:true});
    };
    // v2.10.15: 优先加载多时次动画 GIF，失败则回退单帧
    fetch('/api/fy4-irx/anim',{cache:'no-store'}).then(r=>{
      if(!r.ok)throw Error('动画不可用');
      return r.blob();
    }).then(blob=>{
      const url=URL.createObjectURL(blob);
      apply(url,'风云四号红外云图动画 · 近 8 时次 · 白色为云（15 分钟/帧，循环播放）');
      _wmAnimUrl=url;
    }).catch(()=>{
      apply('/api/fy4-irx','风云四号红外云图 · 时次 '+(new Date(d.time_bj).toLocaleString('zh-CN',{hour12:false}))+' · 白色为云');
    });
  }).catch(err=>{setWmStatus('风云四号红外加载失败: '+err.message);_wmSat=null;if(btn)btn.classList.remove('on')});
}
function loadWxblSat(btn){
  setWmStatus('正在加载风云四号真彩色云图…');
  fetch('/api/sat').then(r=>{if(!r.ok)throw Error('云图暂不可用（可能为夜间）');return r.json()}).then(j=>{
    if(!j.ok)throw Error(j.error||'加载失败');
    const d=j.data;
    // 中央气象台 ACHN 等距投影全国图（西起 65E 东至 145E，南 10N 北 60N）
    const bounds=[[10,65],[60,145]];
    const img=L.imageOverlay(d.url,bounds,{opacity:.85,pane:'wmPane',interactive:false});
    if(_wmSatLayer)_wmMap.removeLayer(_wmSatLayer);
    _wmSatLayer=img.addTo(_wmMap);
    setWmStatus('风云四号真彩色云图 · 时次 '+new Date(d.time).toLocaleString('zh-CN',{hour12:false}));
    if(_wmMap.getZoom()<6)_wmMap.setZoom(6,{animate:true});
  }).catch(err=>{setWmStatus('真彩色云图加载失败: '+err.message);_wmSat=null;if(btn)btn.classList.remove('on')});
}
// v2.5: 预报附带的历史（过去7天，随预报数据返回）
function drawHistory(d){
  const box=$('histBox');if(!box)return;
  const any=d.mountains.some(m=>m.history&&m.history.length);
  if(!any){box.innerHTML='<div class="hist-note">暂无历史数据（需要过去 7 天的天气再分析数据，可能因数据源限制暂缺）。</div>';return}
  let html='<div class="hist-note">过去 7 天 · <b style="color:var(--gold)">历史回算</b>（ERA5 再分析数据，非预报，可对照当天实际天气验证「雨后洗尘」等经验规律）。点击日期可展开当日明细；也可在下方选择任意日期范围查询。</div>';
  html+=renderHistDays(d.mountains.filter(m=>!m.terrain_blocked).map(m=>({name:m.name,daily:m.history.map(h=>({date:h.date,morning:h.morning,evening:h.evening}))})));
  box.innerHTML=html;
}
// v2.6: 任意日期范围查询
async function queryHistory(){
  const box=$('histBox'),s=$('hStart').value,e=$('hEnd').value;
  if(!s||!e){box.innerHTML='<div class="hist-note" style="color:var(--red)">请先选择起始和结束日期。</div>';return}
  box.innerHTML='<div class="hist-note">正在从 ERA5 再分析数据回算 '+s+' ~ '+e+' 的观山评分…</div>';
  try{
    const q=new URLSearchParams({lat:+$('lat').value,lon:+$('lon').value,start:s,end:e});
    const r=await fetch('/api/history?'+q);
    // v2.9.4: 网关/代理超时时可能返回纯文本而非 JSON，先校验类型再解析
    const ct=r.headers.get('content-type')||'';
    if(!ct.includes('application/json'))throw Error('历史数据源响应异常（上游超时或不可用），请稍后重试');
    const j=await r.json();
    if(!j.ok)throw Error(j.error);
    const d=j.data;
    let warn='';
    if(!d.aod_available)warn='<div class="hist-note" style="color:var(--gold)">⚠ 该时段超过气溶胶数据范围（近 92 天），AOD 评分项已自动降级，仍可参考云量/能见度结果。</div>';
    const head=`<div class="hist-note">${d.span.start} ~ ${d.span.end} · ${d.observer.name} · <b style="color:var(--gold)">历史回算</b>（ERA5 再分析，非预报）· 点击日期展开明细</div>`;
    const grid=renderHistDays(d.mountains.filter(m=>!m.terrain_blocked).map(m=>({name:m.name,daily:m.daily})));
    box.innerHTML=warn+head+grid;
  }catch(err){box.innerHTML='<div class="hist-note" style="color:var(--red)">查询失败：'+err.message+'</div>'}
}
// v2.6: 通用历史渲染（m: [{name, daily:[{date,morning,evening}]}]）
// v2.10.5: 历史卡片增加「晨/晚剖面」按钮，复用天空形态模拟弹窗渲染剖面图。
let _histCache=[];  // 供历史剖面按钮读取数据（避免 onclick 内嵌大 JSON）
function renderHistDays(items){
  _histCache=items;
  let html='';
  items.forEach((m,mi)=>{
    if(!m.daily||!m.daily.length)return;
    html+=`<div class="hist-head">${m.name} · ${m.daily.length} 天</div><div class="hist-grid">`;
    m.daily.slice().reverse().forEach((h,ri)=>{
      const idx=m.daily.length-1-ri;  // 原数组索引（与 _histCache 对应）
      const mo=h.morning,ev=h.evening;
      const rev=(mo.reasons&&mo.reasons.length)?mo.reasons.join('、'):'';
      html+=`<div class="hist-day" onclick="this.querySelector('.hx').style.display=this.querySelector('.hx').style.display==='block'?'none':'block'" style="cursor:pointer"><strong>${h.date.slice(5)}</strong><span class="hist-tag">回算</span><br>晨 <b style="color:${color(mo.score)}">${mo.score}</b> · 晚 <b style="color:${color(ev.score)}">${ev.score}</b><br><span style="color:var(--muted)">AOD ${mo.aod==null?'—':mo.aod.toFixed(2)} · 低云 ${mo.low}% · 能见 ${mo.visibility}km</span><div class="hx" style="display:none;margin-top:4px;color:var(--muted);line-height:1.5">${rev||'云量与通透度较好'}<br>经验 ${mo.empirical_bonus>=0?'+':''}${mo.empirical_bonus}${synTags(mo.synoptic)}${fogTag(mo.fog)}</div><div class="hist-sim"><button onclick="event.stopPropagation();openHistProfile(${mi},${idx},'morning')">晨间剖面</button><button onclick="event.stopPropagation();openHistProfile(${mi},${idx},'evening')">傍晚剖面</button></div></div>`;
    });
    html+='</div>';
  });
  return html||'<div class="hist-note">该时段无数据。</div>';
}
// v2.10.5: 历史回算剖面图（复用天空形态模拟弹窗与 drawPathProfile）
function openHistProfile(mi,di,period){
  const it=_histCache[mi];if(!it||!it.daily||!it.daily[di])return;
  const m=it.daily[di],x=m[period];
  if(!x||!x.profile||x.profile.length<2){alert('该时段剖面数据缺失');return}
  $('simTitle').textContent=`${it.name} · 路径云层剖面（历史回算）`;
  $('simSub').textContent=`${period==='morning'?'清晨':'傍晚'} ${new Date(x.time).toLocaleString('zh-CN')} · ERA5 再分析回算`;
  $('simMetrics').innerHTML=`<div><b>${x.low}%</b>路径低云峰值</div><div><b>${x.mid}%</b>峰区中云</div><div><b>${x.high}%</b>路径高云均值</div><div><b>${x.rh}%</b>${humidityName(x.rh)}</div><div><b>${x.visibility} km</b>最低能见度</div><div><b>${x.aod==null?'—':x.aod.toFixed(2)}</b>AOD550</div>`;
  $('simNote').innerHTML=`${fogName(x)}；${x.reasons.join('、')}。<br>该剖面为历史回算（ERA5 再分析），非预报；云层高度带为近似范围，用于判断视线遮挡。<br>云属按高度/云量/湿度/降水类型/层结综合推断：高云=卷云·卷积云·卷层云，中云=高积云·高层云，低云=积云·层积云·层云·碎积云；积雨云与雨层云为垂直跨层云塔。`;
  $('simModal').classList.add('open');
  requestAnimationFrame(()=>drawPathProfile(x,{name:it.name,elev:x.profile[x.profile.length-1].terrain}));
}
// v2.9.4: 真实观测三源（METAR 机场实况 + 风云四号卫星云图 + 全国雷达拼图）
function fmtT(iso){if(!iso)return'';const d=new Date(iso);return d.toLocaleString('zh-CN',{month:'numeric',day:'numeric',hour:'2-digit',minute:'2-digit'})}
function loadObs(){
  const box=$('obsBox');if(!box)return;
  fetch('/api/obs',{cache:'no-store'}).then(r=>r.json()).then(j=>{
    if(!j.ok||!j.data){box.style.display='none';return}
    const d=j.data,meta=d.metar,radar=d.radar,sc=d.sc_radar,sat=d.sat;
    if((!meta||!meta.length)&&!radar&&!sc&&!sat){box.style.display='none';return}
    box.style.display='';
    let html='<div class="obs-title">真实观测 <span class="tag2">METAR 机场实况 · 风云四号卫星云图 · 全国雷达拼图 · 四川单站雷达（免费 API，自动刷新）</span></div>';
    if(meta&&meta.length){
      html+='<div class="obs-grid">';
      meta.forEach(m=>{
        const fc=(m.flight||'').split(' ')[0];
        const vis=m.visibility_km!=null?m.visibility_km+'km':(m.wx==='CAVOK'?'≥10km':'—');
        const cb=m.cloud_base_m!=null?m.cloud_base_m+'m':'无';
        const cbTxt=m.cloud_base_m!=null?(m.cloud_base_m<600?'<b style="color:#c98a2b">'+cb+' 低云底</b>':cb):'无';
        const wx=m.wx?('<span style="color:#2d6a8f">'+m.wx+'</span>'):'晴';
        html+=`<div class="metar-card"><div class="ap">${m.name}<small>${m.icao}</small></div>`;
        html+=`<div class="metar-row"><span>能见度</span><b>${vis}</b></div><div class="metar-row"><span>云底</span>${cbTxt}</div>`;
        html+=`<div class="metar-row"><span>湿度/温度</span><b>${m.rh!=null?m.rh+'%':'-'} · ${m.temp!=null?m.temp+'°':'-'}</b></div>`;
        html+=`<div class="metar-row"><span>天气/风</span><b>${wx}${m.wind_kt!=null?' · '+m.wind_kt+'kt':''}</b></div>`;
        html+=`<div class="metar-row"><span>飞行规则</span><span class="flt ${fc}">${m.flight||'-'}</span></div></div>`;
      });
      html+='</div>';
    }
    const imgs=[];
    if(sat&&sat.url)imgs.push(`<figure><div class="sat-wrap"><img src="${sat.url}?_=${Date.now()}" alt="风云四号B星真彩色云图·西南地区" loading="lazy" onload="zoomSat(this);drawCloudOverlay()" onerror="this.parentNode.style.display='none'"><canvas id="satOverlay"></canvas></div><figcaption><b>风云四号B星真彩色云图 · 西南地区</b> ${fmtT(sat.time)}（中央气象台，按 97–110°E / 26–36°N 裁剪放大）<br>彩色线=视线走廊，按识别云型着色（青蓝=高云、绿=低云、橙=层云低云底、深蓝=雨层云、紫=积雨云），线上标注为最严重云型；括号数字为走廊起点总云量</figcaption><div class="sat-legend"><span><i style="background:#7fa3b8"></i>高云</span><span><i style="background:#5fa8ad"></i>中云</span><span><i style="background:#4c9a6c"></i>积云/低云</span><span><i style="background:#d1913f"></i>层云底</span><span><i style="background:#4a647a"></i>雨层云</span><span><i style="background:#7a3fa8"></i>积雨云</span><span><span class="pt"></span>观测点</span><span>▲ 雪山</span></div></figure>`);
    if(radar&&radar.url)imgs.push(`<figure><img src="${radar.url}?_=${Date.now()}" alt="全国雷达组合反射率拼图" loading="lazy" onerror="this.closest('figure').style.display='none'"><figcaption><b>全国雷达组合反射率拼图</b> ${fmtT(radar.time)}（中央气象台）<br>四川盆地回波指示实时降水，可对照「洗尘」评分项</figcaption></figure>`);
    if(imgs.length)html+='<div class="obs-imgs">'+imgs.join('')+'</div>';
    // v2.9.6: 云图趋势 + 当前云层判断（独立请求，不阻塞主面板）
    html+='<div id="obsTrendBox" class="obs-load">☁ 正在获取未来 3 小时云量趋势…</div>';
    // v2.9.5: 四川单站雷达（成都站为主，其余川内站次之）
    if(sc&&sc.stations&&sc.stations.length){
      const st=sc.stations;
      html+='<div class="obs-title" style="margin-top:12px">四川单站雷达 <span class="tag2">组合反射率 · 6 分钟更新 · 中央气象台</span></div>';
      html+=`<div class="obs-sc">${st.map(s=>`<figure class="${s.code==='AZ9280'?'sc-main':''}"><img src="${s.url}?_=${Date.now()}" alt="${s.name}雷达" loading="lazy" onerror="this.closest('figure').style.display='none'"><figcaption><b>${s.name}</b>${s.code==='AZ9280'?' · 主站':''} · ${fmtT(sc.time)}<br>${s.code==='AZ9280'?'成都本站覆盖成都平原及西部山区，直接对应观山方向':s.code==='AZ9816'?'绵阳站覆盖川西北（九顶山/四姑娘山方向）':s.code==='AZ9817'?'南充站覆盖川东北':s.code==='AZ9835'?'雅安站覆盖川西（贡嘎/二郎山方向）':'川内雷达站'}</figcaption></figure>`).join('')}</div>`;
    }
    if(d.sat_note)html+='<div class="obs-note">'+d.sat_note+'</div>';
    box.innerHTML=html;
    loadCloudTrend();
    loadFy4Cloud();
  }).catch(()=>{box.style.display='none'});
}
// v2.10.13: 风云四号实况云况分析（红外云检测统计 + 云顶高度分层）
function loadFy4Cloud(){
  const box=$('fy4CloudBox');if(!box)return;
  const lat=+$('lat').value,lon=+$('lon').value;
  fetch('/api/fy4-cloud?lat='+lat+'&lon='+lon,{cache:'no-store'}).then(r=>r.json()).then(j=>{
    if(!j.ok||!j.data)throw Error((j.error||'暂无数据'));
    box.innerHTML=renderFy4Cloud(j.data);
    loadFy4Gibs();
  }).catch(e=>{box.innerHTML='<div class="hist-note" style="color:var(--red)">云况分析加载失败：'+e.message+'</div>'}).finally(()=>{loadFireCloud()});
}
// v2.10.14: NASA GIBS 卫星云顶高度（Himawari 红外分级 + MODIS CTH 定量）
function loadFy4Gibs(){
  const box=$('fy4CloudBox');if(!box)return;
  const lat=+$('lat').value,lon=+$('lon').value;
  fetch('/api/fy4-cloud/gibs?lat='+lat+'&lon='+lon,{cache:'no-store'}).then(r=>r.json()).then(j=>{
    if(!j.ok||!j.data)throw Error((j.error||'暂无数据'));
    box.insertAdjacentHTML('beforeend',renderFy4Gibs(j.data));
  }).catch(e=>{box.insertAdjacentHTML('beforeend','<div class="hist-note" style="color:var(--muted);margin-top:10px">卫星云顶高度加载失败：'+e.message+'</div>')});
}
function renderFy4Gibs(d){
  const ir=d.ir13,ct=d.cth;
  let html='<div class="obs-title" style="margin-top:16px">卫星云顶高度 <span class="tag2">NASA GIBS · Himawari 红外 + MODIS 定量（免登录）</span></div>';
  if(ir){
    const bar=(v,color,label)=>`<div style="flex:1;min-width:64px"><div style="font-size:11px;color:var(--muted);margin-bottom:4px">${label}</div><div style="height:9px;border-radius:5px;background:#edf2f6;overflow:hidden"><div style="height:100%;width:${v}%;background:${color};border-radius:5px"></div></div><div style="font-size:12px;margin-top:3px"><b>${v}%</b></div></div>`;
    html+=`<div style="display:flex;gap:14px;flex-wrap:wrap;margin-top:8px">${bar(ir.high_rate,'#8a6bbd','高云 ≥7km')}${bar(ir.mid_rate,'#5fa8ad','中云 3-7km')}${bar(ir.low_rate,'#7fa3b8','低云/晴 <3km')}</div>`;
    html+=`<div class="obs-note" style="margin-top:8px">${ir.note}（平均亮温 ${ir.mean_tbb}℃）</div>`;
  }
  if(ct){
    html+=`<div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:10px">`;
    html+=`<div style="background:var(--soft);border-radius:8px;padding:8px 12px;font-size:12px"><div style="color:var(--muted);font-size:11px">云顶最高</div><b style="font-size:16px;color:var(--accent)">${ct.max} km</b></div>`;
    html+=`<div style="background:var(--soft);border-radius:8px;padding:8px 12px;font-size:12px"><div style="color:var(--muted);font-size:11px">平均</div><b style="font-size:16px">${ct.mean} km</b></div>`;
    html+=`<div style="background:var(--soft);border-radius:8px;padding:8px 12px;font-size:12px"><div style="color:var(--muted);font-size:11px">中位</div><b style="font-size:16px">${ct.median} km</b></div>`;
    html+=`<div style="background:var(--soft);border-radius:8px;padding:8px 12px;font-size:12px"><div style="color:var(--muted);font-size:11px">有效像元</div><b style="font-size:16px">${ct.total}</b></div></div>`;
    html+=`<div style="display:flex;height:14px;border-radius:7px;overflow:hidden;margin-top:8px;background:#edf2f6">`;
    html+=`<div style="width:${ct.low_rate}%;background:#7fa3b8" title="低云顶<3km"></div><div style="width:${ct.mid_rate}%;background:#5fa8ad" title="中云顶3-7km"></div><div style="width:${ct.high_rate}%;background:#8a6bbd" title="高云顶≥7km"></div></div>`;
    html+=`<div style="display:flex;gap:12px;font-size:11.5px;margin-top:4px;color:var(--muted)"><span>低&lt;3km ${ct.low_rate}%</span><span>中3-7km ${ct.mid_rate}%</span><span>高≥7km ${ct.high_rate}%</span><span>${ct.source} 过境</span></div>`;
    html+=`<div class="obs-note" style="margin-top:8px">${ct.note}</div>`;
  }
  if(!ir&&!ct)html+='<div class="hist-note" style="color:var(--muted)">暂无卫星云顶高度数据</div>';
  return html;
}
function renderFy4Cloud(d){
  const c=d.center,s=d.stats,t=d.top;
  const time=new Date(d.time_bj);
  const cBg=c.cloud?'linear-gradient(135deg,#5b7d94,#3d5c72)':'linear-gradient(135deg,#e8f4ee,#cfe8dc)';
  const cTxt=c.cloud?'#fff':'#2e6b4f';
  const cTag=c.cloud?'有云':'晴空';
  let topHtml='';
  if(t&&(t.low!=null||t.mid!=null||t.high!=null)){
    const bar=(v,color,label)=>{const p=v==null?0:v;return `<div style="flex:1;min-width:70px"><div style="font-size:11px;color:var(--muted);margin-bottom:4px">${label}</div><div style="height:9px;border-radius:5px;background:#edf2f6;overflow:hidden"><div style="height:100%;width:${p}%;background:${color};border-radius:5px"></div></div><div style="font-size:12px;margin-top:3px"><b>${v==null?'—':p+'%'}</b></div></div>`};
    topHtml=`<div class="obs-title" style="margin-top:12px">云顶高度分层 <span class="tag2">Open-Meteo 数值模式 · 低/中/高云量（近似）</span></div><div style="display:flex;gap:14px;flex-wrap:wrap;margin-top:6px">${bar(t.low,'#7fa3b8','低云 &lt;2km')}${bar(t.mid,'#5fa8ad','中云 2–6km')}${bar(t.high,'#4c9a6c','高云 &gt;6km')}</div>`;
  }
  const rows=[
    ['定位','<b>'+c.lat.toFixed(2)+'°N, '+c.lon.toFixed(2)+'°E</b>（卫星 4km 网格 '+c.row+','+c.col+'）'],
    ['中心像元','<span style="color:'+(c.cloud?'#c25548':'#3e8e6e')+';font-weight:700">'+cTag+'</span>（风云四号红外云检测）'],
    ['有云像元','<b style="color:#c25548">'+s.cloudy+' / '+s.total+'</b>（'+s.cloudy_rate+'%）'],
    ['晴空像元','<b style="color:#3e8e6e">'+s.clear+' / '+s.total+'</b>（'+s.clear_rate+'%）'],
    ['云区比例','<b>'+s.cloud_ratio+'%</b> · 晴空比例 <b>'+s.clear_ratio+'%</b>'],
  ];
  let html=`<div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-top:8px"><div style="min-width:96px;padding:14px 10px;border-radius:12px;background:${cBg};color:${cTxt};text-align:center"><div style="font-size:11px;opacity:.85">中心像元</div><div style="font-size:20px;font-weight:800;letter-spacing:1px">${cTag}</div><div style="font-size:10.5px;opacity:.8;margin-top:2px">${time.toLocaleString('zh-CN',{month:'numeric',day:'numeric',hour:'2-digit',minute:'2-digit'})} 时次</div></div><div style="flex:1;min-width:220px"><div style="font-size:11px;color:var(--muted);margin-bottom:4px">周围 ${s.total} 像元云况（约 90km 范围）</div><div style="display:flex;height:14px;border-radius:7px;overflow:hidden;background:#edf2f6"><div style="width:${s.cloudy_rate}%;background:#8fa8ba"></div><div style="width:${s.clear_rate}%;background:#7fbf9e"></div></div><div style="display:flex;gap:12px;font-size:11.5px;margin-top:4px"><span><i style="display:inline-block;width:9px;height:9px;border-radius:2px;background:#8fa8ba;vertical-align:middle;margin-right:3px"></i>云 ${s.cloudy_rate}%</span><span><i style="display:inline-block;width:9px;height:9px;border-radius:2px;background:#7fbf9e;vertical-align:middle;margin-right:3px"></i>晴 ${s.clear_rate}%</span><span style="color:var(--muted)">有效率 ${s.valid_rate}%</span></div></div></div>`;
  html+='<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:8px;margin-top:12px">'+rows.map(r=>`<div style="background:var(--soft);border-radius:8px;padding:8px 10px;font-size:12px"><div style="color:var(--muted);font-size:11px;margin-bottom:2px">${r[0]}</div>${r[1]}</div>`).join('')+'</div>';
  html+=topHtml;
  html+=`<div class="obs-note" style="margin-top:10px">${d.note}</div>`;
  return html;
}
// v2.10.19: 实时火烧云潜力预报（风云四号 / KMA 千里眼2A 双数据源 + 太阳扇区中高云识别 + 地图丝滑绘制）
let _fcData=null,_fcLayer=null,_fcSrc='fy4';
function switchFireSrc(src){
  if(src===_fcSrc)return;
  _fcSrc=src;
  const box=$('fireCloudBox');if(box)box.innerHTML='<div class="obs-load">☁ 正在切换数据源分析太阳方向中高云…</div>';
  loadFireCloud();
}
function loadFireCloud(){
  const box=$('fireCloudBox');if(!box)return;
  const lat=+$('lat').value,lon=+$('lon').value;
  fetch('/api/fire-cloud?lat='+lat+'&lon='+lon+'&src='+_fcSrc,{cache:'no-store'}).then(r=>r.json()).then(j=>{
    if(!j.ok||!j.data)throw Error(j.error||'暂无数据');
    _fcData=j.data;
    box.innerHTML=renderFireCloud(j.data);
    drawFireCloudMap(j.data,true);
  }).catch(e=>{box.innerHTML='<div class="hist-note" style="color:var(--red)">火烧云预报加载失败：'+e.message+'</div>'});
}
function renderFireCloud(d){
  const lv=d.level==='高'?'#e05a2b':d.level==='中'?'#c07f2a':d.level==='低'?'#8aa0b0':'#9aa8b5';
  const sun=d.sun,s=d.sector,t=new Date(d.time_bj);
  const pct=Math.max(0,Math.min(100,s.midhigh_rate));
  const litPct=Math.max(0,Math.min(100,s.lit_rate||0));
  let html=`<div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap;margin-top:8px"><span style="font-size:11px;color:var(--muted)">数据源</span><button onclick="switchFireSrc('fy4')" style="padding:5px 10px;font-size:11.5px;${d.src==='fy4'?'background:var(--accent);border-color:var(--accent);color:#fff':'background:#f4f7f9;color:#3b5163'}">风云四号</button><button onclick="switchFireSrc('kma')" style="padding:5px 10px;font-size:11.5px;${d.src==='kma'?'background:var(--accent);border-color:var(--accent);color:#fff':'background:#f4f7f9;color:#3b5163'}">KMA 千里眼2A</button></div>`;
  html+='<div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;margin-top:8px">';
  html+=`<div style="min-width:96px;padding:12px 10px;border-radius:12px;background:linear-gradient(135deg,#fdf0e4,#f6dcc2);text-align:center"><div style="font-size:11px;color:#a06a1f">潜力评分</div><div style="font-size:26px;font-weight:800;color:${lv}">${d.score}</div><div style="font-size:11px;font-weight:700;color:${lv}">${d.level}</div></div>`;
  html+=`<div style="flex:1;min-width:210px"><div style="font-size:11px;color:var(--muted);margin-bottom:4px">太阳方向扇区『云底受光』中高云占比（阳光从云边界下方穿过照亮云底才算）</div><div style="display:flex;height:12px;border-radius:6px;overflow:hidden;background:#edf2f6"><div style="width:${litPct}%;background:linear-gradient(90deg,#ffb27a,#e05a2b)"></div></div><div style="display:flex;gap:12px;font-size:11.5px;margin-top:4px"><span>云底受光 <b>${(s.lit_rate||0).toFixed(1)}%</b></span><span>中高云 ${s.midhigh_rate}%</span><span>总云量 ${s.cloud_rate}%</span><span style="color:var(--muted)">${t.toLocaleString('zh-CN',{month:'numeric',day:'numeric',hour:'2-digit',minute:'2-digit'})} 时次</span></div></div></div>`;
  const rows=[
    ['时段','<b>'+sun.phase+'</b>（太阳方位 '+sun.az+'° · 高度角 '+sun.elev+'°）'],
    ['扇区','太阳方位 ±'+s.half_width+'° 扇形内统计'],
    ['云底受光',(s.lit_rate||0)!==0?'<b style="color:#e05a2b">'+s.lit_rate+'%</b> 中高云可被低角度阳光照亮':'<b>0%</b>（云底处于阴影）'],
    ['云顶冷度',s.avg_gray!=null?'平均灰度 <b>'+s.avg_gray+'</b>（越高越冷）':'无受光中高云'],
    ['黄金窗口',d.window?'<b style="color:#e05a2b">处于日出/日落 ±10° 窗口</b>':'当前不在日出/日落窗口'],
  ];
  html+='<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:8px;margin-top:12px">'+rows.map(r=>`<div style="background:var(--soft);border-radius:8px;padding:8px 10px;font-size:12px"><div style="color:var(--muted);font-size:11px;margin-bottom:2px">${r[0]}</div>${r[1]}</div>`).join('')+'</div>';
  html+='<div class="obs-note" style="margin-top:10px">'+d.note+'</div>';
  html+=`<div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-top:10px"><button onclick="fitFireCloud()" style="padding:7px 12px;font-size:12px">定位到火烧云范围</button><span style="font-size:11px;color:var(--muted)">地图浅黄扇区=太阳观测窗口；橙红渐变=中高云分布范围</span></div>`;
  return html;
}
function destPtWgs(lat,lon,km,azDeg){
  const R=6371,br=azDeg*Math.PI/180,d=km/R,la=lat*Math.PI/180,lo=lon*Math.PI/180;
  const la2=Math.asin(Math.sin(la)*Math.cos(d)+Math.cos(la)*Math.sin(d)*Math.cos(br));
  const lo2=lo+Math.atan2(Math.sin(br)*Math.sin(d)*Math.cos(la),Math.cos(d)-Math.sin(la)*Math.sin(la2));
  return [la2*180/Math.PI,lo2*180/Math.PI];
}
function drawFireCloudMap(d,fast){
  const map=liveMap;if(!map)return;
  if(_fcLayer){map.removeLayer(_fcLayer);_fcLayer=null}
  const grp=L.layerGroup();
  const lat=+$('lat').value,lon=+$('lon').value;
  const az=d.sun.az,hw=d.sector.half_width;
  // 1) 太阳观测扇区（±70° 楔形，900km 半径，圆弧平滑拟合）
  const arc=[[lat,lon]];
  for(let k=0;k<=60;k++){const a=az-hw+(2*hw)*k/60;const p=destPtWgs(lat,lon,900,a);arc.push(wgsToGcj(p[0],p[1]))}
  arc.push([lat,lon]);
  L.polygon(arc,{color:'#d9a83f',weight:1,dashArray:'4 6',fillColor:'#ffd27a',fillOpacity:.13,interactive:false}).addTo(grp);
  // 太阳方位指示线 + 标注
  const sp=destPtWgs(lat,lon,620,az),sg=wgsToGcj(sp[0],sp[1]),og=wgsToGcj(lat,lon);
  L.polyline([og,sg],{color:'#e0a93f',weight:1.6,dashArray:'3 7',interactive:false}).addTo(grp);
  L.marker(sg,{icon:L.divIcon({className:'',html:'<div style="transform:translate(-50%,-100%);white-space:nowrap;font-size:11px;font-weight:700;color:#b07a1f;background:rgba(255,255,255,.88);border:1px solid #ecd9a8;border-radius:6px;padding:2px 7px">太阳方位 '+Math.round(az)+'°</div>',iconSize:[0,0]})}).addTo(grp);
  // 2) 中高云分布范围：Canvas 径向渐变叠加（半透明像元融合出"丝滑"云区）
  const cells=d.cells||[];
  const msz=map.getSize();
  if(cells.length&&msz.x>0&&msz.y>0){
    let minLat=90,maxLat=-90,minLon=181,maxLon=-181;
    cells.forEach(c=>{if(c.lat<minLat)minLat=c.lat;if(c.lat>maxLat)maxLat=c.lat;if(c.lon<minLon)minLon=c.lon;if(c.lon>maxLon)maxLon=c.lon});
    const nw=map.latLngToContainerPoint(wgsToGcj(maxLat,minLon));
    const se=map.latLngToContainerPoint(wgsToGcj(minLat,maxLon));
    let W=Math.abs(se.x-nw.x),H=Math.abs(se.y-nw.y);
    if(W>4&&H>4){
      const sc=Math.min(1,2600/Math.max(W,H));W=Math.round(W*sc);H=Math.round(H*sc);
      const cv=document.createElement('canvas');cv.width=W;cv.height=H;
      const ctx=cv.getContext('2d');
      const mpp=156543.03392*Math.cos(map.getCenter().lat*Math.PI/180)/Math.pow(2,map.getZoom());
      const rPx=Math.max(6,(map.getZoom()>=8?30:44)*1000/mpp*sc);
      const bnds=[[maxLat,minLon],[minLat,maxLon]].map(ll=>wgsToGcj(ll[0],ll[1]));
      cells.forEach(c=>{
        const p=map.latLngToContainerPoint(wgsToGcj(c.lat,c.lon));
        const x=(p.x-nw.x)*sc,y=(p.y-nw.y)*sc;
        if(x<-rPx||y<-rPx||x>W+rPx||y>H+rPx)return;
        const t=Math.max(0,Math.min(1,(c.g-160)/95));
        const rr=Math.round(255-55*t),gg=Math.round(70+60*t),bb=Math.round(58-22*t);
        const g=ctx.createRadialGradient(x,y,0,x,y,rPx);
        g.addColorStop(0,`rgba(${rr},${gg},${bb},.6)`);
        g.addColorStop(.55,`rgba(${rr},${gg},${bb},.24)`);
        g.addColorStop(1,`rgba(${rr},${gg},${bb},0)`);
        ctx.fillStyle=g;ctx.beginPath();ctx.arc(x,y,rPx,0,Math.PI*2);ctx.fill();
      });
      const ov=L.imageOverlay(cv.toDataURL(),bnds,{opacity:fast?0.92:0,interactive:false}).addTo(grp);
      if(!fast)setTimeout(()=>{ov.setOpacity(.92)},30);
    }
  }
  _fcLayer=grp.addTo(map);
}
function fitFireCloud(){
  if(!_fcData||!_fcData.cells||!_fcData.cells.length)return;
  let minLat=90,maxLat=-90,minLon=181,maxLon=-181;
  _fcData.cells.forEach(c=>{if(c.lat<minLat)minLat=c.lat;if(c.lat>maxLat)maxLat=c.lat;if(c.lon<minLon)minLon=c.lon;if(c.lon>maxLon)maxLon=c.lon});
  const b=[[maxLat,minLon],[minLat,maxLon]].map(ll=>wgsToGcj(ll[0],ll[1]));
  switchTab('live');setTimeout(()=>liveMap.fitBounds(b,{padding:[24,24],maxZoom:7}),80);
}
// v2.9.8: 西南区域裁剪放大（CSS 定位+缩放，避免跨域图片污染 canvas）
function zoomSat(img){
  const SRC={lon_min:65,lon_max:145,lat_min:10,lat_max:60};   // 风云四号 ACHN 源图（等距投影）
  const VIEW={lon_min:97,lon_max:110,lat_min:26,lat_max:36};  // 西南视图（与后端 SAT_VIEW_BBOX 一致）
  const wrap=img.closest('.sat-wrap');if(!wrap)return;
  const iw=img.naturalWidth||860,ih=img.naturalHeight||540;
  const sx=(VIEW.lon_min-SRC.lon_min)/(SRC.lon_max-SRC.lon_min)*iw;  // 西南区域左上角像素
  const sy=(SRC.lat_max-VIEW.lat_max)/(SRC.lat_max-SRC.lat_min)*ih;
  const sw=(VIEW.lon_max-VIEW.lon_min)/(SRC.lon_max-SRC.lon_min)*iw; // 西南区域像素宽高
  const sh=(VIEW.lat_max-VIEW.lat_min)/(SRC.lat_max-SRC.lat_min)*ih;
  const vw=wrap.clientWidth||wrap.offsetWidth||600;
  const scale=vw/sw;
  img.style.position='absolute';
  img.style.left=(-sx*scale)+'px';
  img.style.top=(-sy*scale)+'px';
  img.style.width=(iw*scale)+'px';
  img.style.height=(ih*scale)+'px';
  wrap.style.height=(sh*scale)+'px';
  const cv=wrap.querySelector('canvas');
  if(cv){cv.width=Math.round(vw);cv.height=Math.round(sh*scale)}
}
// v2.9.7: 在卫星云图上叠加绘制观测点、雪山与视线走廊（走廊按当前云量着色，直观显示影响范围）
// v2.9.9: 缓存按观测点位置区分，选点变化时自动重新拉取（云图信息跟随地图选点联动）
let _cmCache=null,_cmAt=0,_cmKey='';
function fetchCloudMap(cb){
  const now=Date.now(),lat=+$('lat').value,lon=+$('lon').value,key=lat.toFixed(4)+'|'+lon.toFixed(4);
  if(_cmCache&&now-_cmAt<300000&&key===_cmKey){cb(_cmCache);return}
  const q=new URLSearchParams({lat:lat,lon:lon});
  fetch('/api/cloud-map?'+q,{cache:'no-store'}).then(r=>r.json()).then(j=>{
    if(j.ok&&j.data){_cmCache=j.data;_cmAt=now;_cmKey=key;cb(j.data)}
  }).catch(()=>{});
}
function drawCloudOverlay(){
  const cv=$('satOverlay');if(!cv)return;
  fetchCloudMap(data=>{
    const W=cv.width,H=cv.height;
    if(W<10||H<10)return;
    const c=cv.getContext('2d');c.clearRect(0,0,W,H);
    const bb=data.bbox||{lon_min:65,lon_max:145,lat_min:10,lat_max:60};
    const X=lon=>(lon-bb.lon_min)/(bb.lon_max-bb.lon_min)*W;
    const Y=lat=>(bb.lat_max-lat)/(bb.lat_max-bb.lat_min)*H;
    // v2.10.7: 走廊云型识别着色——按 11 云属多因子判定分级（积雨云>雨层云>层云>层积云>中云>高云）
    const GSEV={ci:1,cc:2,cs:3,ac:3,as:4,cu:4,fn:4,sc:5,st:6,ns:7,cb:8};
    const GCOL={ci:'#9fb8c8',cc:'#8fb0c2',cs:'#7fa3b8',ac:'#5fa8ad',as:'#4f97a0',cu:'#4c9a6c',fn:'#7cb892',sc:'#3f8f74',st:'#d1913f',ns:'#4a647a',cb:'#7a3fa8'};
    const sev=g=>g?GSEV[g]||0:0;
    (data.corridors||[]).forEach(cor=>{
      const pts=cor.points||[];
      let worst=0,worstGd='',worstX=0,worstY=0;
      for(let i=0;i<pts.length-1;i++){
        const a=pts[i],b=pts[i+1];
        const gd=sev(a.genus)>=sev(b.genus)?a.genus:b.genus;
        const s=sev(gd);
        c.strokeStyle=GCOL[gd]||'#27ae60';c.lineWidth=2.4;c.globalAlpha=.9;
        c.beginPath();c.moveTo(X(a.lon),Y(a.lat));c.lineTo(X(b.lon),Y(b.lat));c.stroke();
        if(s>worst){worst=s;worstGd=gd;worstX=(X(a.lon)+X(b.lon))/2;worstY=(Y(a.lat)+Y(b.lat))/2}
      }
      c.globalAlpha=1;
      // 该走廊最严重云型标注（低云及以上才标，避免高云噪声）
      if(worst>=4&&worstGd){
        const zh=(pts.find(p=>p.genus===worstGd)||{}).genus_zh||worstGd;
        c.font='bold 10px system-ui';c.textAlign='center';
        c.lineWidth=3;c.strokeStyle='rgba(255,255,255,.92)';
        c.strokeText(zh,worstX,worstY-6);c.fillStyle='#2c3e50';c.fillText(zh,worstX,worstY-6);
      }
      // 雪山三角标记 + 名称
      c.fillStyle='#e67e22';c.strokeStyle='#ffffff';c.lineWidth=2;
      c.beginPath();c.moveTo(X(cor.lon),Y(cor.lat)-9);c.lineTo(X(cor.lon)-8,Y(cor.lat)+6);c.lineTo(X(cor.lon)+8,Y(cor.lat)+6);c.closePath();c.fill();c.stroke();
      c.font='bold 12px system-ui';c.textAlign='center';c.fillStyle='#6b3d12';
      c.fillText(cor.name.split('·')[0],X(cor.lon),Y(cor.lat)-13);
      // 走廊起点云量标注
      const p0=pts[0];
      if(p0&&p0.cloud!=null){c.font='10px system-ui';c.fillStyle='#33475a';c.fillText(p0.cloud+'%',X(p0.lon),Y(p0.lat)+3)}
    });
    // 观测点
    const ox=X(data.observer.lon),oy=Y(data.observer.lat);
    c.strokeStyle='#d64541';c.lineWidth=3;c.beginPath();c.arc(ox,oy,8,0,Math.PI*2);c.stroke();
    c.fillStyle='#d64541';c.beginPath();c.arc(ox,oy,4,0,Math.PI*2);c.fill();
    c.font='bold 13px system-ui';c.textAlign='left';c.fillStyle='#c0392b';
    c.fillText('观测点',ox+12,oy-9);
  });
}
// v2.9.9: 云图信息随地图选点联动刷新（观测点/雪山/走廊标注与走廊云量 + 云量趋势）
let _cmSyncKey='';
function syncCloudMapOverlay(){
  const key=(+$('lat').value).toFixed(4)+','+(+$('lon').value).toFixed(4);
  const moved=(key!==_cmSyncKey);
  if(moved)_cmSyncKey=key;
  const im=document.querySelector('.sat-wrap img');
  if(im){if(moved)_cmCache=null;drawCloudOverlay()}
  if(moved){const tb=$('obsTrendBox');if(tb)loadCloudTrend()}
}
window.addEventListener('resize',()=>{
  const im=document.querySelector('.sat-wrap img');
  if(im&&im.complete&&im.naturalWidth)zoomSat(im);
  const cv=$('satOverlay');if(cv&&cv.width>10)drawCloudOverlay()
});
function layerCls(l){if(l==='晴空')return'clear';if(l==='低云为主')return'low';if(l==='中云为主')return'mid';if(l==='高云为主')return'high';if(l.includes('雨层')||l.includes('系统性'))return'rain';if(l.includes('积雨')||l.includes('对流'))return'cb';if(l.includes('卷'))return'high';if(l.includes('高积')||l.includes('高层'))return'mid';if(l.includes('积云')||l.includes('层积')||l.includes('层云')||l.includes('碎积'))return'low';return'sct'}
function loadCloudTrend(){
  const box=$('obsTrendBox');if(!box)return;
  const q=new URLSearchParams({lat:+$('lat').value,lon:+$('lon').value});
  fetch('/api/cloud-trend?'+q,{cache:'no-store'}).then(r=>r.json()).then(j=>{
    if(!j.ok||!j.data){box.innerHTML='<div class="obs-note">云量趋势暂不可用（数据源繁忙）。</div>';return}
    const d=j.data,hours=d.hours;
    const hh=t=>{const x=new Date(t);return (x.getHours()+'').padStart(2,'0')+':00'};
    let cards=hours.map((h,i)=>`<div class="ct-h"><div class="t">${i===0?'现在':hh(h.time)}</div><div class="v">${h.cloud}%</div>${h.genus_zh?`<div class="g">${h.genus_zh}</div>`:''}<div class="ct-bar"><i class="low" style="width:${h.low}%"></i><i class="mid" style="width:${h.mid}%"></i><i class="high" style="width:${h.high}%"></i></div>${h.pop>=20?`<div class="pop">☔${h.pop}%</div>`:''}</div>`).join('');
    const tcl=d.trend.dir==='inc'?'#c25548':d.trend.dir==='dec'?'#2f7d5c':'#62798c';
    let html=`<div class="obs-trend"><div class="card"><h4>未来 3 小时云量趋势 <span class="tag2">观测点+走廊方向 · Open-Meteo</span></h4><div class="ct-hours">${cards}</div><div class="ct-trend" style="color:${tcl}">${d.trend.text}</div></div>`;
    html+=`<div class="card"><h4>当前云层判断</h4><div class="ct-judge"><span class="ct-layer ${layerCls(d.judge.layer)}">${d.judge.layer}</span><div class="ct-text">${d.judge.text}</div></div></div></div>`;
    box.innerHTML=html;
  }).catch(()=>{box.innerHTML='<div class="obs-note">云量趋势暂不可用（网络异常）。</div>'});
}
// v2.6: 历史日期默认值：默认近 7 天
 (function(){const d=new Date();const fmt=x=>x.toISOString().slice(0,10);const end=fmt(d);const st=new Date(d.getTime()-6*864e5);const el1=$('hStart'),el2=$('hEnd');if(el1){el1.value=fmt(st);el1.max=end;el2.value=end;el2.max=end}})();
 // v2.7: 导出为 Excel（历史范围 + 历史vs预报对比）
 function exportHistory(){
   const s=$('hStart').value,e=$('hEnd').value;
   if(!s||!e){alert('请先选择起始和结束日期');return}
   const q=new URLSearchParams({lat:+$('lat').value,lon:+$('lon').value,start:s,end:e});
   const a=document.createElement('a');a.href='/api/history/export?'+q;a.download='';document.body.appendChild(a);a.click();a.remove();
 }
 // v2.8: 天气系统标签（冷空气/锋面切变/槽脊/逆温层）
 function synTags(s){
   if(!s||!s.available)return'';
   const t=[];
   if(s.system)t.push(s.system);
   if(s.cold_air)t.push('冷空气·'+s.cold_air);
   if(s.front)t.push(s.front);
   if(s.inversion&&s.inversion!=='无')t.push(s.inversion+'逆温层');
   if(!t.length)return'';
   const pos=(s.synoptic_bonus>=0);
   return `<br><span class="sys-chip${pos?' pos':''}">天气系统 ${t.join(' · ')}</span>`;
 }
 // v2.8: 辐射雾风险标签
 function fogTag(f){
   if(!f||!f.level)return'';
   const icon=f.level==='高'?'⚠':'·';
   return `<span class="fog-chip">${icon}辐射雾风险${f.level}</span>`;
 }
 // v2.9: 实况观测面板
 const WMO_ICON={0:'☀',1:'🌤',2:'⛅',3:'☁',45:'🌫',48:'🌫',51:'🌦',53:'🌦',55:'🌧',61:'🌧',63:'🌧',65:'🌧',80:'🌦',81:'🌧',82:'🌧',95:'⛈',96:'⛈',99:'⛈'};
 function windDirName(d){if(d==null)return'';const a=['北','东北','东','东南','南','西南','西','西北'];return a[Math.round(((d%360)+360)%360/45)%8]+'风'}
 function loadCurrent(){
   const box=$('currentBox');if(!box)return;
   // 今日晨评分（各山最高，传给后端做一致性校验）
   let ts=null;
   if(lastData&&lastData.mountains&&lastData.mountains.length){
     ts=Math.max(...lastData.mountains.map(m=>m.daily&&m.daily.length?m.daily[0].morning.score:-1));
     if(ts<0)ts=null;
   }
   const q=new URLSearchParams({lat:+$('lat').value,lon:+$('lon').value,name:$('name').value||'成都'});
   if(ts!=null)q.set('today_score',ts);
   fetch('/api/current?'+q).then(r=>r.json()).then(j=>{
     if(!j.ok){box.innerHTML='<div class="cur-panel" style="color:var(--muted);font-size:13px">实况获取失败：'+j.error+'</div>';return}
     const d=j.data,om=d.om,cn=d.cn,chk=d.check;
     let html='';
     const icon=om?(WMO_ICON[om.weather_code]||'☁'):'';
     html+=`<div class="cur-panel"><div class="cur-main">${om?`<span class="cur-icon">${icon}</span><div><div class="cur-temp">${om.temperature_2m!=null?om.temperature_2m.toFixed(1)+'°':'—'}</div><div class="cur-cond">${d.wmo_zh||''}${om.is_day==1?' · 白天':om.is_day==0?' · 夜间':''}</div></div>`:'<div class="cur-cond">Open-Meteo 实况暂不可用</div>'}</div>`;
     html+=`<div class="cur-metrics">${om&&om.relative_humidity_2m!=null?`<span>湿度 <b>${om.relative_humidity_2m}%</b></span>`:cn&&cn.humidity!=null?`<span>站点湿度 <b>${cn.humidity}%</b></span>`:''}${om&&om.cloud_cover!=null?`<span>云量 <b>${om.cloud_cover}%</b></span>`:''}${om&&om.visibility!=null?`<span>能见度 <b>${(om.visibility/1000).toFixed(1)}km</b></span>`:cn&&cn.visibility!=null?`<span>站点能见度 <b>${(cn.visibility/1000).toFixed(1)}km</b></span>`:''}${om&&om.wind_speed_10m!=null?`<span>${windDirName(om.wind_direction_10m)} <b>${om.wind_speed_10m.toFixed(1)}m/s</b></span>`:''}${cn&&cn.temperature_2m!=null?`<span>站点 <b>${cn.temperature_2m.toFixed(0)}°</b></span>`:''}${cn&&cn.weather?`<span><b>${cn.weather}</b> ${cn.wind||''}${cn.wind_level||''}</span>`:''}${cn&&cn.aqi!=null?`<span>AQI <b>${cn.aqi}</b></span>`:''}</div></div>`;
     const cls=chk.match?'good':'warn';
     html+=`<div class="cur-check ${cls}">${chk.match?'✓':'⚠'} ${chk.note}</div>`;
     html+=`<div class="cur-src">来源：<b>中央气象台站点观测${cn?`（${cn.city}${cn.time?' · 更新 '+cn.time:''}）`:''}</b> + <b>Open-Meteo 当前时次${om&&om.time?' · '+om.time:''}</b>（免费 API，10 分钟刷新）</div>`;
     box.innerHTML=html;
   }).catch(()=>{box.innerHTML='<div class="cur-panel" style="color:var(--muted);font-size:13px">实况获取失败，稍后自动重试</div>'});
 }
 // v2.7: 评分趋势对比图——最近7天历史(虚线) vs 未来5天预报(实线)
 const TREND_COLORS=['#2d6a8f','#b07a2a','#4a9d78','#8a6bbd'];
 function drawTrend(d){
   const canvas=$('trendCanvas');if(!canvas)return;
   const rect=canvas.getBoundingClientRect(),dpr=Math.min(2,devicePixelRatio||1);
   if(rect.width<10||rect.height<10)return;  // v2.7: 容器隐藏(display:none)时尺寸为0，跳过等切Tab再画
   canvas.width=Math.round(rect.width*dpr);canvas.height=Math.round(rect.height*dpr);
   const c=canvas.getContext('2d');c.scale(dpr,dpr);
   const W=rect.width,H=rect.height,P=48,PL=54,PR=16,PT=16,PB=30;
   const cw=W-PL-PR,ch=H-PT-PB;
   c.fillStyle='#fff';c.fillRect(0,0,W,H);
   if(!d.mountains||!d.mountains.length){c.fillStyle='#8aa0b0';c.font='13px sans-serif';c.textAlign='center';c.fillText('暂无数据（请先加载预报）',W/2,H/2);return}
   const histLen=Math.max(0,...d.mountains.map(m=>m.history?m.history.length:0));
   const fcLen=Math.max(0,...d.mountains.map(m=>m.daily?m.daily.length:0));
   if(!histLen&&!fcLen){c.fillStyle='#8aa0b0';c.font='13px sans-serif';c.textAlign='center';c.fillText('暂无数据（请先加载预报）',W/2,H/2);return}
   const n=histLen+fcLen;
   const labels=[];const hist=d.mountains[0].history||[];const fc=d.mountains[0].daily||[];
   hist.forEach(h=>labels.push(h.date.slice(5)));fc.forEach(x=>labels.push(x.date.slice(5)));
   const X=i=>PL+i*cw/Math.max(1,n-1),Y=v=>PT+(100-v)/100*ch;
   // v2.10.1: 回算/预报分区底色（左=历史回算，右=预报）
   if(histLen&&fcLen){
     const bx=X(histLen-0.5);
     c.fillStyle='rgba(45,106,143,.06)';c.fillRect(PL,PT,bx-PL,ch);
     c.fillStyle='rgba(192,127,42,.07)';c.fillRect(bx,PT,PL+cw-bx,ch);
   }else if(histLen){c.fillStyle='rgba(45,106,143,.06)';c.fillRect(PL,PT,cw,ch)}
   else{c.fillStyle='rgba(192,127,42,.07)';c.fillRect(PL,PT,cw,ch)}
   // 网格与纵轴
   c.font='11px sans-serif';c.textAlign='right';c.fillStyle='#8aa0b0';
   for(let v=0;v<=100;v+=20){const y=Y(v);c.strokeStyle='#e7edf2';c.beginPath();c.moveTo(PL,y);c.lineTo(PL+cw,y);c.stroke();c.fillText(v,PL-6,y+4)}
   c.textAlign='center';
   labels.forEach((t,i)=>{const x=X(i);c.fillStyle='#8aa0b0';c.fillText(t,x,H-PB+16);if(i%2===0){c.strokeStyle='#f0f4f7';c.beginPath();c.moveTo(x,PT);c.lineTo(x,PT+ch);c.stroke()}});
   // v2.10.1: 顶部区域标签 + 分界线：历史 | 预报
   if(histLen&&fcLen){
     const bx=X(histLen-0.5);
     c.font='bold 11px sans-serif';c.textAlign='center';
     c.fillStyle='#2d6a8f';c.fillText('◀ 历史（回算）',PL+(bx-PL)/2,PT+13);
     c.fillStyle='#b07a2a';c.fillText('预报 ▶',bx+(PL+cw-bx)/2,PT+13);
     c.strokeStyle='#c25548';c.lineWidth=1.5;c.setLineDash([5,4]);c.beginPath();c.moveTo(bx,PT);c.lineTo(bx,PT+ch);c.stroke();c.setLineDash([]);
     c.fillStyle='#c25548';c.font='bold 11px sans-serif';c.textAlign='left';c.fillText('今天',bx+4,PT+13);
   }else if(histLen){c.font='bold 11px sans-serif';c.textAlign='center';c.fillStyle='#2d6a8f';c.fillText('历史（回算）',PL+cw/2,PT+13)}
   else if(fcLen){c.font='bold 11px sans-serif';c.textAlign='center';c.fillStyle='#b07a2a';c.fillText('预报',PL+cw/2,PT+13)}
   // 折线（v2.10.1: 回算=实线+实心点；预报=虚线+空心点）
   d.mountains.forEach((m,mi)=>{
     if(m.terrain_blocked)return;  // v2.10.21: 地形遮挡的山不参与趋势展示
     const col=TREND_COLORS[mi%TREND_COLORS.length];
     const histPts=[],fcPts=[];
     (m.history||[]).forEach((h,i)=>{histPts.push({x:X(i),y:Y(h.morning.score),v:h.morning.score,date:h.date})});
     (m.daily||[]).forEach((x,i)=>{fcPts.push({x:X(histLen+i),y:Y(x.morning.score),v:x.morning.score,date:x.date})});
     // 回算段：实线 + 实心点
     if(histPts.length){
       if(histPts.length>1){c.strokeStyle=col;c.lineWidth=2;c.setLineDash([]);c.beginPath();histPts.forEach((p,i)=>{if(i===0)c.moveTo(p.x,p.y);else c.lineTo(p.x,p.y)});c.stroke()}
       histPts.forEach(p=>{c.fillStyle=col;c.beginPath();c.arc(p.x,p.y,3.4,0,Math.PI*2);c.fill()});
     }
     // 预报段：虚线 + 空心点
     if(fcPts.length){
       if(fcPts.length>1){c.strokeStyle=col;c.lineWidth=2;c.setLineDash([6,4]);c.beginPath();fcPts.forEach((p,i)=>{if(i===0)c.moveTo(p.x,p.y);else c.lineTo(p.x,p.y)});c.stroke();c.setLineDash([])}
       fcPts.forEach(p=>{c.fillStyle='#fff';c.strokeStyle=col;c.lineWidth=1.8;c.beginPath();c.arc(p.x,p.y,3.6,0,Math.PI*2);c.fill();c.stroke()});
     }
     // 回算-预报连接线（今天与明天之间，细虚线过渡）
     if(histPts.length&&fcPts.length){const a=histPts[histPts.length-1],b=fcPts[0];c.strokeStyle=col;c.lineWidth=1.4;c.setLineDash([3,3]);c.beginPath();c.moveTo(a.x,a.y);c.lineTo(b.x,b.y);c.stroke();c.setLineDash([])}
     // 最后一点标注山名
     const lp=fcPts.length?fcPts[fcPts.length-1]:(histPts.length?histPts[histPts.length-1]:null);
     if(lp){c.fillStyle=col;c.font='bold 11px sans-serif';c.textAlign='left';c.fillText(m.name,Math.min(lp.x+6,W-PR-4),Math.max(PT+14,lp.y-9))}
   });
   // 图例卡片（左下角，带背景，不与折线混淆）
   c.font='11px sans-serif';c.textAlign='left';
   const lx=PL+8,ly=PT+ch-46,lw=172,lh=38;
   c.fillStyle='rgba(255,255,255,.94)';c.strokeStyle='#d3dce3';c.lineWidth=1;
   c.beginPath();if(c.roundRect)c.roundRect(lx,ly,lw,lh,6);else c.rect(lx,ly,lw,lh);c.fill();c.stroke();
   c.strokeStyle='#2d6a8f';c.lineWidth=2.2;c.setLineDash([]);c.beginPath();c.moveTo(lx+10,ly+12);c.lineTo(lx+28,ly+12);c.stroke();c.fillStyle='#2d6a8f';c.beginPath();c.arc(lx+19,ly+12,3,0,Math.PI*2);c.fill();
   c.fillStyle='#4a6270';c.fillText('历史（回算）',lx+36,ly+15);
   c.strokeStyle='#c07f2a';c.lineWidth=2.2;c.setLineDash([6,4]);c.beginPath();c.moveTo(lx+10,ly+30);c.lineTo(lx+28,ly+30);c.stroke();c.setLineDash([]);c.strokeStyle='#c07f2a';c.lineWidth=1.8;c.beginPath();c.arc(lx+19,ly+30,3.6,0,Math.PI*2);c.stroke();
   c.fillStyle='#4a6270';c.fillText('预报',lx+36,ly+33);
 }

function clamp01(x){return Math.max(0,Math.min(1,x))}
function humidityName(rh){return rh>=95?'近饱和':rh>=85?'高湿':rh>=70?'湿润':rh>=50?'适中':'干燥'}
function fogName(x){if(x.visibility<=1)return'浓雾风险';if(x.visibility<=5)return'雾/轻雾';if(x.visibility<=10||x.rh>=92)return'薄雾/霾感';return'无明显雾'}
function closeSimulation(){$('simModal').classList.remove('open')}
function openSimulation(mi,di,period){if(!lastData)return;const m=lastData.mountains[mi],x=m.daily[di][period];paintBlocked(mi,di,period);$('simTitle').textContent=`${m.name} · 路径云层剖面`;$('simSub').textContent=`${period==='morning'?'清晨':'傍晚'} ${new Date(x.time).toLocaleString('zh-CN')} · ${lastData.weather_model.name}`;$('simMetrics').innerHTML=`<div><b>${x.low}%</b>路径低云峰值</div><div><b>${x.mid}%</b>峰区中云</div><div><b>${x.high}%</b>路径高云均值</div><div><b>${x.rh}%</b>${humidityName(x.rh)}</div><div><b>${x.visibility} km</b>最低能见度</div><div><b>${x.aod==null?'—':x.aod.toFixed(2)}</b>AOD550</div>`;$('simNote').innerHTML=`${fogName(x)}；${x.reasons.join('、')}。<br>低云、中云、高云为模型分层云量；垂直高度带是近似范围，剖面用于判断视线遮挡，不代表实测云底云顶。<br>云属按高度/云量/湿度/降水类型/层结综合推断（v2.10.6）：高云=卷云·卷积云·卷层云，中云=高积云·高层云，低云=积云·层积云·层云·碎积云；积雨云与雨层云为垂直跨层云塔（积雨云=强对流降水、砧状云顶，雨层云=连续性降水、深厚云层）。`;$('simModal').classList.add('open');requestAnimationFrame(()=>drawPathProfile(x,m))}
function drawPathProfile(x,m){const canvas=$('simCanvas'),rect=canvas.getBoundingClientRect(),dpr=Math.min(2,devicePixelRatio||1);canvas.width=Math.round(rect.width*dpr);canvas.height=Math.round(rect.height*dpr);const c=canvas.getContext('2d');c.scale(dpr,dpr);const W=rect.width,H=rect.height,P=x.profile||[];if(P.length<2)return;const L=Math.max(52,W*.08),R=18,T=24,B=44,pw=W-L-R,ph=H-T-B,D=P[P.length-1].distance,maxTerrain=Math.max(...P.map(p=>p.terrain)),maxAlt=Math.max(16000,Math.ceil((maxTerrain+12000)/2000)*2000),Re=6371008.8*7/6,X=d=>L+d/D*pw,Y=z=>T+(maxAlt-z)/maxAlt*ph,bulge=d=>d*1000*(D-d)*1000/(2*Re),effective=p=>p.terrain+bulge(p.distance);c.fillStyle='#ffffff';c.fillRect(0,0,W,H);const sky=c.createLinearGradient(0,T,0,T+ph);sky.addColorStop(0,'#e5eef4');sky.addColorStop(.55,'#f0f6f9');sky.addColorStop(1,'#f8fbfc');c.fillStyle=sky;c.fillRect(L,T,pw,ph);
// 高度与距离网格
c.font=`${Math.max(10,H*.027)}px system-ui`;c.lineWidth=1;c.textAlign='right';for(let z=0;z<=maxAlt;z+=2000){const y=Y(z);c.strokeStyle='rgba(130,160,180,.16)';c.beginPath();c.moveTo(L,y);c.lineTo(W-R,y);c.stroke();c.fillStyle='#5a7280';c.fillText(z+'m',L-7,y+4)}c.textAlign='center';for(let i=0;i<=4;i++){const d=D*i/4,xx=X(d);c.strokeStyle='rgba(130,160,180,.12)';c.beginPath();c.moveTo(xx,T);c.lineTo(xx,H-B);c.stroke();c.fillStyle='#5a7280';c.fillText(Math.round(d)+'km',xx,H-B+18)}
// v2.10.6: 全面云属识别——国际十属 + 碎积云（11 属），按高度层/云量/湿度/降水类型/层结多因子判定
// ci卷云 cc卷积云 cs卷层云 | ac高积云 as高层云 | ns雨层云 sc层积云 st层云 cu积云 cb积雨云 fn碎积云
const GN_ZH={ci:'卷云',cc:'卷积云',cs:'卷层云',ac:'高积云',as:'高层云',ns:'雨层云',sc:'层积云',st:'层云',cu:'积云',cb:'积雨云',fn:'碎积云'};
const GENUS_COLOR={ci:{top:'188,214,226',bot:'146,182,200'},cc:{top:'176,205,220',bot:'128,168,190'},cs:{top:'160,196,212',bot:'116,158,180'},ac:{top:'150,196,200',bot:'96,152,168'},as:{top:'168,205,206',bot:'112,162,174'},ns:{top:'120,148,164',bot:'74,100,120'},sc:{top:'142,192,194',bot:'80,148,160'},st:{top:'160,200,200',bot:'96,150,158'},cu:{top:'130,186,190',bot:'64,138,150'},cb:{top:'104,132,152',bot:'48,74,96'},fn:{top:'150,196,198',bot:'88,150,158'},high:{top:'185,212,224',bot:'142,180,198'},mid:{top:'158,200,202',bot:'102,160,172'},low:{top:'138,190,192',bot:'76,145,158'}};
const genusOf=(key,cover,rh,precip,lowCov)=>{
  // lowCov=低云总云量，用于跨层云判定（雨层云/积雨云往往是低云主体）
  if(key==='high'){
    if((precip||0)>=1.5&&cover>=70)return'cs';        // 系统性降水伴随的卷层云
    if((precip||0)>0.8&&cover>=45)return'cc';         // 对流性降水伴随的卷积云
    if(cover>=80&&rh>=75)return'cs';                  // 大片均匀卷层云
    if(cover>=50&&rh<70)return'cc';                   // 成行/成波状的卷积云
    return'ci';
  }
  if(key==='mid'){
    if((precip||0)>=1.5&&cover>=70)return'ns';        // 连续性降水的中层（雨层云上界）
    if(cover>=75&&rh>=80)return'as';                  // 均匀幕状高层云
    if(cover>=55)return'ac';                          // 块状高积云
    if(cover>=25&&rh<65)return'ac';
    return cover>=15?'as':null;
  }
  // 低云：先判跨层强对流/系统性降水
  if((precip||0)>=2&&cover>=75)return'ns';            // 连续性降水 → 雨层云
  if((precip||0)>0.8)return'cb';                      // 对流性强降水 → 积雨云
  if(cover>=80&&lowCov>=80&&rh>=90)return'ns';        // 低云极厚+近饱和+无强对流 → 雨层云
  if(cover>=78&&rh>=90)return'st';                    // 厚而均匀、近饱和 → 层云
  if(cover>=80&&rh>=80&&(precip||0)>=0.3)return'cb';  // 厚积云伴有降水 → 积雨云
  if(cover>=50&&rh>=78)return'sc';                    // 层积云（较稳定）
  if(cover>=30&&rh<72)return'cu';                     // 低湿度块状 → 积云
  if(cover>=15&&rh>=85)return'fn';                    // 零散碎云+高湿 → 碎积云
  if(cover>=12)return'cu';
  return null;
};
let blocked=[];const genusSeen={};
for(let i=0;i<P.length;i++){const p=P[i],d0=i===0?0:(P[i-1].distance+p.distance)/2,d1=i===P.length-1?D:(p.distance+P[i+1].distance)/2,x0=X(d0),x1=X(d1),ground=effective(p),aod=clamp01((p.aod||0)/.6);if(aod>.03){c.fillStyle=`rgba(214,190,150,${.05+aod*.22})`;c.fillRect(x0,T,x1-x0,ph)}
  const lowCov=p.low||0;let lowGd=null;
  for(let bi=0;bi<3;bi++){const key=['high','mid','low'][bi],lo=[7000,3000,200][bi],hi=[12000,7000,3000][bi],cover=p[key]||0,base=ground+lo,top=Math.min(maxAlt,ground+hi);let gd=null;if(cover>3&&top>base){gd=genusOf(key,cover,p.rh,p.precip,lowCov);const yt=Y(top),yb=Y(base),hh=yb-yt,alpha=.14+cover/100*.5,col=GENUS_COLOR[gd]||GENUS_COLOR[key];
    const g=c.createLinearGradient(0,yt,0,yb);g.addColorStop(0,`rgba(${col.top},${alpha})`);g.addColorStop(1,`rgba(${col.bot},${alpha})`);c.fillStyle=g;c.fillRect(x0,yt,x1-x0,hh);
    c.strokeStyle=`rgba(255,255,255,${.28+cover/240})`;c.lineWidth=1;c.beginPath();c.moveTo(x0,yt);c.lineTo(x1,yt);c.stroke();
    c.strokeStyle=`rgba(60,110,120,${.12+cover/400})`;c.lineWidth=1;c.beginPath();c.moveTo(x0,yb);c.lineTo(x1,yb);c.stroke();
    if(gd&&!genusSeen[gd]){genusSeen[gd]=1;c.font=`${Math.max(9,H*.02)}px system-ui`;c.textAlign='center';c.fillStyle='#2f6a72';c.fillText(GN_ZH[gd],(x0+x1)/2,Math.max(T+12,yt-5))}
    const los=P[0].terrain+(P[P.length-1].terrain-P[0].terrain)*(p.distance/D);if(cover>=45&&los>=base&&los<=top)blocked.push({d:p.distance,z:los,key,cover})}
    if(bi===2)lowGd=gd}
  // v2.10.6: 跨层云塔绘制——积雨云（垂直发展至对流层顶，砧状云顶）/ 雨层云（连续性降水的深厚云层）
  if(lowGd==='cb'||lowGd==='ns'){
    const tTop=lowGd==='cb'?Math.min(maxAlt,ground+13000):Math.min(maxAlt,ground+6000),tBot=ground+150;
    if(tTop>tBot){const yT=Y(tTop),yB=Y(tBot),g3=c.createLinearGradient(0,yT,0,yB);
      if(lowGd==='cb'){g3.addColorStop(0,'rgba(122,142,162,.92)');g3.addColorStop(.35,'rgba(140,160,176,.72)');g3.addColorStop(1,'rgba(66,92,114,.9)')}
      else{g3.addColorStop(0,'rgba(128,152,168,.85)');g3.addColorStop(1,'rgba(78,104,124,.92)')}
      c.fillStyle=g3;c.fillRect(x0,yT,x1-x0,yB-yT);
      c.strokeStyle='rgba(255,255,255,.75)';c.lineWidth=1.6;
      c.beginPath();if(lowGd==='cb'){c.ellipse((x0+x1)/2,yT,Math.max(10,(x1-x0)*.9),5,0,0,Math.PI*2)}else{c.moveTo(x0,yT);c.lineTo(x1,yT)};c.stroke();
      c.strokeStyle='rgba(255,255,255,.18)';c.lineWidth=1;
      c.beginPath();c.moveTo(x0+(x1-x0)*.35,yT+6);c.lineTo(x0+(x1-x0)*.3,yB);c.stroke();
      c.beginPath();c.moveTo(x0+(x1-x0)*.65,yT+6);c.lineTo(x0+(x1-x0)*.7,yB);c.stroke();
    }
  }
  const fog=clamp01((p.rh-86)/14+(8-p.visibility)/10);if(fog>0){const fh=250+fog*900,yFb=Y(ground),yFt=Y(ground+fh);const g2=c.createLinearGradient(0,yFt,0,yFb);g2.addColorStop(0,'rgba(185,215,205,0)');g2.addColorStop(1,`rgba(185,215,205,${.25+fog*.4})`);c.fillStyle=g2;c.fillRect(x0,yFt,x1-x0,yFb-yFt)}}
// 近地面温度曲线（红色，沿线冷暖直观展示）
const temps=P.map(p=>p.temp).filter(v=>v!=null);if(temps.length){const tmin=Math.min(...temps),tmax=Math.max(...temps),tr=(tmax-tmin)||1;c.setLineDash([4,3]);c.strokeStyle='#d64541';c.lineWidth=1.8;c.beginPath();P.forEach((p,i)=>{const tv=(p.temp==null?15:p.temp),yt2=Y(effective(p))-140-(tv-tmin)/tr*260;if(i)c.lineTo(X(p.distance),yt2);else c.moveTo(X(p.distance),yt2)});c.stroke();c.setLineDash([]);c.font=`${Math.max(9,H*.019)}px system-ui`;c.fillStyle='#c0392b';const f0=P[0],fl=P[P.length-1];c.textAlign='left';c.fillText(((f0.temp==null?'':f0.temp)+'℃'),X(f0.distance)+3,Y(effective(f0))-150-((f0.temp==null?15:f0.temp)-tmin)/tr*260);c.textAlign='right';c.fillText(((fl.temp==null?'':fl.temp)+'℃'),X(fl.distance)-3,Y(effective(fl))-150-((fl.temp==null?15:fl.temp)-tmin)/tr*260)}
// 地形剖面（含7/6地球半径折射修正）
c.beginPath();c.moveTo(X(P[0].distance),Y(effective(P[0])));P.forEach(p=>c.lineTo(X(p.distance),Y(effective(p))));c.lineTo(X(D),Y(0));c.lineTo(X(0),Y(0));c.closePath();let tg=c.createLinearGradient(0,Y(maxTerrain+1000),0,Y(0));tg.addColorStop(0,'#9aa88a');tg.addColorStop(1,'#5d7153');c.fillStyle=tg;c.fill();c.beginPath();P.forEach((p,i)=>i?c.lineTo(X(p.distance),Y(effective(p))):c.moveTo(X(p.distance),Y(effective(p))));c.strokeStyle='#3f5a3e';c.lineWidth=1.6;c.stroke();
// 观测视线：观测点直达峰顶
const y0=Y(P[0].terrain),y1=Y(P[P.length-1].terrain);c.setLineDash([8,6]);c.strokeStyle='#e07b39';c.lineWidth=2.2;c.beginPath();c.moveTo(X(0),y0);c.lineTo(X(D),y1);c.stroke();c.setLineDash([]);c.fillStyle='#d64541';c.beginPath();c.arc(X(0),y0,5,0,Math.PI*2);c.fill();c.fillStyle='#e07b39';c.beginPath();c.moveTo(X(D),y1-8);c.lineTo(X(D)-6,y1+5);c.lineTo(X(D)+6,y1+5);c.closePath();c.fill();
// 云层与视线交点
blocked.forEach(b=>{c.fillStyle='#e74c3c';c.strokeStyle='#ffffff';c.lineWidth=1;c.beginPath();c.arc(X(b.d),Y(b.z),4.5,0,Math.PI*2);c.fill();c.stroke()});c.textAlign='left';c.fillStyle='#7a3b2e';c.fillText('观测点',X(0)+8,y0-8);c.textAlign='right';c.fillStyle='#b85c2a';c.fillText(`${m.name} ${m.elev}m`,X(D)-5,y1-10);c.textAlign='left';c.fillStyle='#7d8f9a';c.font=`${Math.max(9,H*.019)}px system-ui`;c.fillText('有效地形：已考虑地球曲率与标准折射',L,T+13);
// 图例（v2.10.6：云属已直接标注于图上；积雨云/雨层云为垂直跨层云塔）
const legend=[['rgba(138,190,192,.75)','低云 0.2–3km'],['rgba(158,200,202,.75)','中云 3–7km'],['rgba(185,212,224,.75)','高云 7–12km'],['rgba(104,132,152,.8)','积雨云/雨层云塔'],['rgba(185,215,205,.8)','雾/高湿层'],['#d64541','温度曲线'],['#e74c3c','视线遮挡']];let lx=L,ly=H-8;c.font=`${Math.max(9,H*.021)}px system-ui`;c.textAlign='left';legend.forEach(([co,tx])=>{c.fillStyle=co;c.fillRect(lx,ly-9,11,7);c.fillStyle='#4a6270';c.fillText(tx,lx+14,ly);lx+=c.measureText(tx).width+30;if(lx>W-120){lx=L;ly-=14}});if(blocked.length)$('simNote').innerHTML+=`<br><span style="color:#e74c3c">检测到 ${blocked.length} 个路径格点的云层高度带与观测视线相交，请重点关注红点位置。</span>`;else $('simNote').innerHTML+=`<br><span style="color:#2e8b6e">当前近似云层高度带未与峰顶观测视线直接相交。</span>`}
let forecastRequest=0;
async function loadAll(force){const chosen=$('model').value,key=chosen+'|'+$('lat').value+'|'+$('lon').value;
// v2.3: 请求互斥 + 同坐标去重（强制刷新除外）
if(_forecastBusy&&!force)return;
if(!force&&key===_lastReqKey){return}
_lastReqKey=key;_forecastBusy=true;
const id=++forecastRequest;localStorage.setItem('snowWeatherModel',chosen);document.body.classList.add('loading');$('modelStatus').textContent='天气模型：正在获取 '+$('model').selectedOptions[0].text+'…';$('aerosol').textContent='气溶胶：正在获取 Open-Meteo AOD550…';let q=new URLSearchParams(obs());try{
// v2.4: fetch 加 60s 超时，防止网关/代理挂起后返回 HTML 错误页
const ac=new AbortController(),to=setTimeout(()=>ac.abort(),60000);let r;try{r=await fetch('/api/forecast?'+q,{signal:ac.signal})}finally{clearTimeout(to)}
// v2.4: 校验响应类型——网关/代理超时或 404 时可能返回 HTML 而非 JSON
const ct=r.headers.get('content-type')||'';if(!ct.includes('application/json'))throw Error(r.status===504?'数据源响应超时(504)，请稍后刷新':'数据源返回异常页面('+r.status+')，可能为网关超时，请稍后刷新');
let j;try{j=await r.json()}catch(e){throw Error('数据源返回格式异常（非JSON），请稍后刷新')}
if(id!==forecastRequest)return;if(!j.ok)throw Error(j.error);_retry429=0;draw(j.data);loadCurrent();syncCloudMapOverlay();const ws=(j.data.warnings||[]);$('modelStatus').textContent=(ws.length?'⚠ '+ws.join('；')+' · ':'')+'天气模型：'+j.data.weather_model.name+' · '+j.data.weather_model.detail;$('modelStatus').style.borderColor=ws.length?'#c98a2b':'#2d6a8f';$('aerosol').textContent='气溶胶：'+j.data.aerosol.message;$('aerosol').style.borderColor='#3e8e6e'}catch(e){if(id!==forecastRequest)return;
// v2.3: 429 限流自动重试（最多2次，间隔递增）
if(String(e.message).includes('429')&&_retry429<2){_retry429++;const w=_retry429*2500;$('modelStatus').textContent=`数据源限流(429)，${w/1000} 秒后自动重试…`;$('modelStatus').style.borderColor='#c98a2b';setTimeout(()=>{_forecastBusy=false;loadAll(true)},w);return}
if(String(e.message).includes('超时')||String(e.message).includes('abort')){$('modelStatus').textContent='数据源响应超时，已恢复上一结果，可稍后重试';$('modelStatus').style.borderColor='#c98a2b';$('aerosol').textContent='请求超时（数据源繁忙）';$('aerosol').style.borderColor='#c98a2b'}else{$('modelStatus').textContent='数据源获取失败：'+e.message+'（已恢复上一结果）';$('modelStatus').style.borderColor='#c25548';$('aerosol').textContent='获取失败：'+e.message;$('aerosol').style.borderColor='#c25548'}}finally{if(id===forecastRequest)_forecastBusy=false;if(id===forecastRequest)document.body.classList.remove('loading')}}
let elevTimer,elevRequest=0,_lastGridAt=0;
async function loadElevationGrid(){const now=Date.now();if(now-_lastGridAt<10000)return;_lastGridAt=now;if(!map.hasLayer(elevationLayer))return;const id=++elevRequest,b=map.getBounds(),sw=gcjToWgs(b.getSouth(),b.getWest()),ne=gcjToWgs(b.getNorth(),b.getEast()),q=new URLSearchParams({south:sw[0],west:sw[1],north:ne[0],east:ne[1],n:10});try{const j=await fetch('/api/elevation-grid?'+q).then(r=>r.json());if(id!==elevRequest||!j.ok)return;elevationLayer.clearLayers();const colors={low:'#4a9d78',mid:'#c98a2b',high:'#8a6bbd',unknown:'#9aa8b3'};j.cells.forEach(c=>{const a=wgsToGcj(c.south,c.west),z=wgsToGcj(c.north,c.east);L.rectangle([a,z],{pane:'terrainPane',stroke:false,fillColor:colors[c.band],fillOpacity:.24,interactive:true}).bindTooltip(`${c.elevation==null?'未知':Math.round(c.elevation)+' m'} · ${c.band==='low'?'低海拔':c.band==='mid'?'中海拔':c.band==='high'?'高海拔':'未知'}`).addTo(elevationLayer)})}catch(e){console.warn('海拔分区获取失败',e)}}
map.on('moveend',()=>{clearTimeout(elevTimer);elevTimer=setTimeout(loadElevationGrid,500)});map.on('overlayadd',e=>{if(e.layer===elevationLayer)loadElevationGrid()});
function locate(){navigator.geolocation?navigator.geolocation.getCurrentPosition(p=>{$('lat').value=p.coords.latitude.toFixed(6);$('lon').value=p.coords.longitude.toFixed(6);$('elev').value='读取中';syncCloudMapOverlay();loadAll()},e=>alert(e.message),{enableHighAccuracy:true}):alert('浏览器不支持定位')}
const savedModel=localStorage.getItem('snowWeatherModel');if(savedModel&&[...$('model').options].some(x=>x.value===savedModel))$('model').value=savedModel;loadAll();loadCurrent();loadObs();
</script></body></html>'''


if __name__ == "__main__":
    print("成都看雪山：http://127.0.0.1:5000")
    APP.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False, threaded=True)
